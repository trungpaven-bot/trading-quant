
import pandas as pd
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.workbook.defined_name import DefinedName
import re

input_excel = r"g:\My Drive\30_RESOURCES (Kho tai nguyen)\30.01_Tinh hieu qua du an\02_Tra dinh muc theo TT 09 va 12\Du_lieu_Co_Tieu_De_Bang.xlsx"
output_excel = r"g:\My Drive\30_RESOURCES (Kho tai nguyen)\30.01_Tinh hieu qua du an\02_Tra dinh muc theo TT 09 va 12\File_Tra_Cuu_Final_4_Levels.xlsx"

print("Reading Data...")
df = pd.read_excel(input_excel, header=None)

processed_data = []
current_header_stack = [] 
current_table_name = ""
last_valid_code = ""

for idx, row in df.iterrows():
    row_strs = [str(x).strip() if not pd.isna(x) else "" for x in row]
    table_name = row_strs[0]
    
    # FILTER: Only process "Bảng X" where X is 1 to 101.
    match = re.search(r'Bảng (\d+)', table_name, re.IGNORECASE)
    if match:
        tbl_num = int(match.group(1))
        if not (1 <= tbl_num <= 101):
            continue # Skip Table 102+, or Table 0?
    else:
        # Skip tables without "Bảng X" number
        continue

    # STRICT RESET per Table
    if table_name != current_table_name:
        current_table_name = table_name
        current_header_stack = [] 
        last_valid_code = ""
    
    code = row_strs[1]
    desc = row_strs[2]
    # Check if data row
    vals = row_strs[3:]
    has_number = any(re.search(r'[\d.,]+', v) for v in vals)

    if has_number and desc:
        # --- DATA ROW (This is the leaf node, goes to C7 usually) ---
        val1, val2, val3 = 0, 0, 0
        nums = []
        for v in vals:
            try: nums.append(float(v.replace('.', '').replace(',', '.')))
            except: pass
        if len(nums) >= 1: val1 = nums[0]
        if len(nums) >= 2: val2 = nums[1]
        if len(nums) >= 3: val3 = nums[2]
        
        # Determine L1, L2, L3 based on Stack + current Desc
        # Stack contains the context. Desc contains the leaf.
        
        l1, l2, l3 = "", "", ""
        
        # Scenario A: Deep Nesting (Table 65)
        # Stack[0]: "Công trình đài..."
        # Stack[1]: "Cột anten... 30m"
        # Desc: "20W"
        # Target: L1=Stack[0], L2=Stack[1], L3=Desc.
        
        # Scenario B: Shallow (Table 1)
        # Stack[0]: "Nhà chung cư 5 tầng"
        # Desc: "Có 1 hầm"
        # Target: L1=Stack[0], L2=Desc, L3="-" (Empty)
        
        # Scenario C: Very Shallow
        # Stack empty.
        # Desc: "Nhà cấp 4"
        # Target: L1=Desc, L2="-", L3="-"
        
        stack_len = len(current_header_stack)
        
        if stack_len == 0:
            l1 = desc
            l2 = "-"
            l3 = "-"
        elif stack_len == 1:
            l1 = current_header_stack[0]
            l2 = desc
            l3 = "-"
        elif stack_len >= 2:
            l1 = current_header_stack[0]
            l2 = current_header_stack[1]
            l3 = desc
            
        if code: last_valid_code = code; final_code = code
        else: final_code = last_valid_code
        
        processed_data.append([table_name, final_code, l1, l2, l3, val1, val2, val3])
        
    else:
        # --- HEADER ROW (Context builder) ---
        if desc:
            if code:
                # Major Header (Likely L1)
                current_header_stack = [desc]
                last_valid_code = code
            else:
                # Minor Header
                if len(current_header_stack) == 0:
                    current_header_stack.append(desc)
                else:
                    # Logic for subsequent headers:
                    # If we are drilling down: Append.
                    # If we are moving sideways: Replace last.
                    # How to distinguish?
                    # Generally, drilling down happens once. "Category -> SubCategory".
                    # A third level usually doesn't exist as a Header, but as Data.
                    # Exception: Very complex tables.
                    
                    # Safer approach for Table 65:
                    # "Công trình..." (L1) -> "Cột anten..." (L2).
                    # "Cột anten 45m" -> Should replace "Cột anten 30m" (L2).
                    
                    if len(current_header_stack) >= 2:
                        current_header_stack[-1] = desc # Replace the deepest level
                    else:
                        current_header_stack.append(desc) # Drill down to Level 2

print(f"Parsed {len(processed_data)} rows.")

# --- GENERATE MENUS ---
processed_data.sort(key=lambda x: (x[0], x[2], x[3], x[4])) # Sort Tbl, L1, L2, L3

# Menu 1: Table
tables = []
seen = set()
for r in processed_data:
    if r[0] not in seen: seen.add(r[0]); tables.append(r[0])

# Menu 2: L1 (Key: Tbl)
menu_l1 = []
seen = set()
for r in processed_data:
    k = r[0]
    val = r[2]
    unique = f"{k}_{val}"
    if unique not in seen:
        seen.add(unique); menu_l1.append([k, val])

# Menu 3: L2 (Key: Tbl_L1)
menu_l2 = []
seen = set()
for r in processed_data:
    k = f"{r[0]}_{r[2]}"
    val = r[3]
    unique = f"{k}_{val}"
    if unique not in seen:
        seen.add(unique); menu_l2.append([k, val])

# Menu 4: L3 (Key: Tbl_L1_L2) -> Actually we can pull L3 from Engine directly if sorted.
# But for Dropdown C7 source, we need a list.
menu_l3 = []
seen = set()
for r in processed_data:
    k = f"{r[0]}_{r[2]}_{r[3]}"
    val = r[4]
    unique = f"{k}_{val}"
    if unique not in seen:
        seen.add(unique); menu_l3.append([k, val])


# --- EXCEL ---
wb = openpyxl.Workbook()

# SHEET: MENU
ws_menu = wb.active; ws_menu.title = "MENU"
ws_menu.append(["Table", "M1_Key", "M1_Val", "M2_Key", "M2_Val", "M3_Key", "M3_Val"])

# Col A: Tables
for i, t in enumerate(tables): ws_menu[f'A{i+2}'] = t
# Col B,C: L1
for i, m in enumerate(menu_l1): ws_menu[f'B{i+2}'] = m[0]; ws_menu[f'C{i+2}'] = m[1]
# Col D,E: L2
for i, m in enumerate(menu_l2): ws_menu[f'D{i+2}'] = m[0]; ws_menu[f'E{i+2}'] = m[1]
# Col F,G: L3
for i, m in enumerate(menu_l3): ws_menu[f'F{i+2}'] = m[0]; ws_menu[f'G{i+2}'] = m[1]

# DEFINED NAMES
def add_dn(name, sheet, col, count):
    wb.defined_names.add(DefinedName(name, attr_text=f"{sheet}!${col}$2:${col}${count+1}"))

add_dn("ListTables", "MENU", "A", len(tables))
add_dn("M1_Key", "MENU", "B", len(menu_l1))
add_dn("M1_Val", "MENU", "C", len(menu_l1))
add_dn("M2_Key", "MENU", "D", len(menu_l2))
add_dn("M2_Val", "MENU", "E", len(menu_l2))
add_dn("M3_Key", "MENU", "F", len(menu_l3))
add_dn("M3_Val", "MENU", "G", len(menu_l3))

# SHEET: ENGINE
ws_eng = wb.create_sheet("ENGINE")
headers = ["Table", "Code", "L1", "L2", "L3", "V1", "V2", "V3", "SuperKey"]
ws_eng.append(headers)
for i, r in enumerate(processed_data):
    # r = [Tbl, Code, L1, L2, L3, V1, V2, V3]
    super_key = f"{r[0]}_{r[2]}_{r[3]}_{r[4]}"
    ws_eng.append([r[0], r[1], r[2], r[3], r[4], r[5], r[6], r[7], super_key])
    
add_dn("Eng_Super", "ENGINE", "I", len(processed_data))


# SHEET: TRA_CUU
ws_ui = wb.create_sheet("TRA_CUU")
ws_ui['B2']="TRA CỨU 4 CẤP (FINAL)"; ws_ui['B2'].font=Font(size=14, bold=True)

# C4: TABLE
ws_ui['B4']="1. Bảng:"; ws_ui['C4'].border=Border(bottom=Side(style='thin'))
dv1 = DataValidation(type="list", formula1="=ListTables", allow_blank=True); ws_ui.add_data_validation(dv1); dv1.add(ws_ui['C4'])

# C5: L1 (Group 1)
ws_ui['B5']="2. Nhóm 1:"; ws_ui['C5'].border=Border(bottom=Side(style='thin')); ws_ui['C5'].alignment=Alignment(wrap_text=True)
dv2_f = "=OFFSET(M1_Val;MATCH($C$4;M1_Key;0)-1;0;COUNTIF(M1_Key;$C$4);1)"
dv2 = DataValidation(type="list", formula1=dv2_f, allow_blank=True); ws_ui.add_data_validation(dv2); dv2.add(ws_ui['C5'])

# C6: L2 (Group 2)
ws_ui['B6']="3. Nhóm 2:"; ws_ui['C6'].border=Border(bottom=Side(style='thin')); ws_ui['C6'].alignment=Alignment(wrap_text=True)
# Key: Tbl_L1
dv3_f = "=OFFSET(M2_Val;MATCH($C$4&\"_\"&$C$5;M2_Key;0)-1;0;COUNTIF(M2_Key;$C$4&\"_\"&$C$5);1)"
dv3 = DataValidation(type="list", formula1=dv3_f, allow_blank=True); ws_ui.add_data_validation(dv3); dv3.add(ws_ui['C6'])

# C7: L3 (Detail)
ws_ui['B7']="4. Chi tiết:"; ws_ui['C7'].border=Border(bottom=Side(style='thin')); ws_ui['C7'].alignment=Alignment(wrap_text=True)
# Key: Tbl_L1_L2
dv4_f = "=OFFSET(M3_Val;MATCH($C$4&\"_\"&$C$5&\"_\"&$C$6;M3_Key;0)-1;0;COUNTIF(M3_Key;$C$4&\"_\"&$C$5&\"_\"&$C$6);1)"
dv4 = DataValidation(type="list", formula1=dv4_f, allow_blank=True); ws_ui.add_data_validation(dv4); dv4.add(ws_ui['C7'])

# RESULTS
ws_ui['B9']="KẾT QUẢ:"; labels=["Mã","Suất","XD","TB"]; cols=["B","F","G","H"] # Engine Cols
for i,l in enumerate(labels):
    r=10+i; ws_ui[f'B{r}']=l; ws_ui[f'B{r}'].font=Font(bold=True)
    # Match SuperKey: C4_C5_C6_C7
    f=f"=INDEX(ENGINE!{cols[i]}:{cols[i]}; MATCH($C$4 & \"_\" & $C$5 & \"_\" & $C$6 & \"_\" & $C$7; Eng_Super; 0))"
    if i>=1: ws_ui[f'C{r}'].number_format='#,##0'
    ws_ui[f'C{r}']=f

ws_menu.sheet_state='hidden'
ws_eng.sheet_state='hidden'

print(f"Saving {output_excel}...")
wb.save(output_excel)
print("Done.")
