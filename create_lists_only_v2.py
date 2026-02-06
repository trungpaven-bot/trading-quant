
import pandas as pd
import openpyxl
from openpyxl.styles import Font
import re

input_excel = r"g:\My Drive\30_RESOURCES (Kho tai nguyen)\30.01_Tinh hieu qua du an\02_Tra dinh muc theo TT 09 va 12\Du_lieu_Co_Tieu_De_Bang.xlsx"
output_excel = r"g:\My Drive\30_RESOURCES (Kho tai nguyen)\30.01_Tinh hieu qua du an\02_Tra dinh muc theo TT 09 va 12\File_Tra_Cuu_Lists_Only.xlsx"

print("Reading Data...")
df = pd.read_excel(input_excel, header=None)

processed_data = []

# Logic State
current_header_stack = [] 
current_table_name = ""
last_valid_code = ""

# FLAG: Track if we have seen data since the last header
# If True: The next header is a SIBLING (Replace last item).
# If False: The next header is a CHILD (Append).
data_seen_since_last_header = False

start_processing = False

for idx, row in df.iterrows():
    row_strs = [str(x).strip() if not pd.isna(x) else "" for x in row]
    table_name = row_strs[0]
    
    # Filter 1-101
    match = re.search(r'Bảng (\d+)', table_name, re.IGNORECASE)
    if match:
        tbl_num = int(match.group(1))
        if not (1 <= tbl_num <= 101): continue
        if not start_processing: start_processing = True
    else:
        continue
    
    # Reset Stack on Table Change
    if table_name != current_table_name:
        current_table_name = table_name
        current_header_stack = [] 
        last_valid_code = ""
        data_seen_since_last_header = False
    
    code = row_strs[1]
    desc = row_strs[2]
    vals = row_strs[3:]
    has_number = any(re.search(r'[\d.,]+', v) for v in vals)

    if has_number and desc:
        # DATA ROW
        data_seen_since_last_header = True # Mark that we have processed data for current header
        
        # Combine Headers
        prefix = " ".join(current_header_stack)
        
        if prefix:
            final_desc = f"{prefix} {desc}"
        else:
            final_desc = desc
            
        if code: last_valid_code = code; final_code = code
        else: final_code = last_valid_code
        
        processed_data.append([tbl_num, table_name, final_code, final_desc])
        
    else:
        # HEADER ROW
        if desc:
            if code:
                # Major header implies new block -> Reset
                current_header_stack = [desc]
                last_valid_code = code
                data_seen_since_last_header = False
            else:
                # Minor header
                if len(current_header_stack) == 0:
                    current_header_stack.append(desc)
                else:
                    # DECISION: Append vs Replace
                    if data_seen_since_last_header:
                        # We have just finished a data block. This new header is a SIBLING.
                        # Replace the last level.
                        current_header_stack[-1] = desc
                    else:
                        # We are still drilling down (Header after Header). This is a CHILD.
                        # Append.
                        current_header_stack.append(desc)
                
                # Reset data flag because we just set a new header context
                data_seen_since_last_header = False

# Sort Natural
processed_data.sort(key=lambda x: x[0]) 

# --- BUILD EXCEL ---
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "LISTS"

# Header
ws['A1'] = "Bảng"
ws['C1'] = "Mã hiệu"
ws['D1'] = "Loại công trình"

# Write Data
for i, r in enumerate(processed_data):
    row_idx = i + 2
    ws[f'A{row_idx}'] = r[1] 
    ws[f'C{row_idx}'] = r[2] 
    ws[f'D{row_idx}'] = r[3] 

# Formatting
for cell in ws[1]: cell.font = Font(bold=True)
ws.column_dimensions['A'].width = 40
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 80

print(f"Saving {output_excel}...")
wb.save(output_excel)
print("Done.")
