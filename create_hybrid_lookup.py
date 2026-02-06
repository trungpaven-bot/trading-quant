
import pandas as pd
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.workbook.defined_name import DefinedName
import re

# Logic: 
# 1. Use the "Du_lieu_Co_Tieu_De_Bang.xlsx" which has Table Names.
# 2. But user wants the DATA sheet to look like the Screenshot (with headers, gaps).
#    So we should use the raw extraction "Du_lieu_Suat_von_dau_tu_2024.xlsx"
#    AND inject the Table Titles back in visually so it looks nice?
#    Or just use the raw excel and build lookup ON TOP of it.

# Let's start with the "Du_lieu_Co_Tieu_De_Bang.xlsx" data but format it nicely with gaps.
# actually user liked the screenshot which HAS gaps and headers. 
# That screenshot came from "Du_lieu_Suat_von_dau_tu_2024.xlsx".
# But that file didn't have Table Titles associated with rows for lookup.

# HYBRID APPROACH:
# Use "Du_lieu_Co_Tieu_De_Bang.xlsx" (Table Name | Code | Desc | Val...) as the HIDDEN ENGINE ("ENGINE_DATA").
# Create a VISUAL DATA sheet that looks like the screenshot for user to read.
# TRA_CUU sheet uses the ENGINE_DATA.

source_excel = r"g:\My Drive\30_RESOURCES (Kho tai nguyen)\30.01_Tinh hieu qua du an\02_Tra dinh muc theo TT 09 va 12\Du_lieu_Co_Tieu_De_Bang.xlsx"
display_excel = r"g:\My Drive\30_RESOURCES (Kho tai nguyen)\30.01_Tinh hieu qua du an\02_Tra dinh muc theo TT 09 va 12\Du_lieu_Suat_von_dau_tu_2024.xlsx" # The pretty one
output_excel = r"g:\My Drive\30_RESOURCES (Kho tai nguyen)\30.01_Tinh hieu qua du an\02_Tra dinh muc theo TT 09 va 12\File_Tra_Cuu_Dep_Va_Thong_Minh.xlsx"

print("Loading Data...")
df_engine = pd.read_excel(source_excel, header=None) # [Table, Code, Desc, Unit, Val1, Val2, Val3]
df_visual = pd.read_excel(display_excel, header=None) # Raw dump with gaps

wb = openpyxl.Workbook()

# --- 1. SHEET: DATA (Visual) ---
ws_visual = wb.active
ws_visual.title = "DATA_GOC" # Rename to imply it's source visual
# Copy raw visual data
for idx, row in df_visual.iterrows():
    ws_visual.append(list(row))

# Formatting Visual Sheet roughly
ws_visual.column_dimensions['A'].width = 15 # Code
ws_visual.column_dimensions['B'].width = 60 # Desc
for row in ws_visual.iter_rows(min_row=1, max_row=100): # Just style top lightly
    for cell in row:
        if cell.value and "BẢNG" in str(cell.value):
            cell.font = Font(bold=True)

# --- 2. SHEET: ENGINE (Hidden) ---
ws_engine = wb.create_sheet("ENGINE")
ws_engine.sheet_state = 'hidden'
ws_engine.append(["Table", "Code", "Desc", "Unit", "Total", "Const", "Equip", "LookupKey"])

# Process Engine Data to build LookupKey
# Row format from previous extraction: [Table, Code, Desc, Unit, Val1, Val2, Val3]
cleaned_engine = []
for idx, row in df_engine.iterrows():
    # Skip header row if exists in df selection (it had 0, 1, 2...)
    # Ensure values
    r = [str(x) if not pd.isna(x) else "" for x in row]
    # Create Lookup Key: Table + Desc (Code might be empty in children)
    # Key = TableName + "_" + Description
    key = f"{r[0]}_{r[2]}"
    cleaned_engine.append(r + [key])
    ws_engine.append(r + [key])

# --- 3. SHEET: LISTS (Helpers) ---
ws_lists = wb.create_sheet("LISTS")
ws_lists.sheet_state = 'hidden'

# Unique Tables
unique_tables = sorted(list(set(r[0] for r in cleaned_engine)))
ws_lists['A1'] = "UniqueTables"
for i, t in enumerate(unique_tables):
    ws_lists[f'A{i+2}'] = t

wb.defined_names.add(DefinedName("List_Tables", attr_text=f"LISTS!$A$2:$A${len(unique_tables)+1}"))

# Helper Cols for Dependent Dropdown (Table -> Desc)
# Sort engine data by Table
sorted_engine = sorted(cleaned_engine, key=lambda x: x[0])
ws_lists['C1'] = "TableCol"
ws_lists['D1'] = "DescCol"
for i, row in enumerate(sorted_engine):
    ws_lists[f'C{i+2}'] = row[0] # Table
    ws_lists[f'D{i+2}'] = row[2] # Desc

max_r = len(sorted_engine) + 1
wb.defined_names.add(DefinedName("Col_Table", attr_text=f"LISTS!$C$2:$C${max_r}"))
wb.defined_names.add(DefinedName("Col_Desc", attr_text=f"LISTS!$D$2:$D${max_r}"))

# --- 4. SHEET: TRA_CUU ---
ws_ui = wb.create_sheet("TRA_CUU")
ws_ui.column_dimensions['B'].width = 25
ws_ui.column_dimensions['C'].width = 60

ws_ui['B2'] = "TRA CỨU SUẤT VỐN ĐẦU TƯ 2024"
ws_ui['B2'].font = Font(size=14, bold=True, color="008000")

# L1: TABLE
ws_ui['B4'] = "1. Chọn Bảng:"
ws_ui['B4'].font = Font(bold=True)
c4 = ws_ui['C4']
c4.border = Border(bottom=Side(style='thin'))
dv1 = DataValidation(type="list", formula1="=List_Tables", allow_blank=True)
ws_ui.add_data_validation(dv1)
dv1.add(c4)

# L2: PROJECT
ws_ui['B5'] = "2. Chọn Công trình:"
ws_ui['B5'].font = Font(bold=True)
c5 = ws_ui['C5']
c5.border = Border(bottom=Side(style='thin'))
c5.alignment = Alignment(wrap_text=True)

# Dependent Formula
dv2_formula = "=OFFSET(LISTS!$D$2;MATCH($C$4;Col_Table;0)-1;0;COUNTIF(Col_Table;$C$4);1)"
dv2 = DataValidation(type="list", formula1=dv2_formula, allow_blank=True)
dv2.errorTitle = ""
ws_ui.add_data_validation(dv2)
dv2.add(c5)

# RESULTS
ws_ui['B8'] = "KẾT QUẢ TRA CỨU:"
ws_ui['B8'].font = Font(bold=True, underline="single")

labels = ["Mã hiệu", "Đơn vị tính", "Suất vốn đầu tư", "Chi phí Xây dựng", "Chi phí Thiết bị"]
# Map to ENGINE columns (A=1...): Code(B)=2, Desc(C)=3, Unit(D)=4, Total(E)=5, Const(F)=6, Equip(G)=7
eng_indices = [2, 4, 5, 6, 7]

for i, lbl in enumerate(labels):
    r = 9 + i
    ws_ui[f'B{r}'] = lbl
    ws_ui[f'B{r}'].font = Font(bold=True)
    
    col_idx = eng_indices[i]
    col_letter = openpyxl.utils.get_column_letter(col_idx)
    
    # Logic: Look up in ENGINE sheet using Key (Table_Desc).
    # Key is in Col H (8) of ENGINE.
    # Return Value from Col X.
    # INDEX(ENGINE!Col_X, MATCH(C4&"_"&C5, ENGINE!Col_H, 0))
    
    f = f"=INDEX(ENGINE!{col_letter}:{col_letter}; MATCH($C$4&\"_\"&$C$5; ENGINE!H:H; 0))"
    
    # For numeric columns (Cost)
    if i >= 2: 
        ws_ui[f'C{r}'].number_format = '#,##0'
        # Wrap in IFERROR/N/A handling?
        f = f"=IFERROR({f}; \"-\")"
        
    ws_ui[f'C{r}'] = f

print(f"Saving to {output_excel}...")
wb.save(output_excel)
print("Done.")
