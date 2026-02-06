
import pandas as pd
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.workbook.defined_name import DefinedName
import re

input_excel = r"g:\My Drive\30_RESOURCES (Kho tai nguyen)\30.01_Tinh hieu qua du an\02_Tra dinh muc theo TT 09 va 12\Du_lieu_Co_Tieu_De_Bang.xlsx"
output_excel = r"g:\My Drive\30_RESOURCES (Kho tai nguyen)\30.01_Tinh hieu qua du an\02_Tra dinh muc theo TT 09 va 12\File_Tra_Cuu_Final_Sorted.xlsx"

print("Reading Data...")
df = pd.read_excel(input_excel, header=None)

processed_data = []
current_header_stack = [] 
current_table_name = ""
last_valid_code = ""

# --- PROCESSING ---
for idx, row in df.iterrows():
    row_strs = [str(x).strip() if not pd.isna(x) else "" for x in row]
    table_name = row_strs[0]
    
    match = re.search(r'Bảng (\d+)', table_name, re.IGNORECASE)
    if match:
        tbl_num = int(match.group(1))
        if not (1 <= tbl_num <= 101): continue
    else: continue

    if table_name != current_table_name:
        current_table_name = table_name
        current_header_stack = [] 
        last_valid_code = ""
    
    code = row_strs[1]
    desc = row_strs[2]
    vals = row_strs[3:]
    has_number = any(re.search(r'[\d.,]+', v) for v in vals)

    if has_number and desc:
        val1, val2, val3 = 0, 0, 0
        nums = []
        for v in vals:
            try: nums.append(float(v.replace('.', '').replace(',', '.')))
            except: pass
        if len(nums) >= 1: val1 = nums[0]
        if len(nums) >= 2: val2 = nums[1]
        if len(nums) >= 3: val3 = nums[2]
        
        stack_len = len(current_header_stack)
        if stack_len == 0:
            l1, l2, l3 = desc, "-", "-"
        elif stack_len == 1:
            l1, l2, l3 = current_header_stack[0], desc, "-"
        elif stack_len >= 2:
            l1, l2, l3 = current_header_stack[0], current_header_stack[1], desc
            
        if code: last_valid_code = code; final_code = code
        else: final_code = last_valid_code
        
        # ADD TABLE NUMBER for Sorting
        processed_data.append([tbl_num, table_name, final_code, l1, l2, l3, val1, val2, val3])
        
    else:
        if desc:
            if code:
                current_header_stack = [desc]
                last_valid_code = code
            else:
                if len(current_header_stack) == 0: current_header_stack.append(desc)
                else:
                    if len(current_header_stack) >= 2: current_header_stack[-1] = desc 
                    else: current_header_stack.append(desc) 

# --- SORTING LOGIC ---
# Sort by: Table Number (Int), then L1, L2, L3
processed_data.sort(key=lambda x: (x[0], x[3], x[4], x[5]))

# Generate Menus (Clean Duplicate Lists)
# Menu 1: Table (unique)
tables = []
seen = set()
for r in processed_data:
    if r[1] not in seen: seen.add(r[1]); tables.append(r[1])

# Menu 2: L1
menu_l1 = []
seen = set()
for r in processed_data:
    k = f"{r[1]}_{r[3]}"
    if k not in seen: seen.add(k); menu_l1.append([r[1], r[3]])

# Menu 3: L2
menu_l2 = []
seen = set()
for r in processed_data:
    k = f"{r[1]}_{r[3]}_{r[4]}"
    if k not in seen: seen.add(k); menu_l2.append([r[1], r[3], r[4]])

# Menu 4: L3
menu_l3 = []
seen = set()
for r in processed_data:
    k = f"{r[1]}_{r[3]}_{r[4]}_{r[5]}"
    if k not in seen: seen.add(k); menu_l3.append([r[1], r[3], r[4], r[5]])


# --- EXCEL ---
wb = openpyxl.Workbook()

# MENU
ws_menu = wb.active; ws_menu.title="MENU"
ws_menu.append(["Tbl", "M1K", "M1V", "M2K", "M2V", "M3K", "M3V"])
for i,t in enumerate(tables): ws_menu[f'A{i+2}'] = t 
for i,m in enumerate(menu_l1): ws_menu[f'B{i+2}']=m[0]; ws_menu[f'C{i+2}']=m[1]
for i,m in enumerate(menu_l2): ws_menu[f'D{i+2}']=f"{m[0]}_{m[1]}"; ws_menu[f'E{i+2}']=m[2]
for i,m in enumerate(menu_l3): ws_menu[f'F{i+2}']=f"{m[0]}_{m[1]}_{m[2]}"; ws_menu[f'G{i+2}']=m[3]

def add_dn(name, sheet, col, count):
    wb.defined_names.add(DefinedName(name, attr_text=f"{sheet}!${col}$2:${col}${count+1}"))

add_dn("ListTbl", "MENU", "A", len(tables))
add_dn("M1K", "MENU", "B", len(menu_l1)); add_dn("M1V", "MENU", "C", len(menu_l1))
add_dn("M2K", "MENU", "D", len(menu_l2)); add_dn("M2V", "MENU", "E", len(menu_l2))
add_dn("M3K", "MENU", "F", len(menu_l3)); add_dn("M3V", "MENU", "G", len(menu_l3))

# ENGINE
ws_eng = wb.create_sheet("ENGINE")
ws_eng.append(["Tbl","Code","L1","L2","L3","V1","V2","V3","Key"])
for r in processed_data: # r has tbl_num at 0
    super_key = f"{r[1]}_{r[3]}_{r[4]}_{r[5]}"
    ws_eng.append([r[1], r[2], r[3], r[4], r[5], r[6], r[7], r[8], super_key])
add_dn("EngKey", "ENGINE", "I", len(processed_data))

# UI
ws_ui = wb.create_sheet("TRA_CUU")
ws_ui['B2']="TRA CỨU (SORTED)"; ws_ui['B2'].font=Font(size=14, bold=True)

ws_ui['B4']="1. Bảng:"; ws_ui['C4'].border=Border(bottom=Side(style='thin'))
ws_ui.add_data_validation(DataValidation(type="list", formula1="=ListTbl", allow_blank=True))
ws_ui['C4'].parent.data_validations.dataValidation[-1].add(ws_ui['C4'])

ws_ui['B5']="2. Nhóm:"; ws_ui['C5'].border=Border(bottom=Side(style='thin')); ws_ui['C5'].alignment=Alignment(wrap_text=True)
ws_ui.add_data_validation(DataValidation(type="list", formula1="=OFFSET(M1V;MATCH($C$4;M1K;0)-1;0;COUNTIF(M1K;$C$4);1)", allow_blank=True))
ws_ui['C5'].parent.data_validations.dataValidation[-1].add(ws_ui['C5'])

ws_ui['B6']="3. Loại 1:"; ws_ui['C6'].border=Border(bottom=Side(style='thin')); ws_ui['C6'].alignment=Alignment(wrap_text=True)
ws_ui.add_data_validation(DataValidation(type="list", formula1="=OFFSET(M2V;MATCH($C$4&\"_\"&$C$5;M2K;0)-1;0;COUNTIF(M2K;$C$4&\"_\"&$C$5);1)", allow_blank=True))
ws_ui['C6'].parent.data_validations.dataValidation[-1].add(ws_ui['C6'])

ws_ui['B7']="4. Loại 2:"; ws_ui['C7'].border=Border(bottom=Side(style='thin')); ws_ui['C7'].alignment=Alignment(wrap_text=True)
ws_ui.add_data_validation(DataValidation(type="list", formula1="=OFFSET(M3V;MATCH($C$4&\"_\"&$C$5&\"_\"&$C$6;M3K;0)-1;0;COUNTIF(M3K;$C$4&\"_\"&$C$5&\"_\"&$C$6);1)", allow_blank=True))
ws_ui['C7'].parent.data_validations.dataValidation[-1].add(ws_ui['C7'])


ws_ui['B9']="KẾT QUẢ:"; labels=["Mã","Suất","XD","TB"]; cols=["B","F","G","H"]
for i,l in enumerate(labels):
    r=10+i; ws_ui[f'B{r}']=l; ws_ui[f'B{r}'].font=Font(bold=True)
    ws_ui[f'C{r}'] = f"=INDEX(ENGINE!{cols[i]}:{cols[i]}; MATCH($C$4 & \"_\" & $C$5 & \"_\" & $C$6 & \"_\" & $C$7; EngKey; 0))" # Note: EngKey is now Col I
    if i>=1: ws_ui[f'C{r}'].number_format='#,##0'

ws_menu.sheet_state='hidden'; ws_eng.sheet_state='hidden'
wb.save(output_excel); print("Done.")
