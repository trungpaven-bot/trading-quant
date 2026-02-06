
import pandas as pd
import openpyxl
from openpyxl.styles import Font
import re

input_excel = r"g:\My Drive\30_RESOURCES (Kho tai nguyen)\30.01_Tinh hieu qua du an\02_Tra dinh muc theo TT 09 va 12\Du_lieu_Co_Tieu_De_Bang.xlsx"
output_excel = r"g:\My Drive\30_RESOURCES (Kho tai nguyen)\30.01_Tinh hieu qua du an\02_Tra dinh muc theo TT 09 va 12\File_Tra_Cuu_Lists_Fixed.xlsx"

print("Reading Data...")
df = pd.read_excel(input_excel, header=None)

processed_data = []
current_header_stack = [] 
current_table_name = ""
last_valid_code = ""
data_seen_since_last_header = False

start_processing = False

for idx, row in df.iterrows():
    row_strs = [str(x).strip() if not pd.isna(x) else "" for x in row]
    table_name = row_strs[0]
    
    match = re.search(r'Bảng (\d+)', table_name, re.IGNORECASE)
    if match:
        tbl_num = int(match.group(1))
        if not (1 <= tbl_num <= 101): continue
        if not start_processing: start_processing = True
    else: continue
    
    if table_name != current_table_name:
        current_table_name = table_name
        current_header_stack = [] 
        last_valid_code = ""
        data_seen_since_last_header = False
    
    code = row_strs[1]
    desc = row_strs[2]
    vals = row_strs[3:]
    has_number = any(re.search(r'[\d.,]+', v) for v in vals)

    if has_number and desc:
        data_seen_since_last_header = True
        prefix = " ".join(current_header_stack)
        final_desc = f"{prefix} {desc}" if prefix else desc
        if code: last_valid_code = code; final_code = code
        else: final_code = last_valid_code
        processed_data.append([tbl_num, table_name, final_code, final_desc])
        
    else:
        if desc:
            if code:
                current_header_stack = [desc]
                last_valid_code = code
                data_seen_since_last_header = False
            else:
                if len(current_header_stack) == 0:
                    current_header_stack.append(desc)
                else:
                    if data_seen_since_last_header:
                        current_header_stack[-1] = desc
                    else:
                        current_header_stack.append(desc)
                data_seen_since_last_header = False

processed_data.sort(key=lambda x: x[0]) 

wb = openpyxl.Workbook()
ws = wb.active; ws.title = "LISTS"
ws['A1'] = "Bảng"; ws['C1'] = "Mã hiệu"; ws['D1'] = "Loại công trình"

for i, r in enumerate(processed_data):
    row = i + 2
    ws[f'A{row}'] = r[1] 
    ws[f'C{row}'] = r[2] 
    ws[f'D{row}'] = r[3] 

for cell in ws[1]: cell.font = Font(bold=True)
ws.column_dimensions['A'].width = 40; ws.column_dimensions['C'].width = 15; ws.column_dimensions['D'].width = 80

print(f"Saving {output_excel}...")
wb.save(output_excel)
print("Done.")
