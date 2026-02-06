
import pandas as pd
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.workbook.defined_name import DefinedName
import re

input_excel = r"g:\My Drive\30_RESOURCES (Kho tai nguyen)\30.01_Tinh hieu qua du an\02_Tra dinh muc theo TT 09 va 12\Du_lieu_Co_Tieu_De_Bang.xlsx"
output_excel = r"g:\My Drive\30_RESOURCES (Kho tai nguyen)\30.01_Tinh hieu qua du an\02_Tra dinh muc theo TT 09 va 12\File_Tra_Cuu_3_Cap.xlsx"

print("Reading Data...")
df = pd.read_excel(input_excel, header=None)

# 1. PARSING DATA INTO 3 LEVELS
# structure: [Table, Group(L1), Detail(L2), Unit, Price...]
processed_data = []

current_header_stack = [] 
current_table_name = ""
last_valid_code = ""

start_processing = False

for idx, row in df.iterrows():
    row_strs = [str(x).strip() if not pd.isna(x) else "" for x in row]
    table_name = row_strs[0]
    
    # Table Reset Logic
    if table_name != current_table_name:
        current_table_name = table_name
        current_header_stack = [] 
        last_valid_code = ""
    
    if not start_processing:
        if "Bảng 1" in table_name or "Bang 1" in table_name:
            start_processing = True
        else:
            continue
            
    code = row_strs[1]
    desc = row_strs[2]
    vals = row_strs[3:]
    has_number = any(re.search(r'[\d.,]+', v) for v in vals)

    if has_number and desc:
        # --- DATA ROW ---
        val1, val2, val3 = 0, 0, 0
        nums = []
        for v in vals:
            try: nums.append(float(v.replace('.', '').replace(',', '.')))
            except: pass
        if len(nums) >= 1: val1 = nums[0]
        if len(nums) >= 2: val2 = nums[1]
        if len(nums) >= 3: val3 = nums[2]
        
        # SPLIT INTO GROUP (L1) AND DETAIL (L2)
        # Logic: 
        # If Stack exists: Stack[0] is Group. Everything else + Desc is Detail.
        # If Stack empty: Desc is Group. Detail is "-".
        
        group_l1 = ""
        detail_l2 = ""
        
        if current_header_stack:
            group_l1 = current_header_stack[0] # The Major Header
            
            # The rest form the detail
            rest = current_header_stack[1:]
            rest.append(desc)
            detail_l2 = " - ".join(rest)
        else:
            group_l1 = desc
            detail_l2 = "Chi tiết" # Or specific text
            
        if code: last_valid_code = code; final_code = code
        else: final_code = last_valid_code
        
        processed_data.append([table_name, final_code, group_l1, detail_l2, "1000đ/Dv", val1, val2, val3])
        
    else:
        # --- HEADER ROW ---
        if desc:
            if code:
                current_header_stack = [desc]
                last_valid_code = code
            else:
                if len(current_header_stack) == 0:
                    current_header_stack.append(desc)
                else:
                    # Replace minor header
                    if len(current_header_stack) > 1:
                        current_header_stack[-1] = desc
                    else:
                        current_header_stack.append(desc)

print(f"Parsed {len(processed_data)} rows.")

# 2. GENERATE AUXILIARY LISTS (SORTED)
# We need 3 datasets:
# A. Unique Tables
# B. Unique Groups (Table -> Group)
# C. Engine Data (Table -> Group -> Detail)

# Sort Data First to ensure contiguity
# Key: Table, Group, Detail
processed_data.sort(key=lambda x: (x[0], x[2], x[3])) # Sort by Table, then Group, then Detail

# List A: Tables
unique_tables = []
seen_tables = set()
for r in processed_data:
    if r[0] not in seen_tables:
        seen_tables.add(r[0])
        unique_tables.append(r[0])

# List B: Groups (for Dropdown 2)
# Structure: [Table, Group, LookupKey]
unique_groups = []
seen_groups = set() # Key = Table_Group
for r in processed_data:
    key = f"{r[0]}_{r[2]}"
    if key not in seen_groups:
        seen_groups.add(key)
        unique_groups.append([r[0], r[2]]) # Table, GroupName

print(f"Generated {len(unique_groups)} unique groups.")

# --- BUILD EXCEL ---
wb = openpyxl.Workbook()

# === SHEET 1: ENGINE (The Data) ===
ws_eng = wb.active # Use default first
ws_eng.title = "ENGINE"
headers = ["Table", "Code", "Group", "Detail", "Unit", "Val1", "Val2", "Val3", "SuperKey"]
ws_eng.append(headers)
for i, r in enumerate(processed_data):
    # r = [Tbl, Code, Grp, Dtl, Unit, V1, V2, V3]
    row = [r[0], r[1], r[2], r[3], r[4], r[5], r[6], r[7]]
    # SuperKey for Final Lookup: Table_Group_Detail
    key = f"{r[0]}_{r[2]}_{r[3]}"
    row.append(key)
    ws_eng.append(row)

# === SHEET 2: MENU (The Dropdown Lists) ===
ws_menu = wb.create_sheet("MENU")
ws_menu.append(["TableList", "", "Helper_Tbl", "Helper_Grp", "Key_TblGrp"]) # Headers

# Col A: Unique Tables
for i, t in enumerate(unique_tables):
    ws_menu[f'A{i+2}'] = t

# Col C, D, E: Helper for Dropdown 2 (Map Table to Group)
for i, g in enumerate(unique_groups):
    # g = [Table, Group]
    ws_menu[f'C{i+2}'] = g[0] # Table
    ws_menu[f'D{i+2}'] = g[1] # Group
    ws_menu[f'E{i+2}'] = f"{g[0]}_{g[1]}" # Key

# === DEFINE RANGES (HARD CODED For Safety) ===
# 1. ListTables (Menu A)
wb.defined_names.add(DefinedName("ListTables", attr_text=f"MENU!$A$2:$A${len(unique_tables)+1}"))

# 2. Ranges for Menu Lookup (Menu C and D)
cnt_grp = len(unique_groups)
wb.defined_names.add(DefinedName("Menu_Tbl", attr_text=f"MENU!$C$2:$C${cnt_grp+1}"))
wb.defined_names.add(DefinedName("Menu_Grp", attr_text=f"MENU!$D$2:$D${cnt_grp+1}"))

# 3. Ranges for Engine Lookup (Engine A, C, D)
cnt_data = len(processed_data)
# Engine A: Table (Not strictly needed if we use SuperKey match, but needed for Dropdown 3 source logic)
# Actually Dropdown 3 needs to filter Groups.
# Wait, Dropdown 3 depends on Group Selection.
# So we need to find lines in ENGINE where (Engine_Tbl == C4) AND (Engine_Grp == C5).
# Since data is Sorted by Table then Group, these lines are contiguous!
# So we can use OFFSET based on (Table + Group).

# Helpers for Engine
wb.defined_names.add(DefinedName("Eng_Tbl", attr_text=f"ENGINE!$A$2:$A${cnt_data+1}"))
wb.defined_names.add(DefinedName("Eng_Grp", attr_text=f"ENGINE!$C$2:$C${cnt_data+1}"))
wb.defined_names.add(DefinedName("Eng_Dtl", attr_text=f"ENGINE!$D$2:$D${cnt_data+1}"))
# Helper Key for OFFSET finding: We need a column in ENGINE that combines Table_Group?
# Let's add Column J in Engine = Table_Group
for i in range(cnt_data):
    ws_eng[f'J{i+2}'] = f"=A{i+2}&\"_\"&C{i+2}"
wb.defined_names.add(DefinedName("Eng_Key_TblGrp", attr_text=f"ENGINE!$J$2:$J${cnt_data+1}"))
wb.defined_names.add(DefinedName("Eng_SuperKey", attr_text=f"ENGINE!$I$2:$I${cnt_data+1}"))


# === SHEET 3: TRA_CUU ===
ws_ui = wb.create_sheet("TRA_CUU")
ws_ui['B2']="TRA CỨU 3 CẤP"; ws_ui['B2'].font=Font(size=14, bold=True)

# --- LEVEL 1: TABLE ---
ws_ui['B4']="1. Chọn Bảng:"; ws_ui['C4'].border=Border(bottom=Side(style='thin')); 
dv1 = DataValidation(type="list", formula1="=ListTables", allow_blank=True)
ws_ui.add_data_validation(dv1); dv1.add(ws_ui['C4'])

# --- LEVEL 2: GROUP (Depends on C4) ---
ws_ui['B5']="2. Chọn Nhóm:"; ws_ui['C5'].border=Border(bottom=Side(style='thin')); 
# Logic: Look in MENU sheet. Match C4 in Menu_Tbl. Return Menu_Grp.
dv2_f = "=OFFSET(MENU!$D$2;MATCH($C$4;Menu_Tbl;0)-1;0;COUNTIF(Menu_Tbl;$C$4);1)"
dv2 = DataValidation(type="list", formula1=dv2_f, allow_blank=True)
ws_ui.add_data_validation(dv2); dv2.add(ws_ui['C5'])

# --- LEVEL 3: DETAIL (Depends on C4 + C5) ---
ws_ui['B6']="3. Chọn Chi tiết:"; ws_ui['C6'].border=Border(bottom=Side(style='thin')); ws_ui['C6'].alignment=Alignment(wrap_text=True)
# Logic: Look in ENGINE sheet. Match (C4_C5) in Eng_Key_TblGrp. Return Eng_Dtl.
# Because Engine is sorted, all items for (Table+Group) are together.
dv3_f = "=OFFSET(ENGINE!$D$2;MATCH($C$4&\"_\"&$C$5;Eng_Key_TblGrp;0)-1;0;COUNTIF(Eng_Key_TblGrp;$C$4&\"_\"&$C$5);1)"
dv3 = DataValidation(type="list", formula1=dv3_f, allow_blank=True)
ws_ui.add_data_validation(dv3); dv3.add(ws_ui['C6'])


# --- RESULTS ---
ws_ui['B9']="KẾT QUẢ:"; labels=["Mã","ĐV","Suất","XD","TB"]; cols=["B","E","F","G","H"] # Engine Cols: B(Code), E(Unit), F, G, H
for i,l in enumerate(labels):
    r=10+i; ws_ui[f'B{r}']=l; ws_ui[f'B{r}'].font=Font(bold=True)
    # INDEX/MATCH on SuperKey columns I
    # Key = C4_C5_C6
    f=f"=INDEX(ENGINE!{cols[i]}:{cols[i]}; MATCH($C$4 & \"_\" & $C$5 & \"_\" & $C$6; Eng_SuperKey; 0))"
    if i>=2: ws_ui[f'C{r}'].number_format='#,##0'
    ws_ui[f'C{r}']=f

# Hide helpers
ws_menu.sheet_state = 'hidden'
ws_eng.sheet_state = 'hidden'

print(f"Saving {output_excel}...")
wb.save(output_excel)
print("Done.")
