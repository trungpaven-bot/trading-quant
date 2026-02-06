
import openpyxl

file_path = r"g:\My Drive\00_INBOX (xu ly nhanh)\tra TMDT.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True) # data_only=True to read values

def search_in_sheet(sheet_name, keywords):
    print(f"\n--- Searching in {sheet_name} ---")
    if sheet_name not in wb.sheetnames:
        print("Sheet not found.")
        return
        
    ws = wb[sheet_name]
    found_count = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.value:
                val_str = str(cell.value).lower()
                for kw in keywords:
                    if kw.lower() in val_str:
                        print(f"Found '{kw}' at {cell.coordinate}: {cell.value}")
                        found_count += 1
                        if found_count > 20: return # Stop after some matches to avoid spam

keywords = [
    "Cấp I", "Cấp II", "Cấp III", "Cấp IV", 
    "1 bước", "2 bước", "3 bước",
    "Thẩm định", "Kiểm toán",
    "Độc lập", "Nội bộ" # Keywords for audit/verification types
]

for sheet in wb.sheetnames:
    search_in_sheet(sheet, keywords)
