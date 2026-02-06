import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

input_file = 'g:/My Drive/30_RESOURCES (Kho tai nguyen)/30.01_Tinh hieu qua du an/02_Tra dinh muc theo TT 09 va 12/2026.01.19 Tien do du an CCN PDP 1 - Chi phi HĐ V1.0_T.xlsx'
output_file = 'g:/My Drive/30_RESOURCES (Kho tai nguyen)/30.01_Tinh hieu qua du an/02_Tra dinh muc theo TT 09 va 12/2026.01.19 Tien do du an CCN PDP 1 - Chi phi HĐ V1.0_T_Professional.xlsx'

# Load the workbook to preserve details
wb_src = openpyxl.load_workbook(input_file)

# Create new workbook
wb_new = openpyxl.Workbook()
wb_new.remove(wb_new.active) # Remove default sheet

# Define Styles
header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid") # Dark Slate Blue
header_font = Font(name='Arial', bold=True, color="FFFFFF", size=10)
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

row_fill_even = PatternFill(start_color="F7F9F9", end_color="F7F9F9", fill_type="solid") # Very Light Grey
row_fill_odd = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")   # White

border_thin = Side(border_style="thin", color="D7DBDD")
border_style = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)

# Currency Format
currency_fmt = '#,##0 "₫"'
date_fmt = 'DD/MM/YYYY'

def apply_professional_style(ws, sheet_name):
    # Determine header row based on sheet
    if sheet_name == 'Tien do':
        header_row_idx = 4
        date_cols = [4, 5]
        cost_cols = [9, 10]
        freeze = 'C5'
    else: # TT12
        header_row_idx = 1
        date_cols = []
        cost_cols = []
        freeze = 'A2'
    
    # Iterate through rows and columns
    for row in ws.iter_rows():
        for cell in row:
            # Apply border to all cells
            cell.border = border_style
            cell.font = Font(name='Arial', size=10)
            
            # Header Styling
            if cell.row == header_row_idx:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_align
            
            # Data Styling (Alternating Rows)
            elif cell.row > header_row_idx:
                if cell.row % 2 == 0:
                    cell.fill = row_fill_even
                else:
                    cell.fill = row_fill_odd
                
                # Sheet Specific Column Formatting
                if sheet_name == 'Tien do':
                    if cell.column in date_cols: 
                        cell.number_format = date_fmt
                        cell.alignment = Alignment(horizontal='center')
                    if cell.column in cost_cols:
                        cell.number_format = currency_fmt
                    # "Nội dung" alignment
                    if cell.column == 2:
                        cell.alignment = Alignment(wrap_text=True, vertical='top')

    # Auto-fit Column Widths (simplified)
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.1
        if adjusted_width > 50: adjusted_width = 50 
        if adjusted_width < 5: adjusted_width = 8
        ws.column_dimensions[column].width = adjusted_width

    # Override widths
    if sheet_name == 'Tien do':
        ws.column_dimensions['B'].width = 50
        ws.freeze_panes = freeze

for sheet_name in wb_src.sheetnames:
    src_ws = wb_src[sheet_name]
    new_ws = wb_new.create_sheet(title=sheet_name)
    
    # Copy Data
    for row in src_ws.iter_rows():
        for cell in row:
            new_val = cell.value
            new_ws.cell(row=cell.row, column=cell.column, value=new_val)
            
    # Copy Merged Cells
    if src_ws.merged_cells:
        for merged_cell in src_ws.merged_cells.ranges:
            new_ws.merge_cells(str(merged_cell))
            
    apply_professional_style(new_ws, sheet_name)

wb_new.save(output_file)
print(f"File created at: {output_file}")
