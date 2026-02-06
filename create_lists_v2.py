
import openpyxl

file_path = r"g:\My Drive\00_INBOX (xu ly nhanh)\tra TMDT.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=False)
ws = wb["TT09"]

data_lists = {
    "LoaiCongTrinh": ["Công trình dân dụng", "Công trình công nghiệp", "Công trình giao thông", "Công trình NN&PTNT", "Công trình hạ tầng kỹ thuật"],
    "KieuCongTrinh": ["Công trình xây dựng mới", "Công trình thiết kế mẫu, điển hình", "Công trình sửa chữa, cải tạo, mở rộng"],
    "ThamDinh": ["Cơ quan chuyên môn về xây dựng thẩm định", "Chủ đầu tư tổ chức thẩm định"],
    "KiemToan": ["Thuê kiểm toán độc lập", "Quyết toán nội bộ (Không thuê)"],
    "CapCongTrinh": ["Cấp đặc biệt", "Cấp I", "Cấp II", "Cấp III", "Cấp IV"],
    "SoBuocThietKe": ["Thiết kế 1 bước (BCKTKT)", "Thiết kế 2 bước (Cơ sở + BVTC)", "Thiết kế 3 bước (Cơ sở + KT + BVTC)"]
}

start_col_idx = 27 # AA
start_row = 1000
source_map = {}

print("Writing data lists...")
for i, (key, values) in enumerate(data_lists.items()):
    col_idx = start_col_idx + i
    col_letter = openpyxl.utils.get_column_letter(col_idx)
    
    # Write Header
    ws.cell(row=start_row, column=col_idx, value=key)
    
    # Write Values
    for j, val in enumerate(values):
        ws.cell(row=start_row + 1 + j, column=col_idx, value=val)
        
    end_row = start_row + len(values)
    range_ref = f"${col_letter}${start_row + 1}:${col_letter}${end_row}"
    source_map[key] = f"TT09!{range_ref}"

wb.save(file_path)
print("Lists created successfully.")
for k, v in source_map.items():
    print(f"{k}: {v}")
