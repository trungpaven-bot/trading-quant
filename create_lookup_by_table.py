
import pandas as pd
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.workbook.defined_name import DefinedName
import re

input_excel = r"g:\My Drive\30_RESOURCES (Kho tai nguyen)\30.01_Tinh hieu qua du an\02_Tra dinh muc theo TT 09 va 12\Du_lieu_Co_Tieu_De_Bang.xlsx"
output_excel = r"g:\My Drive\30_RESOURCES (Kho tai nguyen)\30.01_Tinh hieu qua du an\02_Tra dinh muc theo TT 09 va 12\File_Tra_Cuu_Theo_Bang_Final.xlsx"

print("Reading Data with Tables as Headers...")
df = pd.read_excel(input_excel, header=None)

# 1. CLEAN DATA
refined_data = []

# Assuming Col 0 is Table Name
# Col 1 is Code, Col 2 is Desc... based on extract_with_titles logic
# row_vals = [title] + [cell_txts...]

current_code = ""

for idx, row in df.iterrows():
    row_strs = [str(x).strip() if not pd.isna(x) else "" for x in row]
    
    table_name = row_strs[0] 
    # Skip if table name is suspiciously short or strictly 'Bảng X (Không tên)'? 
    # Keep it, user can ignore.
    
    # Check data content
    code = row_strs[1]
    desc = row_strs[2] if len(row_strs)>2 else ""
    
    # Code filling logic
    if code: current_code = code
    
    vals = row_strs[3:]
    has_number = any(re.search(r'[\d.,]+', v) for v in vals)
    
    if desc and has_number:
        # Standardize 3 col values
        val1, val2, val3 = 0, 0, 0
        nums = []
        for v in vals:
            clean = v.replace('.', '').replace(',', '.')
            try: nums.append(float(clean))
            except: pass
        if len(nums) >= 1: val1 = nums[0]
        if len(nums) >= 2: val2 = nums[1]
        if len(nums) >= 3: val3 = nums[2]
        
        # Structure: TableName, Code, Desc, Unit, Total, Const, Equip
        # Note: Original extraction might have varied column positions.
        # But generally: Code, Desc, Unit, Val... 
        # Let's assume standardized Unit "1000đ..."
        unit = "1000d/m2"
        refined_data.append([table_name, current_code, desc, unit, val1, val2, val3])

print(f"Refined {len(refined_data)} rows.")

# 2. SETUP WORKBOOK
wb = openpyxl.Workbook()

# --- SHEET: DATA ---
ws_data = wb.active
ws_data.title = "DATA"
headers = ["Tên Bảng", "Mã hiệu", "Loại công trình", "Đơn vị tính", "Suất vốn đầu tư", "Chi phí Xây dựng", "Chi phí Thiết bị"]
ws_data.append(headers)
for row in refined_data:
    ws_data.append(row)

# Formatting
header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
for cell in ws_data[1]: 
    cell.fill = header_fill
    cell.font = Font(bold=True)
ws_data.column_dimensions['A'].width = 40
ws_data.column_dimensions['C'].width = 50

# --- SHEET: LISTS ---
ws_lists = wb.create_sheet("LISTS")
ws_lists.sheet_state = 'hidden'

# List 1: Unique Table Names
unique_tables = sorted(list(set(r[0] for r in refined_data)))
ws_lists['A1'] = "UniqueTables"
for i, t in enumerate(unique_tables):
    ws_lists[f'A{i+2}'] = t

wb.defined_names.add(DefinedName("List_Tables", attr_text=f"LISTS!$A$2:$A${len(unique_tables)+1}"))

# List 2: Helper Columns for Dependent Dropdown
# We need sorted lists by Table Name
ws_lists['C1'] = "TableCol"
ws_lists['D1'] = "DescCol"

sorted_data = sorted(refined_data, key=lambda x: x[0]) # Sort by Table Name
for i, row in enumerate(sorted_data):
    ws_lists[f'C{i+2}'] = row[0] # Table
    ws_lists[f'D{i+2}'] = row[2] # Description

max_r = len(sorted_data) + 1
wb.defined_names.add(DefinedName("Col_Table", attr_text=f"LISTS!$C$2:$C${max_r}"))
wb.defined_names.add(DefinedName("Col_Desc", attr_text=f"LISTS!$D$2:$D${max_r}"))

# --- SHEET: TRA_CUU ---
ws_ui = wb.create_sheet("TRA_CUU")
ws_ui.column_dimensions['B'].width = 25
ws_ui.column_dimensions['C'].width = 50

ws_ui['B2'] = "TRA CỨU SUẤT VỐN ĐẦU TƯ 2024"
ws_ui['B2'].font = Font(size=14, bold=True)

# L1: TABLE
ws_ui['B4'] = "1. Chọn Bảng:"
ws_ui['B4'].font = Font(bold=True)
c4 = ws_ui['C4']
c4.border = Border(bottom=Side(style='thin'))
dv1 = DataValidation(type="list", formula1="=List_Tables", allow_blank=True)
ws_ui.add_data_validation(dv1)
dv1.add(c4)

# L2: PROJECT
ws_ui['B5'] = "2. Chọn Loại công trình:"
ws_ui['B5'].font = Font(bold=True)
c5 = ws_ui['C5']
c5.border = Border(bottom=Side(style='thin'))
c5.alignment = Alignment(wrap_text=True)

# Formula: OFFSET/MATCH/COUNTIF on Table Column
dv2_formula = "=OFFSET(LISTS!$D$2;MATCH($C$4;Col_Table;0)-1;0;COUNTIF(Col_Table;$C$4);1)"
dv2 = DataValidation(type="list", formula1=dv2_formula, allow_blank=True)
dv2.errorTitle = ""
ws_ui.add_data_validation(dv2)
dv2.add(c5)

# RESULTS
res_labels = ["Mã hiệu", "Đơn vị tính", "Suất vốn đầu tư", "Chi phí Xây dựng", "Chi phí Thiết bị"]
col_indices = [2, 4, 5, 6, 7] # DATA cols: B(2), D(4), E(5)... NO data has 7 cols?
# DATA: A=Table, B=Code, C=Desc, D=Unit, E=Total, F=Const, G=Equip
# Index: 1, 2, 3, 4, 5, 6, 7
map_indices = [2, 4, 5, 6, 7]

ws_ui['B8'] = "KẾT QUẢ:"
ws_ui['B8'].font = Font(bold=True, underline="single")

for i, lbl in enumerate(res_labels):
    r = 9 + i
    ws_ui[f'B{r}'] = lbl
    ws_ui[f'B{r}'].font = Font(bold=True)
    
    col_idx = map_indices[i]
    col_letter = openpyxl.utils.get_column_letter(col_idx)
    
    # Code and Unit are Strings -> Cannot use SUMIFS. Use INDEX/MATCH (Array) or DGET?
    # Simple workaround: LOOKUP or INDEX match using MULTIPLE Criteria helper key.
    
    # Let's add a Helper Key in DATA sheet (Col H) = Table + Desc
    # But I can't modify DATA sheet logic mid-loop. 
    # Use SUMIFS for Cost (Numbers).
    # Use generic manual filtering for Code/Unit.
    
    if i < 2: # Code & Unit
        # LOOKUP(2, 1/((ColTable=C4)*(ColDesc=C5)), ColResult) -> Requires CSE.
        # Let's try XLOOKUP syntax (Excel 2019+)? No, keep compatibility.
        # Let's just create a Helper Key column in DATA at the end.
        pass
    else:
        # SUMIFS
        f = f"=SUMIFS(DATA!{col_letter}:{col_letter}; DATA!$A:$A; $C$4; DATA!$C:$C; $C$5)"
        ws_ui[f'C{r}'].number_format = '#,##0'
        ws_ui[f'C{r}'] = f

# Add Helper Key to DATA for Code Lookup
ws_data['H1'] = "HelperKey"
for idx in range(2, ws_data.max_row + 1):
    # =A2 & "_" & C2
    ws_data[f'H{idx}'] = f"=A{idx}&\"_\"&C{idx}"

# Fill formulas for Code/Unit using VLOOKUP on Helper Key?
# No, Vlookup needs key first. INDEX/MATCH on Helper Column H.
# MATCH(C4&"_"&C5, H:H, 0)
ws_ui['C9'] = f"=INDEX(DATA!B:B; MATCH($C$4&\"_\"&$C$5; DATA!H:H; 0))" # Code
ws_ui['C10'] = f"=INDEX(DATA!D:D; MATCH($C$4&\"_\"&$C$5; DATA!H:H; 0))" # Unit

print(f"Saving to {output_excel}...")
wb.save(output_excel)
print("Done.")
