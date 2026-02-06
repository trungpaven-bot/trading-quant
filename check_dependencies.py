
import openpyxl

file_path = r"g:\My Drive\00_INBOX (xu ly nhanh)\tra TMDT.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=False)

# Check content of TMDT sheet at C88, C94, C99, C118, C128
ws_tmdt = wb["TMDT"]
cells = [88, 94, 99, 118, 128]
print("--- Checking candidate Config cells in TMDT sheet ---")
for r in cells:
    cell = ws_tmdt.cell(row=r, column=3) # Col C
    print(f"TMDT!C{r}: {cell.value}")

# Check TT09 specific rows for cost calculations
ws_tt09 = wb["TT09"]
print("\n--- Searching for specific Cost Rows in TT09 ---")
target_costs = ["Chi phí thiết kế", "Chi phí thẩm định", "Chi phí kiểm toán", "Chi phí giám sát"]

for r in range(1, 300):
    val = ws_tt09.cell(row=r, column=2).value # Col B
    if val and isinstance(val, str):
        for target in target_costs:
            if target.lower() in val.lower():
                # Print formula in Col E, F, G, H
                formulas = []
                for c in range(3, 10):
                    c_val = ws_tt09.cell(row=r, column=c).value
                    if isinstance(c_val, str) and c_val.startswith("="):
                         formulas.append(f"Col {c}: {c_val}")
                print(f"Row {r}: {val}")
                if formulas:
                    print(f"  Formulas: {formulas}")
