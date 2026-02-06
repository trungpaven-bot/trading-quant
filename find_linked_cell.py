
import openpyxl

file_path = r"g:\My Drive\00_INBOX (xu ly nhanh)\tra TMDT.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)
ws = wb["TT09"]

print("--- Searching for 'Công trình dân dụng' in first 20 rows ---")
for r in range(1, 21):
    for c in range(1, 21):
        cell = ws.cell(row=r, column=c)
        if cell.value == "Công trình dân dụng":
            print(f"FOUND at {cell.coordinate}")
        elif cell.value == 1: # Maybe it's an index?
             print(f"Index 1 found at {cell.coordinate}")

print("--- Dumping E59, F58, E61, E62 values ---")
for coord in ["E59", "F58", "E61", "E62"]:
    print(f"{coord}: {ws[coord].value}")
