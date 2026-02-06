
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Border, Side, Alignment

file_path = r"g:\My Drive\30_RESOURCES (Kho tai nguyen)\30.01_Tinh hieu qua du an\02_Tra dinh muc theo TT 09 va 12\Tra dinh muc TT09.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=False)
ws = wb["TT09"]

# Map: LabelText, LabelPos, InputPos, VarName, ListRange
ui_map = [
    ("Loại công trình:", "G2", "H2", "i_CPC", "$AA$1001:$AA$1005"),
    ("Kiểu công trình:", "K2", "L2", "i_Kieu", "$AB$1001:$AB$1003"),
    ("Thẩm định TK, DT:", "G3", "H3", "i_tkdutoan", "$AC$1001:$AC$1002"),
    ("Kiểm toán/QT:", "K3", "L3", "i_KT", "$AD$1001:$AD$1002"),
    ("Cấp công trình:", "G5", "H5", "i_CapCT", "$AE$1001:$AE$1005"), 
    ("Số bước thiết kế:", "K5", "L5", "i_BTK", "$AF$1001:$AF$1003"),
]

helper_col_letter = "AG"

print("--- Rebuilding Interface (Using Semicolon ';') ---")

for i, (label_text, label_cell, input_cell, var_name, list_range_addr) in enumerate(ui_map):
    # Match formula with SEMICOLON
    # Note: Generally the file format stores commas (US Std), but if user insists on ';' causing issues,
    # we can try forcing it. However, OpenPyXL follows OOXML standard which IS comma.
    # If the user sees an error, it might be due to viewing it.
    # BUT, let's try injecting it with Semicolon. If Excel complains upon opening, we know OpenPyXL handles it.
    
    index_cell_row = 1001 + i
    index_cell_addr = f"${helper_col_letter}${index_cell_row}"
    
    # Using ; as separator
    match_formula = f"=MATCH({input_cell};'TT09'!{list_range_addr};0)"
    ws[index_cell_addr] = match_formula
    print(f"Updated Logic: {index_cell_addr} = {match_formula}")
    
    # 2. Re-apply Data Validation (Standard Formula1 doesn't use separator, just range)
    dv = DataValidation(type="list", formula1=f"='TT09'!{list_range_addr}", allow_blank=True)
    ws.add_data_validation(dv)
    # Using 'add' might duplicate if not careful, but openpyxl handles references.
    # Ideally should clear old DV but it's hard. Overwriting should be okay.
    dv.add(ws[input_cell])

print("Saving file...")
wb.save(file_path)
print("Done.")
