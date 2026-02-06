
import openpyxl

file_path = r"g:\My Drive\00_INBOX (xu ly nhanh)\tra TMDT.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)
ws = wb["TMDT"]

print("--- Inspecting TMDT Sheet (Row 80-140) for Labels ---")
# If TMDT has labels like "Loai Config", "Kieu Config" next to Col C
# It will confirm our mapping.
for r in [88, 94, 99, 118, 128]:
    # Check Col B (Label?) and C (Value)
    label = ws.cell(row=r, column=2).value
    val = ws.cell(row=r, column=3).value
    print(f"Row {r}: {label} | {val}")
    
# Scan nearby rows for the missing ones (Kiểm toán, Kiểu công trình)
print("\n--- Scanning for 'Kiểu', 'Kiểm toán' in TMDT ---")
for r in range(1, 200):
    for c in range(1, 10):
        val = ws.cell(row=r, column=c).value
        s = str(val).lower() if val else ""
        if "kiểu" in s or "kiểm toán" in s:
            print(f"Found '{val}' at ({r},{c})")
