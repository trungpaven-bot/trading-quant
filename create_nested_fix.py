
import pandas as pd
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.workbook.defined_name import DefinedName
import re

input_excel = r"g:\My Drive\30_RESOURCES (Kho tai nguyen)\30.01_Tinh hieu qua du an\02_Tra dinh muc theo TT 09 va 12\Du_lieu_Co_Tieu_De_Bang.xlsx"
output_excel = r"g:\My Drive\30_RESOURCES (Kho tai nguyen)\30.01_Tinh hieu qua du an\02_Tra dinh muc theo TT 09 va 12\File_Tra_Cuu_Final_Nested.xlsx"

print("Reading Data...")
df = pd.read_excel(input_excel, header=None)

processed_data = []

# State variables
current_header_stack = [] # List of header strings found before a data row
last_valid_code = ""

start_processing = False

for idx, row in df.iterrows():
    row_strs = [str(x).strip() if not pd.isna(x) else "" for x in row]
    table_name = row_strs[0]
    
    if not start_processing:
        if "Bảng 1" in table_name or "Bang 1" in table_name:
            start_processing = True
        else:
            continue
            
    code = row_strs[1]
    desc = row_strs[2]
    
    # Check values
    vals = row_strs[3:]
    has_number = any(re.search(r'[\d.,]+', v) for v in vals)

    # --- MULTI-LEVEL LOGIC ---
    if has_number and desc:
        # DATA ROW FOUND
        
        # 1. Calculate Values
        val1, val2, val3 = 0, 0, 0
        nums = []
        for v in vals:
            clean = v.replace('.', '').replace(',', '.')
            try: nums.append(float(clean))
            except: pass
        if len(nums) >= 1: val1 = nums[0]
        if len(nums) >= 2: val2 = nums[1]
        if len(nums) >= 3: val3 = nums[2]
        
        # 2. Construct Full Name
        # Combine everything in stack + current desc
        # Stack might contain "Cong trinh..." and "Cot anten..."
        prefix = " ".join(current_header_stack)
        
        if prefix:
            final_desc = f"{prefix} {desc}"
        else:
            final_desc = desc
            
        # 3. Handle Code
        if code:
            last_valid_code = code # Update code owner
            final_code = code
        else:
            final_code = last_valid_code # Inherit
            
        processed_data.append([table_name, final_code, final_desc, "1000đ/Dv", val1, val2, val3])
        
    else:
        # HEADER ROW (No money)
        # Decision: Does this header REPLACE previous headers or APPEND?
        # In Excel raw extraction, hierarchical data usually flows sequentially.
        # Case A: Indented sub-header -> We don't know indentation.
        # Case B: New major header replaces old major header.
        
        # Strategy:
        # If this row has a Code (e.g. 11120.05), it's a MAJOR header. Reset stack.
        # If this row has NO Code, it's a SUB header? 
        
        if desc:
            if code:
                # New Major Block (e.g. 11120.05 Nhà 4-5 tầng)
                current_header_stack = [desc]
                last_valid_code = code
            else:
                # Sub Header (e.g. "Cột anten 30m" or "Diện tích <50m2")
                # Problem: Does "Cột anten 45m" replace "Cột anten 30m"? Yes.
                # Does it replace "Công trình đài phát"? No.
                
                # Heuristic: If the stack gets too deep (>2), pop the last one?
                # Or simply: just keep the LAST non-empty header line found?
                # In your image:
                # Row 1: Công trình...
                # Row 2: Cột anten...
                # Row 3: Data.
                # Row X: Cột anten new...
                
                # Aggressive Reset Logic to avoid super-long duplicate names:
                # If we encounter a new text line, and we ALREADY have a text line in stack...
                # It's safer to Assume it replaces the last part.
                
                if len(current_header_stack) > 0:
                    # Check if this looks like a continuation or a replacement
                    # For now, let's Append if < 2 items, Replace if >= 2?
                    # Let's try simple Append. User can edit if too long.
                    # Wait, if we just append, "30m" -> "45m", stack becomes "30m 45m". Bad.
                    
                    # Better Logic:
                    # Just keep the "Current Active Context".
                    # For the complex Table 85:
                    # The "Công trình..." line is physically separated.
                    # Let's use the fact that Data Clears Stack? No.
                    
                    # Let's use a "Single Prefix" logic for safety.
                    # Store only the IMMEDIATE previous text line as prefix.
                    # And if there was a Code-Header before that, add that too.
                    
                    # Logic 2.0:
                    # Stack[0] = Last Major Header (With Code)
                    # Stack[1] = Last Minor Header (No Code)
                    
                    if code:
                        current_header_stack = [desc] # Major
                    else:
                         # Append to Minor slot
                         # If we already have a Minor slot (len 2), replace it.
                         if len(current_header_stack) >= 2:
                             current_header_stack[1] = desc
                         elif len(current_header_stack) == 1:
                             current_header_stack.append(desc)
                         else:
                             current_header_stack.append(desc) # No major yet
                else:
                    current_header_stack.append(desc)


# --- BUILD EXCEL ---
wb = openpyxl.Workbook()

# 1. LISTS
ws_lists = wb.active; ws_lists.title="LISTS"
ws_lists['A1']="Bảng"; ws_lists['C1']="Mã"; ws_lists['D1']="Tên CT"

# Write Content
for i, r in enumerate(processed_data):
    row=i+2
    ws_lists[f'A{row}'] = r[0] 
    ws_lists[f'C{row}'] = r[1]
    ws_lists[f'D{row}'] = r[2]

# Unique Tables (Compact Z)
unique_tables = []
seen = set()
for r in processed_data:
    if r[0] not in seen:
        seen.add(r[0]); unique_tables.append(r[0])
for i, t in enumerate(unique_tables): ws_lists[f'Z{i+2}'] = t

# Hardcode Ranges (The "Hard Fix" form previous step)
cnt_data = len(processed_data)
cnt_tbl = len(unique_tables)

wb.defined_names.add(DefinedName("ListTables", attr_text=f"LISTS!$Z$2:$Z${cnt_tbl+1}"))
wb.defined_names.add(DefinedName("ColTbl", attr_text=f"LISTS!$A$2:$A${cnt_data+1}"))
wb.defined_names.add(DefinedName("ColDesc", attr_text=f"LISTS!$D$2:$D${cnt_data+1}"))


# 2. ENGINE
ws_eng = wb.create_sheet("ENGINE")
ws_eng.append(["", "Bảng", "Bảng2", "Mã", "Tên", "ĐV", "Suất", "XD", "TB", "Key"])
for i, r in enumerate(processed_data):
    row=i+2
    ws_eng.cell(row,2,r[0]); ws_eng.cell(row,3,r[0]); ws_eng.cell(row,4,r[1]); ws_eng.cell(row,5,r[2])
    ws_eng.cell(row,6,r[3]); ws_eng.cell(row,7,r[4]); ws_eng.cell(row,8,r[5]); ws_eng.cell(row,9,r[6])
    ws_eng.cell(row,10).value = f'=B{row} & "_" & E{row}'

# 3. TRA_CUU
ws_ui = wb.create_sheet("TRA_CUU")
ws_ui['B2']="TRA CỨU (NESTED FIX)"; ws_ui['B2'].font=Font(size=14, bold=True)

# C4 Table
ws_ui['B4']="1. Chọn Bảng:"; ws_ui['C4'].border=Border(bottom=Side(style='thin'))
dv1 = DataValidation(type="list", formula1="=ListTables", allow_blank=True)
ws_ui.add_data_validation(dv1); dv1.add(ws_ui['C4'])

# C5 Project
ws_ui['B5']="2. Chọn CT:"; ws_ui['C5'].border=Border(bottom=Side(style='thin')); ws_ui['C5'].alignment=Alignment(wrap_text=True)
dv2_f = "=OFFSET(LISTS!$D$2;MATCH($C$4;ColTbl;0)-1;0;COUNTIF(ColTbl;$C$4);1)"
dv2 = DataValidation(type="list", formula1=dv2_f, allow_blank=True)
ws_ui.add_data_validation(dv2); dv2.add(ws_ui['C5'])

# Results
ws_ui['B8']="KẾT QUẢ:"; labels=["Mã","ĐV","Suất","XD","TB"]; cols=["D","F","G","H","I"]
for i,l in enumerate(labels):
    r=9+i; ws_ui[f'B{r}']=l; ws_ui[f'B{r}'].font=Font(bold=True)
    f=f"=INDEX(ENGINE!{cols[i]}:{cols[i]}; MATCH($C$4 & \"_\" & $C$5; ENGINE!$J:$J; 0))"
    if i>=2: ws_ui[f'C{r}'].number_format='#,##0'
    ws_ui[f'C{r}']=f

print(f"Saving {output_excel}...")
wb.save(output_excel)
print("Done.")
