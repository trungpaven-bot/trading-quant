
import openpyxl

file_path = r"g:\My Drive\00_INBOX (xu ly nhanh)\tra TMDT.xlsx"
sheet_name = "TT09"

wb = openpyxl.load_workbook(file_path, data_only=False)
ws = wb[sheet_name]

print("--- Searching for Keywords ---")
keywords = ["nội suy", "Gxd", "Gtb", "CP", "Tỷ lệ", "%"]
found_cells = []

# Scan first 20 columns, first 100 rows for keywords
for r in range(1, 100):
    for c in range(1, 20):
        cell = ws.cell(row=r, column=c)
        val = str(cell.value) if cell.value else ""
        for k in keywords:
            if k.lower() in val.lower():
                found_cells.append(f"{cell.coordinate}: {val}")

print("\n".join(found_cells[:20]))

print("\n--- Searching for unique formulas ---")
formulas = set()
for r in range(1, 200): # Scan more rows
    for c in range(1, 20):
        cell = ws.cell(row=r, column=c)
        val = cell.value
        if isinstance(val, str) and val.startswith("="):
            formulas.add(val)
            if len(formulas) > 10: break
    if len(formulas) > 10: break

for f in formulas:
    print(f)
