
import openpyxl

file_path = r"g:\My Drive\00_INBOX (xu ly nhanh)\tra TMDT.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)
ws = wb["TT09"]

print("--- Locating Dropdown Labels and Target Cells ---")
# The user screenshot shows a layout. Let's find the text and deduce the cell to the right.
labels_to_find = {
    "Loại công trình": "LoaiCongTrinh",
    "Kiểu công trình": "KieuCongTrinh",
    "Thẩm định thiết kế": "ThamDinh", # Partial match needed
    "Kiểm toán": "KiemToan", # Partial
    "Cấp công trình": "CapCongTrinh",
    "Số bước thiết kế": "SoBuocThietKe"
}

target_map = {}

for r in range(1, 10): # Header area
    for c in range(1, 20):
        cell = ws.cell(row=r, column=c)
        val = str(cell.value).lower() if cell.value else ""
        
        for label, key in labels_to_find.items():
            if label.lower() in val:
                # Found label. The dropdown is likely to the RIGHT.
                # Check next 1-2 cells for emptiness or existing validation.
                # In the screenshot, "Loại công trình" label is at ~F2/G2?, Dropdown is wide.
                print(f"Found Label '{label}' at {cell.coordinate}")
                
                # Heuristic: The cell to the right that has a border or is merged?
                # Let's just assume it's the immediate next cell for now, or scan for a merged range start.
                target_cell = ws.cell(row=r, column=c+1)
                print(f"  -> Candidate Target: {target_cell.coordinate} (Value: {target_cell.value})")
                target_map[key] = target_cell.coordinate

print("\n--- Detected Target Map ---")
print(target_map)
