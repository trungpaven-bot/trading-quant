
import pandas as pd
import re
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

input_excel = r"g:\My Drive\30_RESOURCES (Kho tai nguyen)\30.01_Tinh hieu qua du an\02_Tra dinh muc theo TT 09 va 12\Du_lieu_Suat_von_dau_tu_2024.xlsx"
output_excel = r"g:\My Drive\30_RESOURCES (Kho tai nguyen)\30.01_Tinh hieu qua du an\02_Tra dinh muc theo TT 09 va 12\File_Tra_Cuu_Suat_Von_Dau_Tu_2024.xlsx"

print("Reading raw data...")
df = pd.read_excel(input_excel, header=None)

# --- Step 1: Normalize & Extract Data ---
# Logic: Find rows where the 1st column is a Code (e.g. 11000.10) OR just capture valid description rows with costs.
# Word tables are messy. Let's try to assume: 
# Col 0: Code (Optional)
# Col 1: Description
# Col 2: Total Cost / Cost 1
# ...

cleaned_data = []
headers = ["Mã hiệu", "Loại công trình", "Đơn vị tính", "Suất vốn đầu tư (Nghìn đ)", "Chi phí Xây dựng", "Chi phí Thiết bị", "Chi phí Khác/Ghi chú"]

# Heuristic: Valid data rows usually have a number in column 2 or 3.
for idx, row in df.iterrows():
    # Convert row to list of strings
    row_strs = [str(x).strip() if not pd.isna(x) else "" for x in row]
    
    # Skip empty lines or purely divider lines
    if not any(row_strs): continue
    if "--- BẢNG" in row_strs[0]: continue
    
    code = row_strs[0]
    desc = row_strs[1]
    
    # Try to identify values columns. 
    # Usually Suat Von Dau Tu tables have: [Code] [Description] [Unit] [Total] [Construction] [Equipment]
    # Sometimes just [Description] [Value]
    
    # Simple Heuristic: If we find a number in cols 2,3,4, capture it.
    # We will just dump the likely useful rows into the DB.
    
    # Filter: Description must exist
    if not desc: continue
    
    # Merge row into standard format
    # Only pick rows where at least one of col 2,3,4,5 looks like a float number
    is_data_row = False
    for c in row_strs[2:6]:
        if re.search(r'^\d+([.,]\d+)?$', c):
            is_data_row = True
            break
            
    if is_data_row:
        # Normalize Data
        entry = {
            "Mã hiệu": code,
            "Loại công trình": desc,
            "Đơn vị": row_strs[2] if len(row_strs)>2 and not re.match(r'[\d.,]+', row_strs[2]) else "Đơn vị chuẩn",
            "Val1": row_strs[2] if len(row_strs)>2 else "",
            "Val2": row_strs[3] if len(row_strs)>3 else "",
            "Val3": row_strs[4] if len(row_strs)>4 else "",
            "Val4": row_strs[5] if len(row_strs)>5 else ""
        }
        
        # Adjust for shifting columns (sometimes Unit is missing)
        # If Col 2 is number, then Unit is likely implicit or in Description
        if re.match(r'[\d.,]+', entry["Val1"]):
             entry["Val2"] = entry["Val1"] # Shift values
             entry["Val3"] = result = row_strs[3] if len(row_strs)>3 else ""
             entry["Val4"] = row_strs[4] if len(row_strs)>4 else ""
             entry["Val1"] = "Tổng/Suất vốn" # Placeholder for implicit total
             entry["Đơn vị"] = "1000đ/m2/đơn vị"

        cleaned_data.append([entry["Mã hiệu"], entry["Loại công trình"], entry["Đơn vị"], entry["Val2"], entry["Val3"], entry["Val4"]])

print(f"Extracted {len(cleaned_data)} potential data rows.")

# --- Step 2: Create Excel File with 'DATABASE' and 'TRA_CUU' ---
wb = openpyxl.Workbook()

# Sheet 1: DATABASE
ws_db = wb.active
ws_db.title = "DATABASE"

# Write Headers
ws_db.append(headers)
for row in cleaned_data:
    ws_db.append(row)

# Styling DB
for cell in ws_db[1]:
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

# Sheet 2: TRA_CUU
ws_ui = wb.create_sheet("TRA_CUU")

# Layout
ws_ui.column_dimensions['A'].width = 2 # Padding
ws_ui.column_dimensions['B'].width = 25 # Label
ws_ui.column_dimensions['C'].width = 50 # Input
ws_ui.column_dimensions['D'].width = 20 # Result Label
ws_ui.column_dimensions['E'].width = 20 # Result Value

# Function Title
title_cell = ws_ui['B2']
title_cell.value = "TRA CỨU SUẤT VỐN ĐẦU TƯ 2024 (QĐ 409/BXD)"
title_cell.font = Font(size=14, bold=True, color="0000FF")
ws_ui.merge_cells('B2:E2')
title_cell.alignment = Alignment(horizontal='center')

# Search Box (Using MATCH logic similar to previous tasks)
# Since the list is huge (hundreds of rows), a simple dropdown is hard.
# We will create a "Search Key" approach or just a raw list for now.
# To make it user friendly, we will list ALL Descriptions in a hidden column and use Data Validation.

# Copy Descriptions to Hidden Column AA for DV
ws_ui['AA1'] = "Danh sách công trình"
for i, row in enumerate(cleaned_data):
    ws_ui[f'AA{i+2}'] = row[1] # Description

# Named Range for List
list_len = len(cleaned_data)
dv_formula = f"=TRA_CUU!$AA$2:$AA${list_len+1}"

# Label
ws_ui['B4'] = "Chọn loại công trình:"
ws_ui['B4'].font = Font(bold=True)
ws_ui['B4'].alignment = Alignment(horizontal='right')

# Input Cell
input_cell = ws_ui['C4']
input_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
dv = DataValidation(type="list", formula1=dv_formula, allow_blank=True)
ws_ui.add_data_validation(dv)
dv.add(input_cell)

# Result Area
ws_ui['B6'] = "Mã hiệu:"
ws_ui['C6'] = f"=IFERROR(VLOOKUP(C4,DATABASE!B:F,1,0), \"\")" # Find Code by Name (Inverse? No, Name is Col 2 in DB... wait VLOOKUP needs Key first)
# Accessing DB columns: A:Code, B:Name, C:Unit, D:Cost, E:Const, F:Equip
# We need to lookup by NAME (Col B). So range is B:F.
# Col B->B is 1. Col B->A is impossible with VLOOKUP. Use INDEX/MATCH.
ws_ui['C6'] = f"=INDEX(DATABASE!A:A,MATCH(C4,DATABASE!B:B,0))"

ws_ui['B7'] = "Đơn vị tính:"
ws_ui['C7'] = f"=VLOOKUP(C4,DATABASE!B:F,2,0)"

ws_ui['B9'] = "KẾT QUẢ:"
ws_ui['B9'].font = Font(bold=True, underline="single")

headers_result = ["Suất vốn đầu tư", "Chi phí Xây dựng", "Chi phí Thiết bị"]
for i, h in enumerate(headers_result):
    r = 10 + i
    ws_ui[f'B{r}'] = h
    # VLOOKUP cols 3, 4, 5
    col_idx = 3 + i
    ws_ui[f'C{r}'] = f"=VLOOKUP($C$4,DATABASE!$B:$F,{col_idx},0)"
    ws_ui[f'C{r}'].number_format = '#,##0.00'
    ws_ui[f'C{r}'].font = Font(bold=True)

# Instructions
ws_ui['B14'] = "Hướng dẫn: Chọn tên công trình từ danh sách. Do danh sách dài, bạn có thể gõ từ khóa vào ô C4 để tìm nhanh."
ws_ui['B14'].font = Font(italic=True, color="555555")

# Hide Database Sheet (Optional, keeping visible for check first)
# ws_db.sheet_state = 'hidden'

print(f"Saving to {output_excel}...")
wb.save(output_excel)
print("Done.")
