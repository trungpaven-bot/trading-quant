
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Paths
input_file = r'g:\My Drive\30_RESOURCES (Kho tai nguyen)\30.01_Tinh hieu qua du an\02_Tra dinh muc theo TT 09 va 12\2026.01.17 Tien do du an CCN PDP 2 - Chi phi HĐ V1.0.xlsx'
output_file = r'g:\My Drive\30_RESOURCES (Kho tai nguyen)\30.01_Tinh hieu qua du an\02_Tra dinh muc theo TT 09 va 12\2026.01.17 Tien do du an CCN PDP 2 - Chi phi HĐ V1.0_Professional.xlsx'

# Load workbook
wb = openpyxl.load_workbook(input_file)
ws = wb['Tien do']

# Define Styles
header_font = Font(bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid") # Dark Blue
sub_header_fill = PatternFill(start_color="95B3D7", end_color="95B3D7", fill_type="solid") # Lighter Blue (Quarters)

# Parent Style (Blue Rows) - Keep similar to input or standardize
# Input Blue: FF5B9BD5 (kind of medium blue)
# We will use a nice standardized light blue for parents
parent_font = Font(bold=True, size=11)
parent_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid") # Very Light Blue (Standardized)
# OR if user matched Blue, maybe they want THAT blue?
# User said "blue items are large items", implying identification.
# I will style them with my "Professional" Parent style to be consistent with previous work.

border_thin = Side(border_style="thin", color="BFBFBF") 
border_dotted = Side(border_style="dotted", color="BFBFBF")
simple_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)

# Currency Format
currency_format = '#,##0'

# Constants for Identification
BLUE_COLOR_PREFIX = 'FF5B9BD5' # approximate check

# --- 1. Fix Headers (Row 4) ---
# Main Header is Row 4
for col in range(1, 100):
    cell = ws.cell(row=4, column=col)
    cell_val = cell.value
    if cell_val:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = simple_border
    
    # Months center
    if col >= 19:
        cell.alignment = Alignment(horizontal='center', vertical='center')

# Years / Quarters (Rows 2, 3 likely, based on previous file logic, although we saw None in Col 1/2)
# We scan row 2 for "NĂM" and row 3 for "QUÝ" or similar
for col in range(1, 100):
    # Row 2 Check (Years)
    cell_r2 = ws.cell(row=2, column=col)
    if cell_r2.value and "NĂM" in str(cell_r2.value):
        # Merge next 11
        ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col+11)
        cell_r2.alignment = Alignment(horizontal='center', vertical='center')
        cell_r2.font = Font(bold=True)
        cell_r2.fill = sub_header_fill
        cell_r2.border = simple_border
    
    # Row 3 Check (Quarters)
    # If explicit quarters are there. Previous file row 2 was quarters? No, row 2 year, row 3 quarter?
    # Actually in previous log: Row 1 Year, Row 2 Device Cost... wait.
    # Previous scan of old file: Row 1 (Index 2) "NĂM 2026". Row 2 (Index 3) "QUÝ 1".
    # Same here likely.

# --- 2. Process Data Rows (Row 5 onwards) ---
max_row = ws.max_row

# We need to know where grouping starts. 
# Row 5 (Yellow) -> Title. Level 0.
# Row 6 (Green) -> Summary. Level 0.
# Row 7 (Blue) -> Parent. Level 0.
# Row 8.. (White) -> Child. Level 1.
# Row X (Blue) -> Parent. Level 0.

current_outline_level = 0

for row_idx in range(5, max_row + 1):
    cell_b = ws.cell(row=row_idx, column=2) # Check Name col for color
    
    # Determine color
    fill_color = None
    if cell_b.fill and cell_b.fill.start_color and cell_b.fill.start_color.index:
        fill_color = cell_b.fill.start_color.index
        # Convert to string just in case
        fill_color = str(fill_color)

    # Logic:
    # If Blue -> Parent. Reset Group.
    # If Yellow/Green -> Special Header (Treat as Parent/Top Level).
    # If White/None -> Child (if inside a group).
    
    is_blue_parent = False
    if fill_color and ('5B9BD5' in fill_color): # Fuzzy match the Blue
        is_blue_parent = True
    
    # Note: Row 5, 6 are unique.
    if row_idx == 5 or row_idx == 6:
        # Keep them as Level 0. Format them.
        # Maybe bold them.
        for col in range(1, ws.max_column + 1):
            c = ws.cell(row=row_idx, column=col)
            c.border = simple_border
            c.font = Font(bold=True) # Bold for Title/Summary
            
            # Formats
            if col >= 19 or col in [8, 9, 17, 18]:
                c.number_format = currency_format
            if col in [4, 5]:
                 c.number_format = 'DD/MM/YYYY'
                 c.alignment = Alignment(horizontal='center', vertical='center')
            if col in [3, 6]:
                 c.alignment = Alignment(horizontal='center', vertical='center')
        
        ws.row_dimensions[row_idx].outlineLevel = 0

    elif is_blue_parent:
        # Parent Row
        current_outline_level = 1 # Next rows will be level 1
        ws.row_dimensions[row_idx].outlineLevel = 0 # Parent is visible
        
        # Style Parent
        for col in range(1, ws.max_column + 1):
            c = ws.cell(row=row_idx, column=col)
            c.font = parent_font
            c.fill = parent_fill # Apply our standardized Parent Blue
            c.border = simple_border
            
            if col >= 19 or col in [8, 9, 17, 18]:
                c.number_format = currency_format
            if col in [4, 5]:
                 c.number_format = 'DD/MM/YYYY'
                 c.alignment = Alignment(horizontal='center', vertical='center')
            if col in [3, 6]:
                 c.alignment = Alignment(horizontal='center', vertical='center')

    else:
        # Child Row
        # Check if we assume everything else is a child
        # Just maintain outline level 1
        ws.row_dimensions[row_idx].outlineLevel = 1
        
        for col in range(1, ws.max_column + 1):
            c = ws.cell(row=row_idx, column=col)
            c.border = simple_border
            # No special fill (white)
            
            if col >= 19 or col in [8, 9, 17, 18]:
                c.number_format = currency_format
            if col in [4, 5]:
                 c.number_format = 'DD/MM/YYYY'
                 c.alignment = Alignment(horizontal='center', vertical='center')
            if col in [3, 6]:
                 c.alignment = Alignment(horizontal='center', vertical='center')

# --- 3. Final Touches ---
# Freeze Panes (Row 5, Col 3 - C)
ws.freeze_panes = "C5"

# Auto-fit columns (Approximate)
for col_idx in range(1, 10):
    col_letter = get_column_letter(col_idx)
    ws.column_dimensions[col_letter].width = 15 
ws.column_dimensions['B'].width = 50 

for col_idx in range(19, ws.max_column + 1):
    col_letter = get_column_letter(col_idx)
    ws.column_dimensions[col_letter].width = 12

wb.save(output_file)
print(f"File saved to: {output_file}")
