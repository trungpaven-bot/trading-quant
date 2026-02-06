
import openpyxl

file_path = r"g:\My Drive\00_INBOX (xu ly nhanh)\tra TMDT.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=False)
ws = wb["TT09"]

print("--- Mapping Item Names and Formulas ---")
# Usually Column B has names like "Chi phí quản lý dự án"
# and Column E, F or H has the Rate/Value.
# Let's scan Column B for text and check the corresponding formulas in Row E, F, G, H

items = {}

for r in range(20, 150): # Assuming main calc table starts after header
    name_cell = ws.cell(row=r, column=2) # Col B
    name = str(name_cell.value).strip() if name_cell.value else ""
    
    if len(name) > 5 and "-" in name: # Likely a cost item e.g. "- Chi phí ..."
        # Check potential formula columns
        formulas = []
        for c in range(3, 10): # Check C to J
            cell = ws.cell(row=r, column=c)
            val = cell.value
            if isinstance(val, str) and val.startswith("="):
                formulas.append(f"Col {openpyxl.utils.get_column_letter(c)}: {val}")
        
        if formulas:
            print(f"Row {r}: {name}")
            for f in formulas:
                print(f"  {f}")

print("\n--- Checking for Control/Driver Cells in Headers ---")
# Look for cells referenced frequently in the formulas above
