
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import re

input_excel = r"g:\My Drive\30_RESOURCES (Kho tai nguyen)\30.01_Tinh hieu qua du an\02_Tra dinh muc theo TT 09 va 12\Du_lieu_Co_Tieu_De_Bang.xlsx"
output_excel = r"g:\My Drive\30_RESOURCES (Kho tai nguyen)\30.01_Tinh hieu qua du an\02_Tra dinh muc theo TT 09 va 12\File_Tra_Cuu_User_Design.xlsx"

print("Reading Data...")
df = pd.read_excel(input_excel, header=None)

processed_data = []
current_header_stack = [] 
current_table_name = ""
last_valid_code = ""

# --- PROCESSING ---
for idx, row in df.iterrows():
    row_strs = [str(x).strip() if not pd.isna(x) else "" for x in row]
    table_name = row_strs[0]
    
    # Filter Tables 1-101
    match = re.search(r'Bảng (\d+)', table_name, re.IGNORECASE)
    if match:
        tbl_num = int(match.group(1))
        if not (1 <= tbl_num <= 101): continue
    else: continue

    if table_name != current_table_name:
        current_table_name = table_name
        current_header_stack = [] 
        last_valid_code = ""
    
    code = row_strs[1]
    desc = row_strs[2]
    vals = row_strs[3:]
    has_number = any(re.search(r'[\d.,]+', v) for v in vals)

    if has_number and desc:
        val1, val2, val3 = 0, 0, 0
        nums = []
        for v in vals:
            try: nums.append(float(v.replace('.', '').replace(',', '.')))
            except: pass
        if len(nums) >= 1: val1 = nums[0]
        if len(nums) >= 2: val2 = nums[1]
        if len(nums) >= 3: val3 = nums[2]
        
        # 3-Level Logic
        t = table_name
        stack = current_header_stack
        
        l1, l2, l3 = "", "", ""
        if len(stack) == 0:
            l1, l2, l3 = desc, "-", "-"
        elif len(stack) == 1:
            l1, l2, l3 = stack[0], desc, "-"
        elif len(stack) >= 2:
            l1, l2, l3 = stack[0], stack[1], desc
            
        if code: last_valid_code = code; final_code = code
        else: final_code = last_valid_code
        
        processed_data.append([tbl_num, t, l1, l2, l3, final_code, val1, val2, val3])
        
    else:
        if desc:
            if code:
                current_header_stack = [desc]
                last_valid_code = code
            else:
                if not current_header_stack: current_header_stack.append(desc)
                else: 
                     if len(current_header_stack) >= 2: current_header_stack[-1] = desc
                     else: current_header_stack.append(desc)

# Sort Natural
processed_data.sort(key=lambda x: (x[0], x[2], x[3], x[4]))

# --- BUILD EXCEL ---
wb = openpyxl.Workbook()

# Function to write data with "Blank Layout"
def write_formatted_sheet(ws, data):
    headers = ["Bảng", "Nhóm CT (Cấp 1)", "Chi tiết (Cấp 2)", "Chi tiết (Cấp 3)", "Mã hiệu", "Suất vốn", "CP Xây dựng", "CP Thiết bị"]
    ws.append(headers)
    
    prev = ["", "", "", ""] # Tbl, L1, L2, L3
    
    for r in data:
        # r: [Num, Tbl, L1, L2, L3, Code, V1, V2, V3]
        curr = [r[1], r[2], r[3], r[4]]
        
        row_out = []
        # Tbl
        if curr[0] != prev[0]: 
            row_out.append(curr[0])
            prev = ["", "", "", ""] # Reset context
        else: row_out.append("")
        
        # L1
        if curr[1] != prev[1]: row_out.append(curr[1])
        else: row_out.append("")
        
        # L2
        if curr[2] != prev[2]: row_out.append(curr[2])
        else: row_out.append("")
        
        # L3 (Always show? Or blank if duplicate? Usually L3 is unique data line, so always show)
        row_out.append(curr[3])
        
        # Code & Vals
        row_out.append(r[5])
        row_out.append(r[6]); row_out.append(r[7]); row_out.append(r[8])
        
        ws.append(row_out)
        prev = [r[1], r[2], r[3], r[4]] # Update state for next row

    # Format
    for cell in ws[1]: cell.font = Font(bold=True)
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 15

# 1. SHEET LISTS
ws_lists = wb.active; ws_lists.title = "LISTS"
write_formatted_sheet(ws_lists, processed_data)

# 2. SHEET ENGINE (Same Custom Layout as requested)
ws_eng = wb.create_sheet("ENGINE")
write_formatted_sheet(ws_eng, processed_data)

# 3. HIDDEN RAW DATA (Backup)
ws_raw = wb.create_sheet("RAW_DATA")
ws_raw.sheet_state = 'hidden'
ws_raw.append(["Num", "Table", "L1", "L2", "L3", "Code", "V1", "V2", "V3"])
for r in processed_data: ws_raw.append(r)

# 4. SHEET TRA_CUU (Empty for user)
ws_ui = wb.create_sheet("TRA_CUU")
ws_ui['B2'] = "TRA CỨU (TỰ THIẾT KẾ)"

print(f"Saving {output_excel}...")
wb.save(output_excel)
print("Done.")
