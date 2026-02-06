
import pandas as pd
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils import quote_sheetname
import re

input_excel = r"g:\My Drive\30_RESOURCES (Kho tai nguyen)\30.01_Tinh hieu qua du an\02_Tra dinh muc theo TT 09 va 12\Du_lieu_Co_Tieu_De_Bang.xlsx"
output_excel = r"g:\My Drive\30_RESOURCES (Kho tai nguyen)\30.01_Tinh hieu qua du an\02_Tra dinh muc theo TT 09 va 12\File_Tra_Cuu_Final_SmartFix_NoBlanks.xlsx"

print("Reading Data...")
df = pd.read_excel(input_excel, header=None)

processed_data = []

current_table = ""
current_base_prefix = "" 
parent_code = ""

start_processing = False

for idx, row in df.iterrows():
    row_strs = [str(x).strip() if not pd.isna(x) else "" for x in row]
    table_name = row_strs[0]
    
    if not start_processing:
        if "Bảng 1" in table_name or "Bang 1" in table_name:
            start_processing = True
        else:
            continue 
            
    code = row_strs[1]
    desc = row_strs[2]
    vals = row_strs[3:]
    has_number = any(re.search(r'[\d.,]+', v) for v in vals)

    if desc and has_number:
        val1, val2, val3 = 0, 0, 0
        nums = []
        for v in vals:
            clean = v.replace('.', '').replace(',', '.')
            try: nums.append(float(clean))
            except: pass
        if len(nums) >= 1: val1 = nums[0]
        if len(nums) >= 2: val2 = nums[1]
        if len(nums) >= 3: val3 = nums[2]
        
        final_desc = desc
        if "không có" in desc.lower():
            parts = re.split(r'không có', desc, flags=re.IGNORECASE)
            if len(parts) > 0: current_base_prefix = parts[0].strip()
            parent_code = code 
        elif desc.lower().startswith("có ") and current_base_prefix:
            final_desc = f"{current_base_prefix} {desc}"
            if not code and parent_code: code = parent_code
        if not code: code = parent_code
        processed_data.append([table_name, code, final_desc, "1000đ/m2", val1, val2, val3])

# EXCEL
wb = openpyxl.Workbook()

# 1. LISTS
ws_lists = wb.active; ws_lists.title="LISTS"
ws_lists['A1']="Bảng"; ws_lists['C1']="Mã"; ws_lists['D1']="Tên CT"
for i, r in enumerate(processed_data):
    row=i+2; ws_lists[f'A{row}']=r[0]; ws_lists[f'C{row}']=r[1]; ws_lists[f'D{row}']=r[2]
    
# Make Unique Tables List in Col Z
ws_lists['Z1']="UnqTbls"
tables = sorted(list(set(r[0] for r in processed_data)))
for i,t in enumerate(tables): ws_lists[f'Z{i+2}']=t

for col in ['A','C','D','Z']: ws_lists.column_dimensions[col].width=20

# 2. DEFINED NAMES (DYNAMIC RANGES to avoid Blanks)
# Formula: OFFSET(Sheet!$Col$2, 0, 0, COUNTA(Sheet!$Col:$Col)-1, 1)
# Note: Use COUNTA-1 to exclude Header
dn_tables = DefinedName("ListTables", attr_text="OFFSET(LISTS!$Z$2,0,0,COUNTA(LISTS!$Z:$Z)-1,1)")
wb.defined_names.add(dn_tables)

dn_col_tbl = DefinedName("ColTbl", attr_text="OFFSET(LISTS!$A$2,0,0,COUNTA(LISTS!$A:$A)-1,1)")
wb.defined_names.add(dn_col_tbl)

dn_col_desc = DefinedName("ColDesc", attr_text="OFFSET(LISTS!$D$2,0,0,COUNTA(LISTS!$D:$D)-1,1)") # Col D desc
wb.defined_names.add(dn_col_desc)

# 3. ENGINE
ws_eng = wb.create_sheet("ENGINE")
ws_eng.append(["", "Bảng", "Bảng2", "Mã", "Tên", "ĐV", "Suất", "XD", "TB", "Key"])
for i, r in enumerate(processed_data):
    row=i+2
    ws_eng.cell(row,2,r[0]); ws_eng.cell(row,3,r[0]); ws_eng.cell(row,4,r[1]); ws_eng.cell(row,5,r[2])
    ws_eng.cell(row,6,r[3]); ws_eng.cell(row,7,r[4]); ws_eng.cell(row,8,r[5]); ws_eng.cell(row,9,r[6])
    ws_eng.cell(row,10).value = f'=B{row} & "_" & E{row}'

# 4. TRA_CUU
ws_ui = wb.create_sheet("TRA_CUU")
ws_ui['B2']="TRA CỨU SUẤT VỐN (NO BLANKS)"; ws_ui['B2'].font=Font(size=14, bold=True)
ws_ui['B4']="1. Chọn Bảng:"; ws_ui['C4'].border=Border(bottom=Side(style='thin'))

# Validation 1 (Dynamic ListTables)
ws_ui.add_data_validation(DataValidation(type="list", formula1="=ListTables", allow_blank=True))
ws_ui['C4'].parent.data_validations.dataValidation[-1].add(ws_ui['C4'])

# Validation 2 (Dependent)
ws_ui['B5']="2. Chọn CT:"; ws_ui['C5'].border=Border(bottom=Side(style='thin')); ws_ui['C5'].alignment=Alignment(wrap_text=True)
dv2_f = "=OFFSET(LISTS!$D$2;MATCH($C$4;ColTbl;0)-1;0;COUNTIF(ColTbl;$C$4);1)"
ws_ui.add_data_validation(DataValidation(type="list", formula1=dv2_f, allow_blank=True))
ws_ui['C5'].parent.data_validations.dataValidation[-1].add(ws_ui['C5'])

# Results
ws_ui['B8']="KẾT QUẢ:"; labels=["Mã","ĐV","Suất","XD","TB"]; cols=["D","F","G","H","I"]
for i,l in enumerate(labels):
    r=9+i; ws_ui[f'B{r}']=l; ws_ui[f'B{r}'].font=Font(bold=True)
    f=f"=INDEX(ENGINE!{cols[i]}:{cols[i]}; MATCH($C$4 & \"_\" & $C$5; ENGINE!$J:$J; 0))"
    if i>=2: ws_ui[f'C{r}'].number_format='#,##0'
    ws_ui[f'C{r}']=f

print(f"Saving {output_excel}...")
wb.save(output_excel)
print("Done.")
