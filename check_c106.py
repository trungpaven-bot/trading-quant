
import openpyxl

file_path = r"g:\My Drive\00_INBOX (xu ly nhanh)\tra TMDT.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)
ws = wb["TT09"]

print("--- Value at C106 ---")
print(ws["C106"].value)
print("--- Value at C107 ---")
print(ws["C107"].value)

print("\n--- Inspecting Row 106 more closely ---")
for c in range(1, 10):
    print(f"({106},{c}): {ws.cell(row=106, column=c).value}")
