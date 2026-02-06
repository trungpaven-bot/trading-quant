
import openpyxl

file_path = r"g:\My Drive\00_INBOX (xu ly nhanh)\tra TMDT.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)
ws = wb["TT09"]

print("--- Checking named ranges starting with 'QKB' or 'WVM' from previous dump ---")
# Earlier dump showed weird named ranges like QKB483 that pointed to list ranges.
# Let's see if we can identify the cell that holds the INDEX for the dropdowns.
# Usually linked cells are hidden or near the dropdown. The dropdowns are in Rows 4-6.
# Let's check cells in Row 4-6, Columns F, I, L, O for integer values 1,2,3.

for r in range(4, 7):
    for c in range(5, 20): # Check columns E to T
        cell = ws.cell(row=r, column=c)
        val = cell.value
        # If integer, it might be the linked cell
        if isinstance(val, int) or (isinstance(val, float) and val.is_integer()):
            print(f"Possible Linked Cell at {cell.coordinate}: {val}")

print("\n--- Checking formulas in hidden rows 48+ that refer to these possible linked cells ---")
# We saw formulas like =HLOOKUP(..., $F$58+1, )
# F58 was 8. It might be related to "Loại công trình" index.

print(f"Value at F58: {ws['F58'].value}")
print(f"Structure at 48+:")
# Read headers at row 48
header_vals = []
for c in range(7, 15):
    header_vals.append(str(ws.cell(row=48, column=c).value))
print("Row 48 (G-N):", header_vals)

# Read Row 1 of the table (Row 49?)
row1_vals = []
for c in range(7, 15):
    row1_vals.append(str(ws.cell(row=49, column=c).value))
print("Row 49 (G-N):", row1_vals)
