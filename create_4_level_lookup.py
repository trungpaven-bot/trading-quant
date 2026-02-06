
import pandas as pd
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.workbook.defined_name import DefinedName
import re

input_excel = r"g:\My Drive\30_RESOURCES (Kho tai nguyen)\30.01_Tinh hieu qua du an\02_Tra dinh muc theo TT 09 va 12\Du_lieu_Co_Tieu_De_Bang.xlsx"
output_excel = r"g:\My Drive\30_RESOURCES (Kho tai nguyen)\30.01_Tinh hieu qua du an\02_Tra dinh muc theo TT 09 va 12\File_Tra_Cuu_4_Cap.xlsx"

print("Reading Data...")
df = pd.read_excel(input_excel, header=None)

# 1. PARSING TO 3 CONTEXT LEVELS (L1, L2, L3)
processed_data = []
current_header_stack = [] 
current_table_name = ""
last_valid_code = ""
start_processing = False

for idx, row in df.iterrows():
    row_strs = [str(x).strip() if not pd.isna(x) else "" for x in row]
    table_name = row_strs[0]
    
    if table_name != current_table_name:
        current_table_name = table_name
        current_header_stack = [] 
        last_valid_code = ""
    
    if not start_processing:
        if "Bảng 1" in table_name or "Bang 1" in table_name:
            start_processing = True
        else:
            continue
            
    code = row_strs[1]
    desc = row_strs[2]
    vals = row_strs[3:]
    has_number = any(re.search(r'[\d.,]+', v) for v in vals)

    if has_number and desc:
        # DATA ROW
        val1, val2, val3 = 0, 0, 0
        nums = []
        for v in vals:
            try: nums.append(float(v.replace('.', '').replace(',', '.')))
            except: pass
        if len(nums) >= 1: val1 = nums[0]
        if len(nums) >= 2: val2 = nums[1]
        if len(nums) >= 3: val3 = nums[2]
        
        # Flatten Hierarchy
        # Full chain = Stack + [Desc]
        chain = current_header_stack + [desc]
        
        # Normalize to 3 slots: L1, L2, L3
        l1, l2, l3 = "", "-", "-"
        
        if len(chain) == 1:
            l1 = chain[0]
        elif len(chain) == 2:
            l1 = chain[0]
            l2 = chain[1]
        elif len(chain) == 3:
            l1 = chain[0]
            l2 = chain[1]
            l3 = chain[2]
        else:
            # > 3 items? Merge the extras into the last one
            l1 = chain[0]
            l2 = chain[1]
            l3 = " - ".join(chain[2:])
            
        if code: last_valid_code = code; final_code = code
        else: final_code = last_valid_code
        
        processed_data.append([table_name, final_code, l1, l2, l3, "1000đ/Dv", val1, val2, val3])
        
    else:
        # HEADER ROW
        if desc:
            if code:
                current_header_stack = [desc]
                last_valid_code = code
            else:
                if len(current_header_stack) == 0:
                    current_header_stack.append(desc)
                else:
                    # Append or Replace logic
                    # To support deep nesting, we Append.
                    # Use a max depth check?
                    current_header_stack.append(desc)
                    # Note: Previous "Replace" logic might have been too aggressive.
                    # With 4 levels, "Append" is safer for "Drill Down".

print(f"Parsed {len(processed_data)} rows.")

# 2. GENERATE AUXILIARY MENU SHEETS
# We need menus for intermediate dropdowns to avoid duplicates.

# SORT KEY: Table, L1, L2, L3
processed_data.sort(key=lambda x: (x[0], x[2], x[3], x[4]))

# Unique Tables
tables = []
seen = set()
for r in processed_data:
    if r[0] not in seen:
        seen.add(r[0]); tables.append(r[0])

# Unique L1 (Key: Table) -> For C5
menu_l1 = []
seen = set()
for r in processed_data:
    k = f"{r[0]}_{r[2]}"
    if k not in seen:
        seen.add(k); menu_l1.append([r[0], r[2], k])

# Unique L2 (Key: Table_L1) -> For C6
menu_l2 = []
seen = set()
for r in processed_data:
    k = f"{r[0]}_{r[2]}_{r[3]}" # Full uniqueness
    lookup_k = f"{r[0]}_{r[2]}" # Key for filtering
    if k not in seen:
        seen.add(k); menu_l2.append([r[0], r[2], r[3], lookup_k])


# === EXCEL CONSTRUCTION ===
wb = openpyxl.Workbook()

# SHEET: MENU (Stores all helper lists)
ws_menu = wb.active; ws_menu.title = "MENU"
ws_menu.append(["Table", "L1_Tbl", "L1_Val", "L1_Key", "L2_TblL1", "L2_Val"])

# Col A: Tables
for i, t in enumerate(tables): ws_menu[f'A{i+2}'] = t

# Col C,D,E: Menu L1
for i, item in enumerate(menu_l1):
    ws_menu[f'C{i+2}'] = item[0] # Tbl
    ws_menu[f'D{i+2}'] = item[1] # Val
    ws_menu[f'E{i+2}'] = item[2] # Key

# Col G,H: Menu L2
for i, item in enumerate(menu_l2):
    ws_menu[f'G{i+2}'] = item[3] # Key (Tbl_L1)
    ws_menu[f'H{i+2}'] = item[2] # Val (L2)

# SHEET: ENGINE (Stores Data)
ws_eng = wb.create_sheet("ENGINE")
headers = ["Table", "Code", "L1", "L2", "L3", "Unit", "V1", "V2", "V3", "Key_L3", "SuperKey"]
ws_eng.append(headers)
for i, r in enumerate(processed_data):
    # r = [Tbl, Code, L1, L2, L3, Unit, V1, V2, V3]
    # Key for L3 Lookup (Dropdown C7): Tbl_L1_L2
    k_l3 = f"{r[0]}_{r[2]}_{r[3]}"
    # SuperKey for Result: Tbl_L1_L2_L3
    super_k = f"{r[0]}_{r[2]}_{r[3]}_{r[4]}"
    
    ws_eng.append([r[0], r[1], r[2], r[3], r[4], r[5], r[6], r[7], r[8], k_l3, super_k])


# === NAMED RANGES ===
# Helper function to add name
def add_name(name, sheet, col, count):
    wb.defined_names.add(DefinedName(name, attr_text=f"{sheet}!${col}$2:${col}${count+1}"))

add_name("ListTables", "MENU", "A", len(tables))

add_name("M1_Key", "MENU", "C", len(menu_l1)) # Tbl column
add_name("M1_Val", "MENU", "D", len(menu_l1))

add_name("M2_Key", "MENU", "G", len(menu_l2)) # Tbl_L1 column
add_name("M2_Val", "MENU", "H", len(menu_l2))

add_name("Eng_Key", "ENGINE", "J", len(processed_data)) # Tbl_L1_L2 column
add_name("Eng_Val", "ENGINE", "E", len(processed_data)) # L3 column
add_name("Eng_Super", "ENGINE", "K", len(processed_data)) # Full Key


# === TRA_CUU UI ===
ws_ui = wb.create_sheet("TRA_CUU")
ws_ui['B2']="TRA CỨU 4 CẤP (FINAL)"; ws_ui['B2'].font=Font(size=14, bold=True)

# C4: Table
ws_ui['B4']="1. Bảng:"; ws_ui['C4'].border=Border(bottom=Side(style='thin'))
dv1 = DataValidation(type="list", formula1="=ListTables", allow_blank=True); ws_ui.add_data_validation(dv1); dv1.add(ws_ui['C4'])

# C5: L1 (Group) -> Filter Menu1 by C4
ws_ui['B5']="2. Nhóm:"; ws_ui['C5'].border=Border(bottom=Side(style='thin')); ws_ui['C5'].alignment=Alignment(wrap_text=True)
dv2_f = "=OFFSET(M1_Val;MATCH($C$4;M1_Key;0)-1;0;COUNTIF(M1_Key;$C$4);1)"
dv2 = DataValidation(type="list", formula1=dv2_f, allow_blank=True); ws_ui.add_data_validation(dv2); dv2.add(ws_ui['C5'])

# C6: L2 (Detail 1) -> Filter Menu2 by C4_C5
ws_ui['B6']="3. Loại 1:"; ws_ui['C6'].border=Border(bottom=Side(style='thin')); ws_ui['C6'].alignment=Alignment(wrap_text=True)
dv3_f = "=OFFSET(M2_Val;MATCH($C$4&\"_\"&$C$5;M2_Key;0)-1;0;COUNTIF(M2_Key;$C$4&\"_\"&$C$5);1)"
dv3 = DataValidation(type="list", formula1=dv3_f, allow_blank=True); ws_ui.add_data_validation(dv3); dv3.add(ws_ui['C6'])

# C7: L3 (Detail 2) -> Filter Engine by C4_C5_C6
ws_ui['B7']="4. Loại 2:"; ws_ui['C7'].border=Border(bottom=Side(style='thin')); ws_ui['C7'].alignment=Alignment(wrap_text=True)
dv4_f = "=OFFSET(Eng_Val;MATCH($C$4&\"_\"&$C$5&\"_\"&$C$6;Eng_Key;0)-1;0;COUNTIF(Eng_Key;$C$4&\"_\"&$C$5&\"_\"&$C$6);1)"
dv4 = DataValidation(type="list", formula1=dv4_f, allow_blank=True); ws_ui.add_data_validation(dv4); dv4.add(ws_ui['C7'])


# Results
ws_ui['B9']="KẾT QUẢ:"; labels=["Mã","ĐV","Suất","XD","TB"]; cols=["B","F","G","H","I"] # Engine Cols
for i,l in enumerate(labels):
    r=10+i; ws_ui[f'B{r}']=l; ws_ui[f'B{r}'].font=Font(bold=True)
    # Match SuperKey: C4_C5_C6_C7
    f=f"=INDEX(ENGINE!{cols[i]}:{cols[i]}; MATCH($C$4 & \"_\" & $C$5 & \"_\" & $C$6 & \"_\" & $C$7; Eng_Super; 0))"
    if i>=2: ws_ui[f'C{r}'].number_format='#,##0'
    ws_ui[f'C{r}']=f

ws_menu.sheet_state = 'hidden'
ws_eng.sheet_state = 'hidden'

print(f"Saving {output_excel}...")
wb.save(output_excel)
print("Done.")
