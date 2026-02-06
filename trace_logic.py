
import openpyxl

file_path = r"g:\My Drive\00_INBOX (xu ly nhanh)\tra TMDT.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=False) # Get formulas
ws = wb["TT09"]

print("--- Checking Formula at F58 ---")
# F58 seems to be the row index for HLOOKUP
print(f"F58: {ws['F58'].value}")
# If F58 is a formula, it will tell us what drives the Row Index.
# My guess: =MATCH(LinkedCell_LoaiCongTrinh, Row_Titles, 0)

print("\n--- Checking Formulas for Cost Items ---")
# Let's check some cost items around row 60-100 to see if they multiply by any flags.
# Example: "Chi phí thẩm tra" might be = ... * IF(LinkedCell_ThamDinh=1, 1, 0)
# We need to find cells that refer to the linked cells of the dropdowns.

# Since we don't know the exact linked cells yet (they might be hidden), 
# let's look for IFs in the main column (I or J usually has the final calc).

for r in range(60, 150):
    for c in range(1, 15):
        cell = ws.cell(row=r, column=c)
        val = str(cell.value)
        if "IF(" in val and ("=1" in val or "=2" in val):
            print(f"Row {r} Form: {val}")
