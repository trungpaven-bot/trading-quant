
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation

file_path = r"g:\My Drive\00_INBOX (xu ly nhanh)\tra TMDT.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=False)
ws = wb["TT09"]

# Hardcoded Manual Mapping based on visual inspection or educated guess if labels are images/shapes (unlikely but possible if text scan failed)
# Actually, the user's screenshot clearly shows "Loại công trình:", "Cấp công trình:" in the header area.
# But my text scan didn't find them in the cells!
# This implies they might be:
# 1. Inside Text Boxes / Shapes (floating), not in cells.
# 2. In hidden rows enabled by grouping? No, dump included hidden content usually... wait, scan was data_only=True.
# 3. The screenshot is from a different view or the file I'm reading is not the one they are looking at? 
# The user said "xem file excel kia", implying the one I have.
#
# Let's assume the dropdowns ARE the cells that have the variables.
# From the formula trace:
# i_CPC -> F58. "CPC" might be "Chi phí Chung"? No.
# i_CapCT -> used in CHOOSE.
#
# Let's Look at the NAMED RANGES again to find the LINKED CELLS.
# i_CapCT -> [('[2]Config', '$C$88')] -> This points to Sheet '[2]Config' which identifies an external ref or a sheet named "Config"?
# But `wb.sheetnames` didn't show "Config". 
# Ah! [2]Config means valid for the 2nd sheet? Or Sheet Index 2?
#
# Let's check `wb.sheetnames` again.
# ['CHI_TIEU', 'TT09', 'TMDT', ...]
# 'TMDT' is index 2 (0-based: 0, 1, 2). So maybe it is TMDT!C88?
# I checked TMDT!C88 before and it was None.

# PLAN B: Create a new sheet "DATA_LISTS" and just tell the user to link them manually if I can't find the target cells.
# BUT the user asked for a script to "tự động tạo lại".
# I must find where to put the validation.
#
# Let's look for "Combo Box" objects (drawing objects).
# Openpyxl doesn't support reading Form Control Combo Boxes well.
# If they are Form Controls, they float over cells. 
#
# If they are Data Validation dropdowns, they are attached to cells.
# My `inspect_dropdowns.py` found:
# Validation range: E483 E499 ... Formula: $D$480:$D$482.
# That's way down.
#
# Maybe I should just blindly APPLY validation to the cell ranges that "Look like" the header locations.
# 
# Let's try to infer from the structure:
# Row 2,3,4 are data.
# Row 2 Cols F-I? 
# In the screenshot:
# "GXD" is at B2. "GTB" at B3. "TMDT" at B4.
# "Loại công trình" is to the right of GXD (Row 2).
# So "Loại công trình" label is likely at E2 or F2. Dropdown at G2/H2.
# "Kiểu công trình" is further right.
# "Thẩm định..." is below "Loại công trình". Row 3?
# "Kiểm toán..." is right of Thẩm định.
# "Chi phí dự phòng" is at E5. So "Cấp công trình" is to the right of it.
# Screenshot shows "Cấp công trình" same row as "Dự phòng". So Row 5.
# "Số bước thiết kế" right of "Cấp công trình". 

# Hypothesized Locations:
# Row 2: "Loại CT" (Label ~E2, Dropdown ~G2?) | "Kiểu CT" (Label ~J2, Dropdown ~L2?)
# Row 3: "Thẩm định" (Label ~E3, Dropdown ~G3?) | "Kiểm toán" (Label ~J3, Dropdown ~L3?)
# Row 5: "Dự phòng" (E5), "Cấp CT" (Label ~H5?, Dropdown ~J5?) | "Số bước" (Label ~K5?, Dropdown ~M5?)
#
# I will write the lists to hidden area, then Update the file.
# Then I will print instructions telling the user exactly which cells I *think* they are, 
# and how to link them if they are Form Controls.
#
# BUT if they are Form Controls (which they look like in the screenshot - grey 3D styling),
# I cannot edit their "Input Range" via openpyxl. Openpyxl only handles Data Validation (in-cell dropdowns).
# If the user says "Chuột phải... Format Control", they ARE Form Controls (Objects).
#
# Openpyxl CANNOT modify existing Form Controls (Buttons/Combos).
#
# ACTION: I should create the Source Lists (Data) in the sheet so the user exists.
# Then I tell the user: "I have created the reference lists at [Address]. Now please right click the controls and set Input Range to these addresses."

# Writing script to GENERATE THE LISTS.

start_col_idx = 27 # AA
start_row = 1000
source_map = {}

for i, (key, values) in enumerate(data_lists.items()):
    col_idx = start_col_idx + i
    col_letter = openpyxl.utils.get_column_letter(col_idx)
    
    # Write Header
    ws.cell(row=start_row, column=col_idx, value=key)
    
    # Write Values
    for j, val in enumerate(values):
        ws.cell(row=start_row + 1 + j, column=col_idx, value=val)
        
    end_row = start_row + len(values)
    range_ref = f"${col_letter}${start_row + 1}:${col_letter}${end_row}"
    source_map[key] = f"TT09!{range_ref}"

wb.save(file_path)
print("Lists created.")
for k, v in source_map.items():
    print(f"{k}: {v}")
