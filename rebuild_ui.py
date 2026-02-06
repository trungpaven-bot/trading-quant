
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Border, Side, Alignment

file_path = r"g:\My Drive\30_RESOURCES (Kho tai nguyen)\30.01_Tinh hieu qua du an\02_Tra dinh muc theo TT 09 va 12\Tra dinh muc TT09.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=False)
ws = wb["TT09"]

# 1. Setup UI Layout (Labels & Dropdown Cells in the Empty Red Box area)
# Area: Rows 2-5, Cols G-M.

# Map: Label Cell -> Input Cell -> Variable Name -> List Range
ui_map = [
    # LabelText, LabelPos, InputPos, VarName, ListRange
    ("Loại công trình:", "G2", "H2", "i_CPC", "$AA$1001:$AA$1005"),
    ("Kiểu công trình:", "K2", "L2", "i_Kieu", "$AB$1001:$AB$1003"), # i_Kieu is dummy name if not exists
    ("Thẩm định TK, DT:", "G3", "H3", "i_tkdutoan", "$AC$1001:$AC$1002"),
    ("Kiểm toán/QT:", "K3", "L3", "i_KT", "$AD$1001:$AD$1002"),
    ("Cấp công trình:", "G5", "H5", "i_CapCT", "$AE$1001:$AE$1005"), 
    ("Số bước thiết kế:", "K5", "L5", "i_BTK", "$AF$1001:$AF$1003"),
]

# Helper Columns for Index Calculation (where we convert Text -> 1,2,3)
# We will use Column AG (33) for storing the calculated INDEX.
# Start Row for Helpers: 1000
helper_col_idx = 33 # AG
helper_col_letter = "AG"

print("--- Rebuilding Interface with Data Validation ---")

for i, (label_text, label_cell, input_cell, var_name, list_range_addr) in enumerate(ui_map):
    # 1. Write Label
    ws[label_cell].value = label_text
    ws[label_cell].alignment = Alignment(horizontal='right', vertical='center')
    
    # 2. Setup Input Cell (Data Validation)
    dv = DataValidation(type="list", formula1=f"='TT09'!{list_range_addr}", allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(ws[input_cell])
    
    # Set default value to avoid breaking formulas initially
    # Get the first item from the list range to populate
    # We cheat a bit: reading the list range from the known address
    # list_range_addr is like $AA$1001:$AA$1005. 
    # Let's just pick the top cell of that range.
    list_start_cell = list_range_addr.split(":")[0].replace("$", "")
    ws[input_cell].value = ws[list_start_cell].value
    
    # Style the Input Cell
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    ws[input_cell].border = thin_border
    
    # 3. Create Backend Formula for Index (MATCH)
    # Formula: =MATCH(InputCell, ListRange, 0)
    # We place this in AG1001, AG1002...
    index_cell_row = 1001 + i
    index_cell_addr = f"${helper_col_letter}${index_cell_row}" # e.g. $AG$1001
    
    # Formula needs to be written to the cell
    match_formula = f"=MATCH({input_cell},'TT09'!{list_range_addr},0)"
    ws[index_cell_addr] = match_formula
    print(f"Created Logic: {index_cell_addr} = {match_formula}")
    
    # 4. Update Named Range to point to this Index Cell
    # Only if it's one of the known variables
    if var_name in ["i_CPC", "i_CapCT", "i_BTK", "i_tkdutoan", "i_VB"]:
        # Delete old
        if var_name in wb.defined_names:
            del wb.defined_names[var_name]
        
        # Create new pointing to TT09!IndexCell
        new_dest = f"'TT09'!{index_cell_addr}"
        d = openpyxl.workbook.defined_name.DefinedName(var_name, attr_text=new_dest)
        wb.defined_names[var_name] = d
        print(f"Remapped Variable {var_name} -> {new_dest}")
    else:
        # Check if we should create a new name or just leave it calculate
        # For "Kiểm toán" and "Kiểu công trình", we might not have a variable name yet.
        # But having the index in AG column is useful if the user needs to link it later.
        pass

print("Saving file...")
wb.save(file_path)
print("Done. Interface restored.")
