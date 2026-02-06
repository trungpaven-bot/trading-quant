
import pandas as pd
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.workbook.defined_name import DefinedName
import re

input_excel = r"g:\My Drive\30_RESOURCES (Kho tai nguyen)\30.01_Tinh hieu qua du an\02_Tra dinh muc theo TT 09 va 12\Du_lieu_Co_Tieu_De_Bang.xlsx"
output_excel = r"g:\My Drive\30_RESOURCES (Kho tai nguyen)\30.01_Tinh hieu qua du an\02_Tra dinh muc theo TT 09 va 12\File_Tra_Cuu_Final_Compact.xlsx"

print("Reading Data...")
df = pd.read_excel(input_excel, header=None)

# --- PROCESSING DATA ---
processed_data = []
current_base_prefix = "" 
parent_code = ""
start_processing = False

# Filter out None/Blank rows immediately to ensure compactness
clean_rows = []
for idx, row in df.iterrows():
    # Helper to check if row has ANY content
    if not any(pd.notna(x) and str(x).strip() != "" for x in row):
        continue
    clean_rows.append(row)

for row in clean_rows:
    row_strs = [str(x).strip() if not pd.isna(x) else "" for x in row]
    table_name = row_strs[0]
    
    if not start_processing:
        if "Bảng 1" in table_name or "Bang 1" in table_name:
            start_processing = True
        else:
            continue 
            
    code = row_strs[1]
    desc = row_strs[2]
    # Ensure Table/Desc text exists
    if not table_name: continue

    vals = row_strs[3:]
    has_number = any(re.search(r'[\d.,]+', v) for v in vals)

    if desc and has_number:
        val1, val2, val3 = 0, 0, 0
        nums = []
        for v in vals:
            clean = v.replace('.', '').replace(',', '.')
            try: nums.append(float(clean))
            except: pass
        if len(nums) >= 1: val1 = nums[0]
        if len(nums) >= 2: val2 = nums[1]
        if len(nums) >= 3: val3 = nums[2]
        
        final_desc = desc
        if "không có" in desc.lower():
            parts = re.split(r'không có', desc, flags=re.IGNORECASE)
            if len(parts) > 0: current_base_prefix = parts[0].strip()
            parent_code = code 
        elif desc.lower().startswith("có ") and current_base_prefix:
            final_desc = f"{current_base_prefix} {desc}"
            if not code and parent_code: code = parent_code
        if not code: code = parent_code
        
        processed_data.append([table_name, code, final_desc, "1000đ/m2", val1, val2, val3])

# --- EXCEL ---
wb = openpyxl.Workbook()

# 1. LISTS SHEET
ws_lists = wb.active; ws_lists.title="LISTS"
ws_lists['A1']="Bảng"; ws_lists['C1']="Mã"; ws_lists['D1']="Tên CT"

# Write data continuously (NO GAPS guaranteed by loop)
for i, r in enumerate(processed_data):
    row=i+2
    ws_lists[f'A{row}'] = r[0] 
    ws_lists[f'C{row}'] = r[1]
    ws_lists[f'D{row}'] = r[2]

data_count = len(processed_data)

# Create Unique List for Tables in Col Z (COMPACTED)
ws_lists['Z1'] = "UniqueTables_Compacted"
unique_tables = []
seen = set()
for r in processed_data:
    t = r[0]
    if t and t not in seen:
        seen.add(t)
        unique_tables.append(t)
        
for i, t in enumerate(unique_tables):
    ws_lists[f'Z{i+2}'] = t

unique_count = len(unique_tables)

# DEFINED NAMES with EXACT RANGES (No Offset needed because we know the exact count)
# ListTables -> Points to Compact Z column
dn1 = DefinedName("ListTables", attr_text=f"LISTS!$Z$2:$Z${unique_count+1}")
wb.defined_names.add(dn1)

# Helper Columns for Dependent List -> Points to Compact A and D columns
dn2 = DefinedName("ColTbl", attr_text=f"LISTS!$A$2:$A${data_count+1}")
wb.defined_names.add(dn2)
dn3 = DefinedName("ColDesc", attr_text=f"LISTS!$D$2:$D${data_count+1}")
wb.defined_names.add(dn3)


# 2. ENGINE SHEET
ws_eng = wb.create_sheet("ENGINE")
ws_eng.append(["", "Bảng", "Bảng2", "Mã", "Tên", "ĐV", "Suất", "XD", "TB", "Key"])
for i, r in enumerate(processed_data):
    row=i+2
    # Write engine data contiguously
    ws_eng.cell(row,2,r[0]); ws_eng.cell(row,3,r[0]); ws_eng.cell(row,4,r[1]); ws_eng.cell(row,5,r[2])
    ws_eng.cell(row,6,r[3]); ws_eng.cell(row,7,r[4]); ws_eng.cell(row,8,r[5]); ws_eng.cell(row,9,r[6])
    ws_eng.cell(row,10).value = f'=B{row} & "_" & E{row}'

# 3. TRA_CUU SHEET (Validation Setup)
ws_ui = wb.create_sheet("TRA_CUU")
ws_ui['B2']="TRA CỨU SUẤT VỐN (COMPACT)"; ws_ui['B2'].font=Font(size=14, bold=True)
ws_ui['B4']="1. Chọn Bảng:"; ws_ui['C4'].border=Border(bottom=Side(style='thin'))

# Validation 1: Points to ListTables (exact range)
dv1 = DataValidation(type="list", formula1="=ListTables", allow_blank=True) # allow_blank=True means user can delete cell content, not that list has blanks
ws_ui.add_data_validation(dv1)
dv1.add(ws_ui['C4'])

# Validation 2: Dependent
ws_ui['B5']="2. Chọn CT:"; ws_ui['C5'].border=Border(bottom=Side(style='thin')); ws_ui['C5'].alignment=Alignment(wrap_text=True)
dv2_f = "=OFFSET(LISTS!$D$2;MATCH($C$4;ColTbl;0)-1;0;COUNTIF(ColTbl;$C$4);1)"
dv2 = DataValidation(type="list", formula1=dv2_f, allow_blank=True)
ws_ui.add_data_validation(dv2)
dv2.add(ws_ui['C5'])

# Results
ws_ui['B8']="KẾT QUẢ:"; labels=["Mã","ĐV","Suất","XD","TB"]; cols=["D","F","G","H","I"]
for i,l in enumerate(labels):
    r=9+i; ws_ui[f'B{r}']=l; ws_ui[f'B{r}'].font=Font(bold=True)
    f=f"=INDEX(ENGINE!{cols[i]}:{cols[i]}; MATCH($C$4 & \"_\" & $C$5; ENGINE!$J:$J; 0))"
    if i>=2: ws_ui[f'C{r}'].number_format='#,##0'
    ws_ui[f'C{r}']=f

print(f"Saving {output_excel}...")
wb.save(output_excel)
print("Done.")
