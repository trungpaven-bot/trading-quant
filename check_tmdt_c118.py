
import openpyxl

file_path = r"g:\My Drive\00_INBOX (xu ly nhanh)\tra TMDT.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)
ws = wb["TMDT"]

print("--- Data in TMDT!C118 ---")
print(f"Value: {ws['C118'].value}")

print("\n--- Data in TMDT Rows 115-125, Cols A-E ---")
for r in range(115, 126):
    vals = []
    for c in range(1, 6):
        vals.append(str(ws.cell(row=r, column=c).value))
    print(f"Row {r}: {vals}")
