
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation

file_path = r"g:\My Drive\00_INBOX (xu ly nhanh)\tra TMDT.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=False)
ws = wb["TT09"]

# 1. Define the Data Lists
data_lists = {
    "LoaiCongTrinh": ["Công trình dân dụng", "Công trình công nghiệp", "Công trình giao thông", "Công trình NN&PTNT", "Công trình hạ tầng kỹ thuật"],
    "KieuCongTrinh": ["Công trình xây dựng mới", "Công trình thiết kế mẫu, điển hình", "Công trình sửa chữa, cải tạo, mở rộng"],
    "ThamDinh": ["Cơ quan chuyên môn về xây dựng thẩm định", "Chủ đầu tư tổ chức thẩm định"],
    "KiemToan": ["Thuê kiểm toán độc lập", "Quyết toán nội bộ (Không thuê)"],
    "CapCongTrinh": ["Cấp đặc biệt", "Cấp I", "Cấp II", "Cấp III", "Cấp IV"],
    "SoBuocThietKe": ["Thiết kế 1 bước (BCKTKT)", "Thiết kế 2 bước (Cơ sở + BVTC)", "Thiết kế 3 bước (Cơ sở + KT + BVTC)"]
}

# 2. Write these lists to a safe place in the sheet (e.g., hidden Columns AA-AF, starting row 1000)
# We avoid overwriting existing data.
start_col_idx = 27 # Column AA
start_row = 1000

list_ranges = {} # Store "Sheet!$A$1:$A$5" for validation

print("--- Writing Data Lists to Hidden Area (Row 1000+) ---")
for i, (key, values) in enumerate(data_lists.items()):
    col_idx = start_col_idx + i
    col_letter = openpyxl.utils.get_column_letter(col_idx)
    
    # Write Header
    ws.cell(row=start_row, column=col_idx, value=key.upper())
    
    # Write Values
    for j, val in enumerate(values):
        ws.cell(row=start_row + 1 + j, column=col_idx, value=val)
    
    # Define Range String
    end_row = start_row + len(values)
    range_str = f"'{ws.title}'!${col_letter}${start_row + 1}:${col_letter}${end_row}"
    list_ranges[key] = range_str
    print(f"  Written {key} to {col_letter}{start_row+1}:{col_letter}{end_row}")

# 3. Apply Data Validation (Input Source repair)
# We need to map the Visual Input Cells to the List Keys.
# Based on the user's screenshot and previous dumps (Labels in Rows 4-6):
# Row 4: Loại công trình (G4?), Kiểu CT (K4?)
# Row 5: Thẩm định (E5/F5?), Kiểm toán (J5/K5?) ... wait, let's verify positions.
# My previous "dump_wide" showed:
# (5,2): '- Tỷ lệ làm tròn...' | (5,4): 3 | (5,5): '- Chi phí dự phòng' ...
# Wait, the screenshot has labels like:
# "Loại công trình:" (Row ?) -> Dropdown
# "Thẩm định thiết kế...": (Row ?) -> Dropdown
# "Kiểm toán/QT": (Row ?) -> Dropdown
# "Cấp công trình": (Row ?) -> Dropdown
# "Số bước thiết kế": (Row ?) -> Dropdown

# I need to find the EXACT cell addresses for these dropdowns first.
# Using a quick search again for the label locations to be precise.
