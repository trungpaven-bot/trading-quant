
import openpyxl

file_path = r"g:\My Drive\30_RESOURCES (Kho tai nguyen)\30.01_Tinh hieu qua du an\02_Tra dinh muc theo TT 09 va 12\Tra dinh muc TT09.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=False)

print("--- Checking External Links ---")
# Openpyxl stores external links in the workbook object
try:
    print(f"External Links: {wb.external_links}")
except:
    print("Could not read wb.external_links directly.")

print("\n--- Scanning Formulas for External References ---")
# Look for '[' which denotes an external workbook reference in formulas
external_refs = []
for sheet in wb.sheetnames:
    ws = wb[sheet]
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and "=" in cell.value:
                if "[" in cell.value and "]" in cell.value:
                    external_refs.append(f"{sheet}!{cell.coordinate}: {cell.value}")
                    if len(external_refs) > 10: break
    if len(external_refs) > 10: break

if external_refs:
    print(f"Found {len(external_refs)}+ external references:")
    for ref in external_refs:
        print(ref)
else:
    print("No obvious external formulas found.")

print("\n--- Checking Defined Names for External Refs ---")
for name in wb.defined_names:
    defn = wb.defined_names[name]
    if "[" in str(defn.attr_text) or "http" in str(defn.attr_text):
        print(f"External Name: {name} -> {defn.attr_text}")
