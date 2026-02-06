
import openpyxl

file_path = r"g:\My Drive\30_RESOURCES (Kho tai nguyen)\30.01_Tinh hieu qua du an\02_Tra dinh muc theo TT 09 va 12\Tra dinh muc TT09.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=False) # Get Formulas
ws = wb["TT09"]

print("--- Inspecting Row 725 (Header?) and below ---")
# Usually calculation is below header.
for r in range(725, 740):
    col2 = ws.cell(row=r, column=2).value # Name
    col5 = ws.cell(row=r, column=5).value # Formula E
    print(f"Row {r}: {str(col2)[:50]}... | Formula E: {col5}")
    
# Specifically look for i_tkdutoan in this area
