
import openpyxl

file_path = r"g:\My Drive\30_RESOURCES (Kho tai nguyen)\30.01_Tinh hieu qua du an\02_Tra dinh muc theo TT 09 va 12\Tra dinh muc TT09.xlsx"
# Note: User path might have changed or they opened a different file "Tra dinh muc TT09.xlsx" vs "tra TMDT.xlsx"
# The metadata says Active Document: ...\Tra dinh muc TT09.xlsx
# The previous file was "tra TMDT.xlsx".
# I should check if this new file has the same issue.

print(f"Checking file: {file_path}")
try:
    wb = openpyxl.load_workbook(file_path, data_only=False)
except FileNotFoundError:
    print("File not found at new path. Trying old path...")
    file_path = r"g:\My Drive\00_INBOX (xu ly nhanh)\tra TMDT.xlsx"
    wb = openpyxl.load_workbook(file_path, data_only=False)

print(f"Loaded: {file_path}")

# Check Named Ranges
names_to_fix = ["i_CPC", "i_CapCT", "i_BTK", "i_tkdutoan", "i_VB"]
print("\n--- Current Named Ranges ---")
for name in names_to_fix:
    if name in wb.defined_names:
        d = wb.defined_names[name]
        print(f"{name} -> {d.attr_text} (Dest: {list(d.destinations)})")
    else:
        print(f"{name} -> Not Found")

print("\n--- Sheet Names ---")
print(wb.sheetnames)
