
import pandas as pd
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.workbook.defined_name import DefinedName
import re

input_excel = r"g:\My Drive\30_RESOURCES (Kho tai nguyen)\30.01_Tinh hieu qua du an\02_Tra dinh muc theo TT 09 va 12\Du_lieu_Suat_von_dau_tu_2024.xlsx"
output_excel = r"g:\My Drive\30_RESOURCES (Kho tai nguyen)\30.01_Tinh hieu qua du an\02_Tra dinh muc theo TT 09 va 12\File_Tra_Cuu_Final.xlsx"

print("Reading Data...")
df = pd.read_excel(input_excel, header=None)

# 1. CLEAN DATA
refined_data = []
current_code = ""

for idx, row in df.iterrows():
    row_strs = [str(x).strip() if not pd.isna(x) else "" for x in row]
    if "--- BẢNG" in row_strs[0]: continue
    
    code = row_strs[0]
    desc = row_strs[1]
    
    if code: current_code = code 
    vals = row_strs[2:6]
    has_number = any(re.search(r'[\d.,]+', v) for v in vals)
    
    if desc and has_number:
        unit = "1000đ/m2/đơn vị"
        val1, val2, val3 = 0, 0, 0
        nums = []
        for v in vals:
            clean = v.replace('.', '').replace(',', '.')
            try: nums.append(float(clean))
            except: pass
        if len(nums) >= 1: val1 = nums[0]
        if len(nums) >= 2: val2 = nums[1]
        if len(nums) >= 3: val3 = nums[2]
        refined_data.append([current_code, desc, unit, val1, val2, val3])

print(f"Refined {len(refined_data)} rows.")

# 2. SETUP WORKBOOK
wb = openpyxl.Workbook()

# --- SHEET: DATA ---
ws_data = wb.active
ws_data.title = "DATA"
headers = ["Mã hiệu", "Loại công trình", "Đơn vị tính", "Suất vốn đầu tư", "Chi phí Xây dựng", "Chi phí Thiết bị"]
ws_data.append(headers)
for row in refined_data:
    ws_data.append(row)

# Formatting
header_font = Font(bold=True)
header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
for cell in ws_data[1]: 
    cell.font = header_font
    cell.fill = header_fill
ws_data.column_dimensions['B'].width = 50

# --- SHEET: LISTS ---
ws_lists = wb.create_sheet("LISTS")
ws_lists.sheet_state = 'hidden' 

unique_codes = sorted(list(set(r[0] for r in refined_data)))
ws_lists['A1'] = "UniqueCodes"
for i, c in enumerate(unique_codes):
    ws_lists[f'A{i+2}'] = c

# Defined Name: List_Codes
code_rng_str = f"LISTS!$A$2:$A${len(unique_codes)+1}"
d = DefinedName("List_Codes", attr_text=code_rng_str)
wb.defined_names.add(d) # Changed from append to add

# Defined Name: Col_Code, Col_Desc for Dependent Logic
ws_lists['C1'] = "Code"
ws_lists['D1'] = "Description"
sorted_data = sorted(refined_data, key=lambda x: x[0])
for i, row in enumerate(sorted_data):
    ws_lists[f'C{i+2}'] = row[0] 
    ws_lists[f'D{i+2}'] = row[1] 

max_r = len(sorted_data) + 1
wb.defined_names.add(DefinedName("Col_Code", attr_text=f"LISTS!$C$2:$C${max_r}"))
wb.defined_names.add(DefinedName("Col_Desc", attr_text=f"LISTS!$D$2:$D${max_r}"))

# --- SHEET: TRA_CUU ---
ws_ui = wb.create_sheet("TRA_CUU")
ws_ui.column_dimensions['B'].width = 20
ws_ui.column_dimensions['C'].width = 40

ws_ui['B2'] = "TRA CỨU SUẤT VỐN ĐẦU TƯ"
ws_ui['B2'].font = Font(size=14, bold=True, color="0000FF")

# L1 Input
ws_ui['B4'] = "1. Chọn Mã hiệu:"
ws_ui['B4'].font = Font(bold=True)
input_code = ws_ui['C4']
input_code.border = Border(bottom=Side(style='thin'))
dv1 = DataValidation(type="list", formula1="=List_Codes", allow_blank=True)
ws_ui.add_data_validation(dv1)
dv1.add(input_code)

# L2 Input
ws_ui['B5'] = "2. Chọn Loại công trình:"
ws_ui['B5'].font = Font(bold=True)
input_desc = ws_ui['C5']
input_desc.border = Border(bottom=Side(style='thin'))
input_desc.alignment = Alignment(wrap_text=True)

# Using Semicolon separator
dv2_formula = "=OFFSET(LISTS!$D$2;MATCH($C$4;Col_Code;0)-1;0;COUNTIF(Col_Code;$C$4);1)"
dv2 = DataValidation(type="list", formula1=dv2_formula, allow_blank=True)
dv2.errorTitle = "" 
ws_ui.add_data_validation(dv2)
dv2.add(input_desc)

# Results
res_labels = ["Đơn vị tính", "Suất vốn đầu tư", "Chi phí Xây dựng", "Chi phí Thiết bị"]
col_map = [3, 4, 5, 6]
ws_ui['B8'] = "KẾT QUẢ:"
ws_ui['B8'].font = Font(bold=True, underline="single")

for i, lbl in enumerate(res_labels):
    r = 9 + i
    ws_ui[f'B{r}'] = lbl
    ws_ui[f'B{r}'].font = Font(bold=True)
    col_idx = col_map[i]
    col_letter = openpyxl.utils.get_column_letter(col_idx)
    
    if i == 0: 
        f = "1000đ/m2 (hoặc theo ĐV)"
    else:
        # Use simple SUMIFS with semicolon
        f = f"=SUMIFS(DATA!{col_letter}:{col_letter}; DATA!$A:$A; $C$4; DATA!$B:$B; $C$5)"
        ws_ui[f'C{r}'].number_format = '#,##0'
    ws_ui[f'C{r}'] = f

print(f"Saving to {output_excel}...")
wb.save(output_excel)
print("Done.")
