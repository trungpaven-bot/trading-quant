
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Paths
input_file = r'g:\My Drive\30_RESOURCES (Kho tai nguyen)\30.01_Tinh hieu qua du an\02_Tra dinh muc theo TT 09 va 12\2026.01.17 Tien do du an CCN PDP 2 - Chi phi HĐ V1.0.xlsx'
output_file = r'g:\My Drive\30_RESOURCES (Kho tai nguyen)\30.01_Tinh hieu qua du an\02_Tra dinh muc theo TT 09 va 12\2026.01.17 Tien do du an CCN PDP 2 - Chi phi HĐ V1.0_Professional.xlsx'

# Load workbook
wb = openpyxl.load_workbook(input_file)
ws = wb['Tien do']

# Define Styles
header_font = Font(bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid") # Dark Blue
sub_header_fill = PatternFill(start_color="95B3D7", end_color="95B3D7", fill_type="solid") # Lighter Blue (Quarters)

parent_font = Font(bold=True, size=11)
parent_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid") # Very Light Blue
grand_total_fill = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid") # Slightly Darker Blue

border_thin = Side(border_style="thin", color="BFBFBF") # Gray border
border_dotted = Side(border_style="dotted", color="BFBFBF")
simple_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)

# Currency Format
currency_format = '#,##0'

# --- 1. Fix Headers (Rows 1-4 in Excel, 0-3 in code) ---
# Assuming Headers are row 4 (index 4 in openpyxl 1-based)
# Row 0 (index 1): Title
# Row 1 (index 2): Years
# Row 2 (index 3): Quarters
# Row 3 (index 4): Main Columns

# Style the main header row (Row 4)
for col in range(1, 100): # Iterate enough columns
    cell = ws.cell(row=4, column=col)
    cell_val = cell.value
    if cell_val:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = simple_border
    
    # Style Months (Col 19+)
    if col >= 19:
        cell = ws.cell(row=4, column=col)
        cell.alignment = Alignment(horizontal='center', vertical='center')

# Header Merges Correction (Years and Quarters)
# We need to ensure 'NĂM 2026' spans 12 columns
# Start col for months is 19 (S is 19th letter). Code uses 1-based index. 
# "1" is at col 19.
# Let's find "NĂM 2026" in Row 2
found_year = False
for col in range(1, 100):
    cell = ws.cell(row=2, column=col)
    if cell.value and "NĂM" in str(cell.value):
        # Assuming standard year structure
        # Merge this cell with next 11 cells
        ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col+11)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(bold=True)
        cell.fill = sub_header_fill
        cell.border = simple_border

# --- 2. Process Data Rows (Row 5 onwards) ---
# Start from Row 5 (index 5)
max_row = ws.max_row
cols_to_format_money = [8, 9, 17, 18] # Contract Value, etc. + Cash flows
# Note: Cash flows are from col 19 onwards

# Find the Grand Total Row (Row 5)
grand_total_row_idx = 5 
# Style Grand Total
for col in range(1, ws.max_column + 1):
    cell = ws.cell(row=grand_total_row_idx, column=col)
    cell.fill = grand_total_fill
    cell.font = Font(bold=True)
    cell.border = simple_border
    if col >= 19 or col in [8, 9, 18]:
         cell.number_format = currency_format

# Iterate tasks
current_parent_row = None
for row_idx in range(6, max_row + 1):
    row_num_cell = ws.cell(row=row_idx, column=1) # Column A (STT)
    task_name_cell = ws.cell(row=row_idx, column=2) # Column B
    
    # Check if this is a Parent Row (Has number in Col A)
    if row_num_cell.value is not None:
        # It's a Parent
        for col in range(1, ws.max_column + 1):
            c = ws.cell(row=row_idx, column=col)
            c.font = parent_font
            c.fill = parent_fill
            c.border = simple_border
            if col >= 19 or col in [8, 9, 18]:
                c.number_format = currency_format
            
            # Date Formatting (Cols 4 and 5)
            if col in [4, 5]:
                 c.number_format = 'DD/MM/YYYY'
                 c.alignment = Alignment(horizontal='center', vertical='center')
            
            # Duration and Status Center (Col 3 and 6)
            if col in [3, 6]:
                 c.alignment = Alignment(horizontal='center', vertical='center')

    else:
        # It's a Child
        # Group it (Outline Level 1)
        ws.row_dimensions[row_idx].outlineLevel = 1
        
        # Style Child
        for col in range(1, ws.max_column + 1):
            c = ws.cell(row=row_idx, column=col)
            c.border = simple_border
            if col >= 19 or col in [8, 9, 18]:
                 c.number_format = currency_format
            
            # Date Formatting (Cols 4 and 5)
            if col in [4, 5]:
                 c.number_format = 'DD/MM/YYYY'
                 c.alignment = Alignment(horizontal='center', vertical='center')
            
            # Duration and Status Center (Col 3 and 6)
            if col in [3, 6]:
                 c.alignment = Alignment(horizontal='center', vertical='center')

# --- 3. Final Touches ---
# Freeze Panes (Row 5, Col 3 - C)
ws.freeze_panes = "C5"

# Auto-fit columns (Approximate)
for col_idx in range(1, 10):
    col_letter = get_column_letter(col_idx)
    ws.column_dimensions[col_letter].width = 15 # Default width for info cols
ws.column_dimensions['B'].width = 50 # Task name wide

# Set Month columns width
for col_idx in range(19, ws.max_column + 1):
    col_letter = get_column_letter(col_idx)
    ws.column_dimensions[col_letter].width = 12

wb.save(output_file)
print(f"File saved to: {output_file}")
