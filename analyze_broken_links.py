
import openpyxl

file_path = r"g:\My Drive\30_RESOURCES (Kho tai nguyen)\30.01_Tinh hieu qua du an\02_Tra dinh muc theo TT 09 va 12\Tra dinh muc TT09.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=False)

print("--- Analyzing Broken Links Context ---") 
# User is right. Nuke to 0 is destructive.
# We need to list WHERE the links were and WHAT they likely mean so user can fix them.
# I will scan again (similar to before but focusing on context).

# My previous scan result:
# TT09!D942: ='[2]THCP TVXD'!D9
# ...
# These look like "Tổng hợp chi phí tư vấn xây dựng" (THCP TVXD).
# It seems this sheet TT09 tries to PULL results FROM another workbook "THCP...".
# OR does it push?
# Usually formulas PULL. =OtherBook!Cell

broken_links_report = []

for sheet in wb.sheetnames:
    ws = wb[sheet]
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and "=" in cell.value:
                if "[" in cell.value: # External
                    # Context: Get headers of this column?
                    # Get nearby text?
                    broken_links_report.append({
                        "Sheet": sheet,
                        "Cell": cell.coordinate,
                        "Formula": cell.value,
                        "Value": "UNK" # Can't get value without data_only=True load separately
                    })
                    if len(broken_links_report) > 50: break
    if len(broken_links_report) > 50: break

print(f"Found {len(broken_links_report)} broken links. Sample:")
for item in broken_links_report[:20]:
    print(f"{item['Sheet']}!{item['Cell']} -> {item['Formula']}")

# Also check Defined Names again if I haven't deleted them yet? 
# (I deleted them in previous turn, so they are gone from the file on disk unless user reverted).
# But assuming user MIGHT have reverted or I should explain based on memory/logs.
# The logs showed:
# CHI_TIEU!B12: =+[6]DAU_VAO!B39
# This implies there is a "DAU_VAO" file (Input Data).
