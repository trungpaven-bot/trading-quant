
import pandas as pd
import openpyxl
from openpyxl.styles import Font
import re

input_excel = r"g:\My Drive\30_RESOURCES (Kho tai nguyen)\30.01_Tinh hieu qua du an\02_Tra dinh muc theo TT 09 va 12\Du_lieu_Co_Tieu_De_Bang.xlsx"
output_excel = r"g:\My Drive\30_RESOURCES (Kho tai nguyen)\30.01_Tinh hieu qua du an\02_Tra dinh muc theo TT 09 va 12\File_Tra_Cuu_Lists_Only.xlsx"

print("Reading Data...")
df = pd.read_excel(input_excel, header=None)

processed_data = []

# Logic State
current_header_stack = [] 
current_table_name = ""
last_valid_code = ""

start_processing = False

# --- PROCESSING LOOP (Robust Logic from StrictReset) ---
for idx, row in df.iterrows():
    row_strs = [str(x).strip() if not pd.isna(x) else "" for x in row]
    table_name = row_strs[0]
    
    # Filter 1-101
    match = re.search(r'Bảng (\d+)', table_name, re.IGNORECASE)
    if match:
        tbl_num = int(match.group(1))
        if not (1 <= tbl_num <= 101): continue
        if not start_processing: start_processing = True
    else:
        # If we haven't started, skip. If we have started, check if it's a valid table line
        # Actually Strict Filter on "Bảng starts with" is safer
        continue
    
    # Reset Stack on Table Change
    if table_name != current_table_name:
        current_table_name = table_name
        current_header_stack = [] 
        last_valid_code = ""
    
    code = row_strs[1]
    desc = row_strs[2]
    vals = row_strs[3:]
    has_number = any(re.search(r'[\d.,]+', v) for v in vals)

    if has_number and desc:
        # DATA ROW -> Generate Output
        
        # Combine Headers
        prefix = " - ".join(current_header_stack) # Use " - " for clearer separation? Or just space?
        # User accepted "Space" before, but " - " might be cleaner for deep nesting. 
        # Let's use Space to match "Custom Fix" visual exactness.
        prefix = " ".join(current_header_stack)
        
        if prefix:
            final_desc = f"{prefix} {desc}"
        else:
            final_desc = desc
            
        if code: last_valid_code = code; final_code = code
        else: final_code = last_valid_code
        
        # Store: [TableNum, TableName, Code, FinalDesc]
        processed_data.append([tbl_num, table_name, final_code, final_desc])
        
    else:
        # HEADER ROW -> Build Stack
        if desc:
            if code:
                # Major header implies new block
                current_header_stack = [desc]
                last_valid_code = code
            else:
                # Minor header
                if len(current_header_stack) == 0:
                    current_header_stack.append(desc)
                else:
                    # Heuristic: If we are deep (len>=2), replace last. If shallow, append.
                    # Table 65: "Ct..." -> "Anten 30m". Stack len 2.
                    # Next: "Anten 45m". Replace "Anten 30m".
                    if len(current_header_stack) >= 2:
                        current_header_stack[-1] = desc
                    else:
                        current_header_stack.append(desc)

# Sort Natural
processed_data.sort(key=lambda x: x[0]) # Sort by Tbl Num

# --- BUILD EXCEL ---
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "LISTS"

# Header
ws['A1'] = "Bảng"
ws['C1'] = "Mã hiệu"
ws['D1'] = "Loại công trình"

# Write Data
for i, r in enumerate(processed_data):
    row_idx = i + 2
    ws[f'A{row_idx}'] = r[1] # TableName
    ws[f'C{row_idx}'] = r[2] # Code
    ws[f'D{row_idx}'] = r[3] # Desc

# Formatting
for cell in ws[1]: cell.font = Font(bold=True)
ws.column_dimensions['A'].width = 40
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 80

print(f"Saving {output_excel}...")
wb.save(output_excel)
print("Done.")
