
import openpyxl

file_path = r"g:\My Drive\00_INBOX (xu ly nhanh)\tra TMDT.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=False)

print("--- Sheets including hidden ---")
for sheet in wb.sheetnames:
    ws = wb[sheet]
    print(f"Sheet: {sheet}, State: {ws.sheet_state}")

print("\n--- Searching for usages of variables in TT09 ---")
ws_tt09 = wb["TT09"]
variables = ["i_BTK", "i_CapCT", "i_CPC", "i_tkdutoan", "i_VB"]
# Scan formulas in TT09
found_usages = {}

for row in ws_tt09.iter_rows():
    for cell in row:
        val = str(cell.value)
        if hasattr(cell, 'value') and isinstance(cell.value, str) and "=" in cell.value:
            for v in variables:
                if v in val:
                    if v not in found_usages: found_usages[v] = []
                    found_usages[v].append(f"Row {cell.row}: {cell.coordinate} = {val}")

for v in variables:
    print(f"\nUsages of {v}:")
    if v in found_usages:
        # Print first 5 usages
        for usage in found_usages[v][:5]:
            print(f"  {usage}")
    else:
        print("  None found")
