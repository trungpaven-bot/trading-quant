
import openpyxl

file_path = r"g:\My Drive\00_INBOX (xu ly nhanh)\tra TMDT.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)

if "CHI_TIEU" in wb.sheetnames:
    ws = wb["CHI_TIEU"]
    print("--- Reading CHI_TIEU ---")
    for r in range(1, 50):
        row_vals = []
        for c in range(1, 10):
            val = ws.cell(row=r, column=c).value
            if val:
                row_vals.append(str(val))
        if row_vals:
            print(f"Row {r}: {row_vals}")
else:
    print("CHI_TIEU not found")
