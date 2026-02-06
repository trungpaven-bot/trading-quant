
import pandas as pd
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.workbook.defined_name import DefinedName
import re

input_excel = r"g:\My Drive\30_RESOURCES (Kho tai nguyen)\30.01_Tinh hieu qua du an\02_Tra dinh muc theo TT 09 va 12\Du_lieu_Co_Tieu_De_Bang.xlsx"
output_excel = r"g:\My Drive\30_RESOURCES (Kho tai nguyen)\30.01_Tinh hieu qua du an\02_Tra dinh muc theo TT 09 va 12\File_Tra_Cuu_Final_SmartFix.xlsx"

print("Reading Data...")
df = pd.read_excel(input_excel, header=None)

processed_data = []

# State variables for hierarchical parsing
current_parent_desc = "" # Description of the last row having a Code
current_parent_code = ""

start_processing = False

for idx, row in df.iterrows():
    row_strs = [str(x).strip() if not pd.isna(x) else "" for x in row]
    
    table_name = row_strs[0]
    
    # Start filtering from Table 1
    if not start_processing:
        if "Bảng 1" in table_name or "Bang 1" in table_name:
            start_processing = True
        else:
            continue
            
    code = row_strs[1]
    desc = row_strs[2]
    
    # Parse Values
    vals = row_strs[3:]
    nums = []
    for v in vals:
        clean = v.replace('.', '').replace(',', '.')
        try: nums.append(float(clean))
        except: pass
        
    val1 = nums[0] if len(nums) >= 1 else 0
    val2 = nums[1] if len(nums) >= 2 else 0
    val3 = nums[2] if len(nums) >= 3 else 0
    
    has_values = (len(nums) > 0)
    
    # --- LOGIC MỚI: BÁM THEO MÃ HIỆU ---
    
    final_desc = desc
    final_code = code

    # Case 1: Dòng CÓ Mã hiệu (Đây là Dòng Cha hoặc Dòng Độc lập)
    if code:
        current_parent_code = code
        current_parent_desc = desc # Cập nhật Cha mới
        final_code = code
        
        # Nếu dòng này CÓ Tiền -> Nó là dòng độc lập -> Không cần sửa Desc, giữ nguyên
        # Nếu dòng này KHÔNG có Tiền -> Nó chỉ là Tiêu đề Cha -> Không thêm vào Data (hoặc thêm nhưng ko có tiền)
        # Nhưng đợi đã, trong Bảng 2, dòng 11120.06 "Nhà... có 1 tầng hầm" là Cha, KHÔNG có tiền.
        # Các dòng con "DT <50m2" bên dưới sẽ lấy tiền và ghép tên ông Cha này.
        
        if not has_values:
            # Đây thuần túy là Header Nhóm -> Không add vào database tra cứu (vì ko có số để tra)
            # Chỉ lưu lại state để dòng con dùng
            continue 
        else:
            # Có mã, Có tiền -> Dòng chuẩn -> Add luôn
            pass

    # Case 2: Dòng KHÔNG Mã hiệu (Đây là Dòng Con)
    else:
        # Nếu dòng này có Tiền -> Chắc chắn là item con cần tra cứu
        if has_values:
            # Ghép tên Cha + tên Con
            # Tuy nhiên, cần tránh ghép ngớ ngẩn. 
            # Ví dụ: Cha="Nhà 1 tầng", Con="Nhà 1 tầng mái ngói". Ghép lại thừa.
            # Nhưng ở đây: Cha="Nhà... ko hầm", Con="Có 1 hầm" -> Ghép OK.
            # Cha="Nhà... ko hầm", Con="DT <50m2" -> Ghép OK.
            
            # Logic ghép: Thêm dấu cách
            if current_parent_desc:
                # Kiểm tra xem con có lặp lại từ đầu của cha không? (ít khi)
                final_desc = f"{current_parent_desc} - {desc}"
            
            # Gán mã của cha cho con
            final_code = current_parent_code
            
        else:
            # Không mã, không tiền -> Dòng rác hoặc Helper text -> Bỏ qua
            continue

    # Add to result
    processed_data.append([table_name, final_code, final_desc, "1000đ/m2", val1, val2, val3])

print(f"Processed {len(processed_data)} items.")

# --- BUILD EXCEL ---
wb = openpyxl.Workbook()

# 1. LISTS
ws_lists = wb.active; ws_lists.title="LISTS"
ws_lists['A1']="Bảng"; ws_lists['C1']="Mã"; ws_lists['D1']="Tên CT"
for i, r in enumerate(processed_data):
    row=i+2; ws_lists[f'A{row}']=r[0]; ws_lists[f'C{row}']=r[1]; ws_lists[f'D{row}']=r[2]

# 2. ENGINE
ws_eng = wb.create_sheet("ENGINE")
ws_eng.append(["", "Bảng", "Bảng2", "Mã", "Tên", "ĐV", "Suất", "XD", "TB", "Key"])
for i, r in enumerate(processed_data):
    row=i+2
    ws_eng.cell(row,2,r[0]); ws_eng.cell(row,3,r[0]); ws_eng.cell(row,4,r[1]); ws_eng.cell(row,5,r[2])
    ws_eng.cell(row,6,r[3]); ws_eng.cell(row,7,r[4]); ws_eng.cell(row,8,r[5]); ws_eng.cell(row,9,r[6])
    ws_eng.cell(row,10).value = f'=B{row} & "_" & E{row}'

# 3. TRA_CUU
ws_ui = wb.create_sheet("TRA_CUU")
ws_ui['B2']="TRA CỨU SUẤT VỐN (SMART FIX)"; ws_ui['B2'].font=Font(size=14, bold=True)
ws_ui['B4']="1. Chọn Bảng:"; ws_ui['C4'].border=Border(bottom=Side(style='thin')); 
ws_ui['B5']="2. Chọn CT:"; ws_ui['C5'].border=Border(bottom=Side(style='thin')); ws_ui['C5'].alignment=Alignment(wrap_text=True)

# Validation 1
tbls = sorted(list(set(r[0] for r in processed_data)))
ws_lists['Z1']="UnqTbls"; 
for i,t in enumerate(tbls): ws_lists[f'Z{i+2}']=t
wb.defined_names.add(DefinedName("ListTables", attr_text=f"LISTS!$Z$2:$Z${len(tbls)+1}"))
ws_ui.add_data_validation(DataValidation(type="list", formula1="=ListTables", allow_blank=True)); ws_ui['C4'].parent.data_validations.dataValidation[-1].add(ws_ui['C4'])

# Validation 2
cnt = len(processed_data)
wb.defined_names.add(DefinedName("ColTbl", attr_text=f"LISTS!$A$2:$A${cnt+1}"))
wb.defined_names.add(DefinedName("ColDesc", attr_text=f"LISTS!$D$2:$D${cnt+1}"))
dv2_f = "=OFFSET(LISTS!$D$2;MATCH($C$4;ColTbl;0)-1;0;COUNTIF(ColTbl;$C$4);1)"
ws_ui.add_data_validation(DataValidation(type="list", formula1=dv2_f, allow_blank=True)); ws_ui['C5'].parent.data_validations.dataValidation[-1].add(ws_ui['C5'])

# Results
ws_ui['B8']="KẾT QUẢ:"; labels=["Mã","ĐV","Suất","XD","TB"]; cols=["D","F","G","H","I"]
for i,l in enumerate(labels):
    r=9+i; ws_ui[f'B{r}']=l; ws_ui[f'B{r}'].font=Font(bold=True)
    f=f"=INDEX(ENGINE!{cols[i]}:{cols[i]}; MATCH($C$4 & \"_\" & $C$5; ENGINE!$J:$J; 0))"
    if i>=2: ws_ui[f'C{r}'].number_format='#,##0'
    ws_ui[f'C{r}']=f

print(f"Saving {output_excel}...")
wb.save(output_excel)
print("Done.")
