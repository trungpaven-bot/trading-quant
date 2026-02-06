
import openpyxl

file_path = r"g:\My Drive\00_INBOX (xu ly nhanh)\tra TMDT.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)
ws = wb["TT09"]

labels = ["Thẩm định thiết kế, dự toán:", "Kiểm toán/QT:", "Cấp công trình:", "Số bước thiết kế:"]
label_positions = {}

print("--- Locating Labels ---")
for r in range(1, 20):
    for c in range(1, 20):
        cell = ws.cell(row=r, column=c)
        val = str(cell.value).strip() if cell.value else ""
        for label in labels:
            # Flexible match
            if label.replace(":", "").lower() in val.lower():
                print(f"Found '{val}' at {cell.coordinate}")
                label_positions[val] = (r, c)

print("\n--- Checking Data Validation in potential target cells ---")
# Assuming the input is to the right of the label
for label, (r, c) in label_positions.items():
    target_cells = [
        ws.cell(row=r, column=c+1), # Right
        ws.cell(row=r, column=c+2), # Right + 1
        ws.cell(row=r+1, column=c)  # Below
    ]
    
    for cell in target_cells:
        print(f"Checking {cell.coordinate} (value: {cell.value})")
        # Check validation
        if cell.coordinate in ws.data_validations.dataValidation:
            print(f"  -> FOUND Validation on {cell.coordinate}!")
        else:
             # openpyxl data_validations is a list of DataValidation objects
             # we need to check if the coordinate falls within one of the ranges
             found_val = False
             for dv in ws.data_validations.dataValidation:
                 if cell.coordinate in dv.sqref:
                     print(f"  -> FOUND Validation in Range {dv.sqref}")
                     print(f"     Type: {dv.type}, Formula1: {dv.formula1}")
                     found_val = True
             if not found_val:
                 print("  -> No Validation")

print("\n--- Checking for Hidden Lists/Ranges ---")
# Check if there are lists in the hidden columns nearby?
# The user image shows them in rows ~5-7. 
# Previous search found "Cấp I" at G98. That's far away.
