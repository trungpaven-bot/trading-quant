
import openpyxl

file_path = r"g:\My Drive\30_RESOURCES (Kho tai nguyen)\30.01_Tinh hieu qua du an\02_Tra dinh muc theo TT 09 va 12\Tra dinh muc TT09.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=False)
ws = wb["TT09"]

print("--- Searching for 'Thẩm định' rows ---")
target_rows = []
for r in range(1, 1000): # Scan deep enough
    val = ws.cell(row=r, column=2).value
    if val and isinstance(val, str) and "thẩm định" in val.lower():
        formula_e = ws.cell(row=r, column=5).value # Col E (Rate?)
        formula_f = ws.cell(row=r, column=6).value
        # Check if formula uses i_tkdutoan or C128
        print(f"Row {r}: {val}")
        if formula_e: print(f"  Col E: {formula_e}")
        if formula_f: print(f"  Col F: {formula_f}")
        
        # Check if this row seems to be the one controlled by the selection
        if "thiết kế" in val.lower() or "dự toán" in val.lower():
            target_rows.append(r)

print("\n--- Checking specific logic ---")
# If selection 2 is "Chủ đầu tư tổ chức", it usually implies internal cost or hiring consultant.
# If user wants 0, we can modify the formula to =IF(i_tkdutoan=2, 0, ...)
