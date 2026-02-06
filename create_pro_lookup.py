
import pandas as pd
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import re

input_excel = r"g:\My Drive\30_RESOURCES (Kho tai nguyen)\30.01_Tinh hieu qua du an\02_Tra dinh muc theo TT 09 va 12\Du_lieu_Suat_von_dau_tu_2024.xlsx"
output_excel = r"g:\My Drive\30_RESOURCES (Kho tai nguyen)\30.01_Tinh hieu qua du an\02_Tra dinh muc theo TT 09 va 12\File_Tra_Cuu_Chuyen_Nghiep.xlsx"

print("Reading and Refining Data...")
df = pd.read_excel(input_excel, header=None)

refined_data = []
current_code = ""

# --- Step 1: Smart Data Extraction & Filling ---
for idx, row in df.iterrows():
    row_strs = [str(x).strip() if not pd.isna(x) else "" for x in row]
    if "--- BẢNG" in row_strs[0]: continue
    
    code = row_strs[0]
    desc = row_strs[1]
    
    # Fill Code Down Logic
    if code:
        current_code = code
    
    # Check if data row (has numbers in val cols)
    vals = row_strs[2:6]
    has_number = any(re.search(r'[\d.,]+', v) for v in vals)
    
    if desc and has_number:
        # Standardize Unit
        unit = "1000đ/m2/đơn vị"
        # If val col 0 is text (unit) instead of number, shift
        val1, val2, val3 = "", "", ""
        
        # Heuristic to grab the 3 numbers: Total, Const, Device
        nums = []
        for v in vals:
            if re.search(r'^\d+([.,]\d+)?$', v):
                nums.append(v)
        
        if len(nums) >= 1: val1 = nums[0]
        if len(nums) >= 2: val2 = nums[1]
        if len(nums) >= 3: val3 = nums[2]
        
        refined_data.append([current_code, desc, unit, val1, val2, val3])

print(f"Refined {len(refined_data)} rows.")

# --- Step 2: Organize for 2-Level Dropdown ---
# We need unique Categories (Level 1) and Items (Level 2).
# Since the raw data flat, we need to infer Level 1.
# Assumption: The "Code" usually groups items.
# Or we can group by the "Header" text found in the Word doc (which we lost in CSV).
# Alternative: Use "Code" as the Grouper Key.
# Map: Code -> [List of Descriptions]
# And we need a readable Name for the Code. 
# Problem: We don't have the "Parent Name" (e.g. "Cong trinh nha o") easily associated with Code "11100".
# 
# Workaround:
# Level 1: Select "Mã hiệu" (Code) + "Generic Name" (First item of that code?)
# Level 2: Select "Description"

groups = {}
group_names = {}

for row in refined_data:
    code = row[0]
    desc = row[1]
    if code not in groups:
        groups[code] = []
        # Try to make a friendly name. 
        # Usually the first entry of a code block is generic? 
        # Or we just use the Code itself.
        group_names[code] = f"Nhóm {code}" 
    
    groups[code].append(row)

# Update Group Names based on the first item's description (often meaningful)
for code, items in groups.items():
    # Heuristic: If there are many items, the first one might not be the title. 
    # But usually Code 11110.01 corresponding to "So tang <= 5..." 
    # The Dictionary (Decision 409) has structure.
    # Let's just use the Code as Level 1 for now to ensure correctness, 
    # or combined "Code - First Desc Snippet"
    snippet = items[0][1][:30] + "..."
    group_names[code] = f"{code} - {snippet}"

# --- Step 3: Excel Construction ---
wb = openpyxl.Workbook()

# DATA Sheet
ws_data = wb.active
ws_data.title = "DATA"
ws_data.append(["Code", "Description", "Unit", "Total", "Const", "Equip", "LookupKey"])

# Write Data and Create Lookup Keys
row_idx = 2
for code, items in groups.items():
    for item in items:
        # Create a unique key for VLOOKUP: Code + Description
        # But Description might be long. 
        # Let's clean description for key: Remove special chars
        clean_desc = re.sub(r'[^a-zA-Z0-9]', '', item[1])
        key = f"{code}_{clean_desc}"
        
        ws_data.append(item + [key])
        row_idx += 1

# LISTS Sheet (For Dropdowns)
ws_lists = wb.create_sheet("LISTS")

# Col A: Level 1 (Codes)
ws_lists['A1'] = "Danh sách Mã"
code_list = list(groups.keys())
for i, code in enumerate(code_list):
    ws_lists[f'A{i+2}'] = group_names[code] # Display Name
    ws_lists[f'B{i+2}'] = code # Real Code value (Helper)

# Create Named Range for Level 1
ws_lists.sheet_state = 'hidden' # Hide later
wb.create_named_range("List_Codes", ws_lists, f"$A$2:$A${len(code_list)+1}")

# For Level 2, we need dynamic lists.
# Excel Indirect requires Named Ranges matching the text. 
# Codes like "11110.01" are not valid Names (start with number, dot).
# We must use OFFSET/MATCH approach in Formula.
# Or just one big list and filter? No.
#
# Better Approach for User:
# 1. TRA_CUU Sheet
# Cell C4: Select Group (Code - Name)
# Cell C5: Select Detail (Description)
#
# To make C5 filtered by C4:
# We put all Descriptions for Code X in a contiguous block in 'LISTS' sheet.
# Then define a specific Name for that block? Too many names.
#
# Use LOOKUP logic in Validation Formula?
# Validation Formula: =OFFSET(StartCell, MATCH(SelectedCode, ColumnCodes, 0)-1, 1, COUNTIF(ColumnCodes, SelectedCode), 1)
# This requires columns: [Code] [Description] sorted by Code.

# Rearrange LISTS sheet for this logic
ws_lists['D1'] = "CodeCol"
ws_lists['E1'] = "DescCol"
r = 2
for code in code_list:
    disp_name = group_names[code]
    for item in groups[code]:
        ws_lists[f'D{r}'] = disp_name # Match the Level 1 selection exactly
        ws_lists[f'E{r}'] = item[1]   # Description
        r += 1

max_r = r - 1
# Define Names for Columns
wb.create_named_range("Col_Group", ws_lists, f"$D$2:$D${max_r}")
wb.create_named_range("Col_Desc", ws_lists, f"$E$2:$E${max_r}")

# --- TRA CUU Sheet ---
ws_ui = wb.create_sheet("TRA_CUU")
ws_ui.column_dimensions['B'].width = 25
ws_ui.column_dimensions['C'].width = 60

ws_ui['B2'] = "TRA CỨU SUẤT VỐN ĐẦU TƯ"
ws_ui['B2'].font = Font(size=14, bold=True)

# L1 Input
ws_ui['B4'] = "1. Chọn Nhóm:"
ws_ui['C4'].border = Border(bottom=Side(style='thin'))
dv1 = DataValidation(type="list", formula1="=List_Codes", allow_blank=True)
ws_ui.add_data_validation(dv1)
dv1.add(ws_ui['C4'])

# L2 Input (DEPENDENT)
ws_ui['B5'] = "2. Chọn Chi tiết:"
ws_ui['C5'].border = Border(bottom=Side(style='thin'))

# Formula explanation:
# OFFSET(StartOfDescColumn, MATCH(C4, GroupColumn, 0)-1, 0, COUNTIF(GroupColumn, C4), 1)
# Note: Separator is semicolon ; for User's machine
dv2_formula = "=OFFSET(LISTS!$E$2;MATCH($C$4;Col_Group;0)-1;0;COUNTIF(Col_Group;$C$4);1)"
dv2 = DataValidation(type="list", formula1=dv2_formula, allow_blank=True)
# Suppress error because C4 might be empty initially
dv2.errorTitle = "" 
ws_ui.add_data_validation(dv2)
dv2.add(ws_ui['C5'])

# Results
labels = ["Mã hiệu", "Đơn vị", "Suất vốn đầu tư", "Chi phí Xây dựng", "Chi phí Thiết bị"]
keys   = ["Code", "Unit", "Total", "Const", "Equip"] # Columns in DATA: A, C, D, E, F

# We use SUMIFS or LOOKUP to fetch values based on C4 and C5.
# Data Sheet: Col A (Code) is insufficient because we selected "Group Name" in C4.
# But we have Description in C5. Description is usually unique within a group.
# Let's use Array Formula or INDEX/MATCH with multiple criteria.
# MATCH(1, (ColCode=C4)*(ColDesc=C5), 0) -> Requires Ctrl+Shift+Enter.
# 
# Simpler: Concatenate formatting in DATA sheet.
# We added "LookupKey" in DATA sheet? That was Code_Desc. C4 is GroupName.
# Let's add a clean lookup column in DATA sheet that matches C5 (Description)
# Assumption: Descriptions are unique enough? Or just lookup Description directly.
# Let's try looking up Description in Col B of Data.

for i, lbl in enumerate(labels):
    r = 8 + i
    ws_ui[f'B{r}'] = lbl
    
    # Formula: LOOKUP C5 in DATA!B, return DATA!Column
    # But C5 description might be duplicated across different Groups? (e.g. "1 tang ham")
    # Yes. So we must match Group AND Description.
    # But Group Name in C4 is "Code - Name". DATA sheet only has Code.
    # We need to extract Code from C4. Left(C4, Find(" ", C4)-1)
    
    col_idx = [1, 3, 4, 5, 6][i] # Target Column Index in DATA
    
    # Extract Code Formula: LEFT(C4; FIND(" ";C4)-1)
    # This is getting complex for Excel formula string injection.
    
    # Alternate Plan:
    # Just use SUMIFS for the Values (Total, Const, Equip).
    # Since Description+Code should be unique.
    # SUMIFS(DataCol, DataCodeCol, ExtractedCode, DataDescCol, C5)
    
    if i >= 2: # Values
        f = f"=SUMIFS(DATA!{get_column_letter(col_idx)}:DATA!{get_column_letter(col_idx)}; DATA!$A:$A; LEFT($C$4;FIND(\" \";$C$4)-1); DATA!$B:$B; $C$5)"
    elif i == 0: # Code
        f = f"=LEFT($C$4;FIND(\" \";$C$4)-1)"
    else: # Unit
         # Lookup Unit is hard with SUMIFS (Text). Use INDEX MATCH/MATCH.
         # INDEX(UnitCol, MATCH(1, (CodeCol=Code)*(DescCol=C5), 0)) -> CSE
         # Let's stick to "1000đ/m2" hardcoded or Simplified Lookup?
         # Most have same unit.
         f = "Original Unit (Check DB)" 

    ws_ui[f'C{r}'] = f 
    if i >= 2: ws_ui[f'C{r}'].number_format = '#,##0.00'

print(f"Saving to {output_excel}...")
wb.save(output_excel)
print("Done.")
