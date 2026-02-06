
import openpyxl

file_path = r"g:\My Drive\30_RESOURCES (Kho tai nguyen)\30.01_Tinh hieu qua du an\02_Tra dinh muc theo TT 09 va 12\Tra dinh muc TT09.xlsx"

print(f"Loading file: {file_path}")
wb = openpyxl.load_workbook(file_path, data_only=False)

# 1. Verify TMDT sheet exists
if "TMDT" not in wb.sheetnames:
    print("Error: Sheet 'TMDT' not found. Cannot remap variables.")
    # Fallback: Create a Config sheet
    print("Creating 'Config' sheet as fallback...")
    ws_config = wb.create_sheet("Config")
    target_sheet_name = "Config"
else:
    print("Sheet 'TMDT' found. Using it as target for variables.")
    target_sheet_name = "TMDT"

# 2. Fix Named Ranges
# Map: Name -> Address
fix_map = {
    "i_CPC": "$C$118",
    "i_CapCT": "$C$88",
    "i_BTK": "$C$94",
    "i_tkdutoan": "$C$128",
    "i_VB": "$C$99"
}

print("\n--- Fixing Named Ranges ---")
for name, address in fix_map.items():
    # Remove old definition if exists
    if name in wb.defined_names:
        del wb.defined_names[name]
        print(f"Removed old definition for {name}")
    
    # Create new definition
    new_dest = f"'{target_sheet_name}'!{address}"
    d = openpyxl.workbook.defined_name.DefinedName(name, attr_text=new_dest)
    wb.defined_names[name] = d
    print(f"Created new definition for {name} -> {new_dest}")

# 3. Re-inject Data Lists into TT09 (Hidden columns AA-AF)
ws_tt09 = wb["TT09"]
print("\n--- Re-injecting Data Lists ---")
data_lists = {
    "LoaiCongTrinh": ["Công trình dân dụng", "Công trình công nghiệp", "Công trình giao thông", "Công trình NN&PTNT", "Công trình hạ tầng kỹ thuật"],
    "KieuCongTrinh": ["Công trình xây dựng mới", "Công trình thiết kế mẫu, điển hình", "Công trình sửa chữa, cải tạo, mở rộng"],
    "ThamDinh": ["Cơ quan chuyên môn về xây dựng thẩm định", "Chủ đầu tư tổ chức thẩm định"],
    "KiemToan": ["Thuê kiểm toán độc lập", "Quyết toán nội bộ (Không thuê)"],
    "CapCongTrinh": ["Cấp đặc biệt", "Cấp I", "Cấp II", "Cấp III", "Cấp IV"],
    "SoBuocThietKe": ["Thiết kế 1 bước (BCKTKT)", "Thiết kế 2 bước (Cơ sở + BVTC)", "Thiết kế 3 bước (Cơ sở + KT + BVTC)"]
}

start_col_idx = 27 # AA
start_row = 1000

for i, (key, values) in enumerate(data_lists.items()):
    col_idx = start_col_idx + i
    col_letter = openpyxl.utils.get_column_letter(col_idx)
    
    # Write Header
    ws_tt09.cell(row=start_row, column=col_idx, value=key)
    
    # Write Values
    for j, val in enumerate(values):
        ws_tt09.cell(row=start_row + 1 + j, column=col_idx, value=val)
    print(f"Written {key} list to {col_letter}{start_row}")

# 4. Save
print("Saving file...")
wb.save(file_path)
print("Done.")
