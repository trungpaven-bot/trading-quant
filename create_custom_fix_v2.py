
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Border, Side, PatternFill
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils import quote_sheetname
import re

input_excel = r"g:\My Drive\30_RESOURCES (Kho tai nguyen)\30.01_Tinh hieu qua du an\02_Tra dinh muc theo TT 09 va 12\Du_lieu_Co_Tieu_De_Bang.xlsx"
output_excel = r"g:\My Drive\30_RESOURCES (Kho tai nguyen)\30.01_Tinh hieu qua du an\02_Tra dinh muc theo TT 09 va 12\File_Tra_Cuu_Final_Custom_Fix.xlsx"

print("Reading Data...")
df = pd.read_excel(input_excel, header=None)

processed_data = []

current_table = ""
current_base_prefix = "" 
parent_code = ""

start_processing = False

for idx, row in df.iterrows():
    row_strs = [str(x).strip() if not pd.isna(x) else "" for x in row]
    
    table_name = row_strs[0]
    
    # Filter: Start from Table 1
    if not start_processing:
        if "Bảng 1" in table_name or "Bang 1" in table_name:
            start_processing = True
        else:
            continue 
            
    code = row_strs[1]
    desc = row_strs[2]
    
    # Check data
    vals = row_strs[3:]
    has_number = any(re.search(r'[\d.,]+', v) for v in vals)

    if desc and has_number:
        # Values
        val1, val2, val3 = 0, 0, 0
        nums = []
        for v in vals:
            clean = v.replace('.', '').replace(',', '.')
            try: nums.append(float(clean))
            except: pass
        if len(nums) >= 1: val1 = nums[0]
        if len(nums) >= 2: val2 = nums[1]
        if len(nums) >= 3: val3 = nums[2]
        
        # --- FIX LOGIC ---
        final_desc = desc
        
        # Parent Logic (Define Base)
        if "không có" in desc.lower():
            # Extract Base: "Số tầng <= 5"
            parts = re.split(r'không có', desc, flags=re.IGNORECASE)
            if len(parts) > 0:
                current_base_prefix = parts[0].strip()
            parent_code = code 
        
        # Child Logic (Prepend Base)
        elif desc.lower().startswith("có ") and current_base_prefix:
            final_desc = f"{current_base_prefix} {desc}"
            if not code and parent_code: # Inherit code
                code = parent_code
        
        if not code: code = parent_code
        
        processed_data.append([table_name, code, final_desc, "1000đ/m2", val1, val2, val3])

print(f"Processed {len(processed_data)} lines.")

# --- EXCEL CONSTRUCTION ---
wb = openpyxl.Workbook()

# 1. LISTS SHEET
ws_lists = wb.active
ws_lists.title = "LISTS"

ws_lists['A1'] = "Bảng"
ws_lists['C1'] = "Mã hiệu"
ws_lists['D1'] = "Loại công trình (Đã sửa)"

for i, row in enumerate(processed_data):
    r = i + 2
    ws_lists[f'A{r}'] = row[0] 
    ws_lists[f'C{r}'] = row[1] 
    ws_lists[f'D{r}'] = row[2] 

ws_lists.column_dimensions['A'].width = 40
ws_lists.column_dimensions['C'].width = 15
ws_lists.column_dimensions['D'].width = 60
for cell in ws_lists[1]: cell.font = Font(bold=True)

# 2. ENGINE SHEET
ws_engine = wb.create_sheet("ENGINE")
headers_eng = ["", "Bảng", "Loại bảng", "Mã hiệu", "Loại công trình", "Đơn vị tính", "Suất vốn đầu tư", "Chi phí xây dựng", "Chi phí thiết bị", "Ghép Lists"]
ws_engine.append(headers_eng)

row_offset = 2
for row in processed_data:
    # A=1(Empty), B=2(Table), C=3(Table), D=4(Code), E=5(Desc), F=6(Unit), G=7(Total), H=8(Const), I=9(Equip), J=10(Key)
    ws_engine.cell(row=row_offset, column=2).value = row[0]
    ws_engine.cell(row=row_offset, column=3).value = row[0]
    ws_engine.cell(row=row_offset, column=4).value = row[1]
    ws_engine.cell(row=row_offset, column=5).value = row[2]
    ws_engine.cell(row=row_offset, column=6).value = row[3] # Unit
    ws_engine.cell(row=row_offset, column=7).value = row[4]
    ws_engine.cell(row=row_offset, column=8).value = row[5]
    ws_engine.cell(row=row_offset, column=9).value = row[6]
    
    # Formula J: B & "_" & E
    key_formula = f'=B{row_offset} & "_" & E{row_offset}'
    ws_engine.cell(row=row_offset, column=10).value = key_formula
    
    row_offset += 1

header_row = ws_engine[1]
for cell in header_row:
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="EFEFEF", end_color="EFEFEF", fill_type="solid")
    cell.border = Border(bottom=Side(style='thin'))
ws_engine.column_dimensions['B'].width = 30
ws_engine.column_dimensions['E'].width = 50
ws_engine.column_dimensions['J'].width = 40

# 3. TRA_CUU SHEET
ws_ui = wb.create_sheet("TRA_CUU")
ws_ui['B2'] = "TRA CỨU SUẤT VỐN (FINAL 2024)"
ws_ui['B2'].font = Font(size=14, bold=True)

# L1: TABLE
ws_ui['B4'] = "1. Chọn Bảng:"
ws_ui['C4'].border = Border(bottom=Side(style='thin'))

tables = sorted(list(set(row[0] for row in processed_data)))
ws_lists['Z1'] = "UniqueTables"
for i, t in enumerate(tables): ws_lists[f'Z{i+2}'] = t
wb.defined_names.add(DefinedName("ListTables", attr_text=f"LISTS!$Z$2:$Z${len(tables)+1}"))

dv1 = DataValidation(type="list", formula1="=ListTables", allow_blank=True)
ws_ui.add_data_validation(dv1)
dv1.add(ws_ui['C4'])

# L2: DESC
ws_ui['B5'] = "2. Chọn Công trình:"
ws_ui['C5'].border = Border(bottom=Side(style='thin'))
ws_ui['C5'].alignment = Alignment(wrap_text=True)

count = len(processed_data)
# Add Named Ranges for OFFSET
wb.defined_names.add(DefinedName("ColTableList", attr_text=f"LISTS!$A$2:$A${count+1}"))
wb.defined_names.add(DefinedName("ColDescList", attr_text=f"LISTS!$D$2:$D${count+1}"))

# Formula List 2
dv2_f = "=OFFSET(LISTS!$D$2;MATCH($C$4;ColTableList;0)-1;0;COUNTIF(ColTableList;$C$4);1)"
dv2 = DataValidation(type="list", formula1=dv2_f, allow_blank=True)
dv2.errorTitle=""
ws_ui.add_data_validation(dv2)
dv2.add(ws_ui['C5'])

# RESULT FORMULAS
ws_ui['B8'] = "KẾT QUẢ:"
res_labels = ["Mã hiệu", "Đơn vị tính", "Suất vốn đầu tư", "Chi phí Xây dựng", "Chi phí Thiết bị"]
# Map to ENGINE columns: Code(D)=4, Unit(F)=6, Total(G)=7, Const(H)=8, Equip(I)=9
map_cols = ["D", "F", "G", "H", "I"]

for i, lbl in enumerate(res_labels):
    r = 9 + i
    ws_ui[f'B{r}'] = lbl
    ws_ui[f'B{r}'].font = Font(bold=True)
    col_ltr = map_cols[i]
    f = f"=INDEX(ENGINE!{col_ltr}:{col_ltr}; MATCH($C$4 & \"_\" & $C$5; ENGINE!$J:$J; 0))"
    if i >= 2: ws_ui[f'C{r}'].number_format = '#,##0'
    ws_ui[f'C{r}'] = f

print(f"Saving to {output_excel}...")
wb.save(output_excel)
print("Done.")
