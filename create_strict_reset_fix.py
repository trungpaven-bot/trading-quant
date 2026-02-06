
import pandas as pd
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.workbook.defined_name import DefinedName
import re

input_excel = r"g:\My Drive\30_RESOURCES (Kho tai nguyen)\30.01_Tinh hieu qua du an\02_Tra dinh muc theo TT 09 va 12\Du_lieu_Co_Tieu_De_Bang.xlsx"
output_excel = r"g:\My Drive\30_RESOURCES (Kho tai nguyen)\30.01_Tinh hieu qua du an\02_Tra dinh muc theo TT 09 va 12\File_Tra_Cuu_Final_StrictReset.xlsx"

print("Reading Data...")
df = pd.read_excel(input_excel, header=None)

processed_data = []

# State variables
current_header_stack = [] 
current_table_name = ""
last_valid_code = ""

start_processing = False

for idx, row in df.iterrows():
    row_strs = [str(x).strip() if not pd.isna(x) else "" for x in row]
    table_name = row_strs[0]
    
    # 1. STRICT TABLE RESET
    # If table name changes compared to stored state -> FLUSH EVERYTHING
    if table_name != current_table_name:
        current_table_name = table_name
        current_header_stack = [] # Wipe memory cleanly
        last_valid_code = ""
    
    if not start_processing:
        if "Bảng 1" in table_name or "Bang 1" in table_name:
            start_processing = True
        else:
            continue
            
    code = row_strs[1]
    desc = row_strs[2]
    
    vals = row_strs[3:]
    has_number = any(re.search(r'[\d.,]+', v) for v in vals)

    if has_number and desc:
        # DATA ROW
        val1, val2, val3 = 0, 0, 0
        nums = []
        for v in vals:
            clean = v.replace('.', '').replace(',', '.')
            try: nums.append(float(clean))
            except: pass
        if len(nums) >= 1: val1 = nums[0]
        if len(nums) >= 2: val2 = nums[1]
        if len(nums) >= 3: val3 = nums[2]
        
        # Combine Parent headers with current Desc
        if current_header_stack:
            # Join with space or dash? Space is safer.
            prefix = " ".join(current_header_stack)
            final_desc = f"{prefix} {desc}"
        else:
            final_desc = desc
            
        if code:
            last_valid_code = code
            final_code = code
        else:
            final_code = last_valid_code
            
        processed_data.append([table_name, final_code, final_desc, "1000đ/Dv", val1, val2, val3])
        
    else:
        # HEADER ROW (No money)
        if desc:
            # Decision Logic:
            # 1. Is this a NEW Top-Level Header (Major)?
            #    Major headers typically have a CODE (like 13480.01 or similar structure, though not always populated in Word extraction)
            #    OR if it's the very first header of the table.
            
            # Logic Update for Table 65:
            # Row 1: "Công trình đài..." -> Pushed to Stack[0]
            # Row 2: "Cột anten... 30m" -> Pushed to Stack[1]
            # Row Data: "20W" -> Uses Stack[0]+[1]
            
            # Row X: "Cột anten... 45m" -> Should replace Stack[1]
            
            if code:
                # Code implies a strong reset or a major block
                current_header_stack = [desc]
                last_valid_code = code
            else:
                # No code. It's a middleware header.
                # If stack is empty, just add it.
                if len(current_header_stack) == 0:
                    current_header_stack.append(desc)
                else:
                    # Stack has content. Does this new line REPLACE the last one or ADD to it?
                    # Generally in these tables derived from Word:
                    # High level indentation is lost.
                    # Assumption: Maximum depth is usually 2 or 3.
                    
                    # Safer heuristic:
                    # If the last item in stack is very long/complex, maybe we are drilling down.
                    # But usually, subsequent headers replace the "Detail" level.
                    
                    # Let's enforce a max stack depth of 2 for safety?
                    # Stack[0] = Broad Category (Cong trinh dai tram...)
                    # Stack[1] = Specific Category (Cot anten 30m)
                    
                    # If we encounter a new header, we assume it's at the same level as the LAST one added.
                    # So we replace the last element.
                    current_header_stack[-1] = desc
                    
                    # Wait, what if it was meant to be a 3rd level?
                    # (e.g. Cong Trinh -> Nha A -> Tang 1)
                    # If we replace, we get "Cong Trinh -> Tang 1". We lose "Nha A".
                    # This is tricky without indentation data.
                    
                    # Let's look at the Table 65 example again.
                    # "Công trình đài trạm..." (Line 1)
                    # "Cột anten ... 30m" (Line 2)
                    # Use both.
                    
                    # "Cột anten ... 45m" (Line X)
                    # This clearly replaces "Cột anten ... 30m".
                    
                    # So "Replace Default" seems correct for Subsequent headers.
                    # BUT how do we get the initial stack of [0, 1] without replacing 0 with 1?
                    # We need to know they are sequential WITHOUT data in between.
                    
                    # Refined Logic:
                    # If we hit Header A, then Header B, WITHOUT any Data Row in between -> They are adding up (Drilling down).
                    # If we hit Header C, AFTER seeing some Data Rows -> It replaces the last header context.
                    pass

# RE-RUN LOOP WITH THE "DATA SEEN" FLAG LOGIC
processed_data = [] # Reset output
current_header_stack = []
current_table_name = ""
last_valid_code = ""

data_seen_since_last_header = False

start_processing = False

for idx, row in df.iterrows():
    row_strs = [str(x).strip() if not pd.isna(x) else "" for x in row]
    table_name = row_strs[0]
    
    # 1. STRICT TABLE RESET
    if table_name != current_table_name:
        current_table_name = table_name
        current_header_stack = [] 
        last_valid_code = ""
        data_seen_since_last_header = False 
    
    if not start_processing:
        if "Bảng 1" in table_name or "Bang 1" in table_name:
            start_processing = True
        else:
            continue
            
    code = row_strs[1]
    desc = row_strs[2]
    vals = row_strs[3:]
    has_number = any(re.search(r'[\d.,]+', v) for v in vals)

    if has_number and desc:
        # --- DATA ROW ---
        data_seen_since_last_header = True
        
        # Parse numbers
        val1, val2, val3 = 0, 0, 0
        nums = []
        for v in vals:
            try: nums.append(float(v.replace('.', '').replace(',', '.')))
            except: pass
        if len(nums) >= 1: val1 = nums[0]
        if len(nums) >= 2: val2 = nums[1]
        if len(nums) >= 3: val3 = nums[2]
        
        # Construct Name
        prefix = " ".join(current_header_stack) if current_header_stack else ""
        final_desc = f"{prefix} {desc}" if prefix else desc
        
        if code: last_valid_code = code; final_code = code
        else: final_code = last_valid_code
        
        processed_data.append([table_name, final_code, final_desc, "1000đ/Dv", val1, val2, val3])
        
    else:
        # --- HEADER ROW ---
        if desc:
            if code:
                # Major header with code -> Reset stack to just this
                current_header_stack = [desc]
                last_valid_code = code
                data_seen_since_last_header = False
            else:
                # Minor header
                if data_seen_since_last_header:
                    # Case: We processed some data, now we see a new header.
                    # This implies a change in context at the LOWEST level.
                    # Replace the last stack item.
                    if current_header_stack:
                        current_header_stack[-1] = desc
                    else:
                        current_header_stack.append(desc)
                    data_seen_since_last_header = False
                else:
                    # Case: We saw a header immediately after another header (Drill down).
                    # Append it.
                    current_header_stack.append(desc)
                    data_seen_since_last_header = False


# --- BUILD EXCEL ---
wb = openpyxl.Workbook()

# LISTS
ws_lists = wb.active; ws_lists.title="LISTS"
ws_lists['A1']="Bảng"; ws_lists['C1']="Mã"; ws_lists['D1']="Tên CT"
for i, r in enumerate(processed_data):
    row=i+2; ws_lists[f'A{row}']=r[0]; ws_lists[f'C{row}']=r[1]; ws_lists[f'D{row}']=r[2]

# Unique Tables and Hardcoded Ranges
unique_tables = []
seen = set()
for r in processed_data:
    if r[0] not in seen: seen.add(r[0]); unique_tables.append(r[0])
for i, t in enumerate(unique_tables): ws_lists[f'Z{i+2}'] = t

cnt_data = len(processed_data)
cnt_tbl = len(unique_tables)
wb.defined_names.add(DefinedName("ListTables", attr_text=f"LISTS!$Z$2:$Z${cnt_tbl+1}"))
wb.defined_names.add(DefinedName("ColTbl", attr_text=f"LISTS!$A$2:$A${cnt_data+1}"))
wb.defined_names.add(DefinedName("ColDesc", attr_text=f"LISTS!$D$2:$D${cnt_data+1}"))

# ENGINE
ws_eng = wb.create_sheet("ENGINE")
ws_eng.append(["", "Bảng", "Bảng2", "Mã", "Tên", "ĐV", "Suất", "XD", "TB", "Key"])
for i, r in enumerate(processed_data):
    row=i+2
    ws_eng.cell(row,2,r[0]); ws_eng.cell(row,3,r[0]); ws_eng.cell(row,4,r[1]); ws_eng.cell(row,5,r[2])
    ws_eng.cell(row,6,r[3]); ws_eng.cell(row,7,r[4]); ws_eng.cell(row,8,r[5]); ws_eng.cell(row,9,r[6])
    ws_eng.cell(row,10).value = f'=B{row} & "_" & E{row}'

# TRA_CUU
ws_ui = wb.create_sheet("TRA_CUU")
ws_ui['B2']="TRA CỨU (STRICT RESET)"; ws_ui['B2'].font=Font(size=14, bold=True)
ws_ui['B4']="1. Chọn Bảng:"; ws_ui['C4'].border=Border(bottom=Side(style='thin')); 
dv1 = DataValidation(type="list", formula1="=ListTables", allow_blank=True); ws_ui.add_data_validation(dv1); dv1.add(ws_ui['C4'])
ws_ui['B5']="2. Chọn CT:"; ws_ui['C5'].border=Border(bottom=Side(style='thin')); ws_ui['C5'].alignment=Alignment(wrap_text=True)
dv2_f = "=OFFSET(LISTS!$D$2;MATCH($C$4;ColTbl;0)-1;0;COUNTIF(ColTbl;$C$4);1)"
dv2 = DataValidation(type="list", formula1=dv2_f, allow_blank=True); ws_ui.add_data_validation(dv2); dv2.add(ws_ui['C5'])

# Results
ws_ui['B8']="KẾT QUẢ:"; labels=["Mã","ĐV","Suất","XD","TB"]; cols=["D","F","G","H","I"]
for i,l in enumerate(labels):
    r=9+i; ws_ui[f'B{r}']=l; ws_ui[f'B{r}'].font=Font(bold=True)
    f=f"=INDEX(ENGINE!{cols[i]}:{cols[i]}; MATCH($C$4 & \"_\" & $C$5; ENGINE!$J:$J; 0))"
    if i>=2: ws_ui[f'C{r}'].number_format='#,##0'
    ws_ui[f'C{r}']=f

print(f"Saving {output_excel}...")
wb.save(output_excel)
print("Done.")
