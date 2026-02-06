
import openpyxl

file_path = r"g:\My Drive\00_INBOX (xu ly nhanh)\tra TMDT.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=False)

print("--- Sheets ---")
print(wb.sheetnames)

print("\n--- Defined Names (Named Ranges) ---")
# Print all named ranges and their destinations
for name, defn in wb.defined_names.items():
    # defined_names might return a dict or object depending on version, let's try to stringify
    try:
        # destinations is a generator of (worksheet_title, coord)
        dests = list(defn.destinations)
        print(f"Name: {name} -> {dests}")
    except Exception as e:
        print(f"Name: {name} -> (Error reading dest: {e})")

print("\n--- Data Validation Check in TT09 ---")
ws = wb["TT09"]
# Iterate over validations
try:
    for dv in ws.data_validations.dataValidation:
        print(f"Validation range: {dv.sqref}")
        print(f"Type: {dv.type}")
        print(f"Formula1: {dv.formula1}")
        print(f"Formula2: {dv.formula2}")
except Exception as e:
    print(f"No standard data validation found or error: {e}")

print("\n--- Checking potential source data areas ---")
# Sometimes lists are stored in hidden columns or far to the right/bottom
print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")
