
import openpyxl

file_path = r"g:\My Drive\30_RESOURCES (Kho tai nguyen)\30.01_Tinh hieu qua du an\02_Tra dinh muc theo TT 09 va 12\Tra dinh muc TT09.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=False)
ws = wb["TT09"]

print("--- Searching for i_tkdutoan Usage ANYWHERE ---")
# The previous search failed, let's try bruteforce again but print context
found = False
for r in range(1, 1200):
    for c in range(1, 20):
        val = str(ws.cell(row=r, column=c).value) 
        if "i_tkdutoan" in val or "$C$128" in val:
            print(f"FOUND at {ws.cell(row=r, column=c).coordinate}: {val}")
            found = True

if not found:
    print("i_tkdutoan not used in TT09 formulas. Checking TMDT formulas again.")
    # Check TMDT for direct usage
    ws_tmdt = wb["TMDT"]
    for r in range(1, 200):
         for c in range(1, 10):
             val = str(ws_tmdt.cell(row=r, column=c).value)
             if "$C$128" in val or "i_tkdutoan" in val:
                 print(f"FOUND in TMDT at {ws_tmdt.cell(row=r, column=c).coordinate}: {val}")
