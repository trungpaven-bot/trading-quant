
import pandas as pd
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import re

input_excel = r"g:\My Drive\30_RESOURCES (Kho tai nguyen)\30.01_Tinh hieu qua du an\02_Tra dinh muc theo TT 09 va 12\Du_lieu_Suat_von_dau_tu_2024.xlsx"
output_excel = r"g:\My Drive\30_RESOURCES (Kho tai nguyen)\30.01_Tinh hieu qua du an\02_Tra dinh muc theo TT 09 va 12\File_Tra_Cuu_Chuyen_Nghiep.xlsx"

print("Reading and Refining Data...")
df = pd.read_excel(input_excel, header=None)

refined_data = []
current_code = ""

for idx, row in df.iterrows():
    row_strs = [str(x).strip() if not pd.isna(x) else "" for x in row]
    if "--- BẢNG" in row_strs[0]: continue
    
    code = row_strs[0]
    desc = row_strs[1]
    
    # Fill Code Down Logic
    if code:
        current_code = code
    
    vals = row_strs[2:6]
    has_number = any(re.search(r'[\d.,]+', v) for v in vals)
    
    if desc and has_number:
        unit = "1000đ/m2/đơn vị"
        val1, val2, val3 = 0, 0, 0
        
        nums = []
        for v in vals:
            # Parse number robustly
            v_clean = v.replace('.', '').replace(',', '.') # Assume VN format 1.000,00 -> 1000.00?
            # Wait, usually Suat Von Dau Tu is like 7.850 (7 million). 
            # If format is 7.850 -> It's float 7.85 or 7850? 
            # Usually input is text. Let's keep text for Excel to parse, or standarize to US float.
            # Let's keep original text string for safety in LOOKUP, or clean for SUM?
            # SUMIFS needs numbers.
            # Let's try to convert to float.
            try:
                # Remove thousand separators (.) and replace decimal (,) if any.
                # Actually in VN: 1.000.000,00
                # But in that Word doc, it looked like "7.850" (7 trieu 8 tram).
                # Which is 7850 k dong.
                # So just remove dots? Or replace dot with empty?
                # If "7.850", float(7.850) = 7.85. If it means 7850, we need to multiply?
                # The unit is "1000d/m2". 7.850 means 7,850,000.
                # Let's simply remove dots to get integer if it looks like coordinate.
                # But 9.177 -> 9177.
                clean = v.replace('.', '').replace(',', '.')
                nums.append(float(clean))
            except:
                pass
                
        if len(nums) >= 1: val1 = nums[0]
        if len(nums) >= 2: val2 = nums[1]
        if len(nums) >= 3: val3 = nums[2]
        
        refined_data.append([current_code, desc, unit, val1, val2, val3])

print(f"Refined {len(refined_data)} rows.")

groups = {}
group_names = {}

for row in refined_data:
    code = row[0]
    desc = row[1]
    if code not in groups:
        groups[code] = []
        group_names[code] = f"Nhóm {code}" 
    groups[code].append(row)

for code, items in groups.items():
    snippet = items[0][1][:40].replace("\"", "").strip() + "..."
    group_names[code] = f"{code} - {snippet}"

wb = openpyxl.Workbook()

# DATA Sheet
ws_data = wb.active
ws_data.title = "DATA"
ws_data.append(["Code", "Description", "Unit", "Total", "Const", "Equip"])
for code, items in groups.items():
    for item in items:
        ws_data.append(item) # Item has 6 elements

# LISTS Sheet
ws_lists = wb.create_sheet("LISTS")
ws_lists['A1'] = "Danh sách Mã"
code_list = list(groups.keys())
for i, code in enumerate(code_list):
    ws_lists[f'A{i+2}'] = group_names[code] 
    ws_lists[f'B{i+2}'] = code 

wb.create_named_range("List_Codes", ws_lists, f"$A$2:$A${len(code_list)+1}")

ws_lists['D1'] = "CodeCol"
ws_lists['E1'] = "DescCol"
r = 2
for code in code_list:
    disp_name = group_names[code]
    for item in groups[code]:
        ws_lists[f'D{r}'] = disp_name 
        ws_lists[f'E{r}'] = item[1]   
        r += 1

max_r = r - 1
wb.create_named_range("Col_Group", ws_lists, f"$D$2:$D${max_r}")
wb.create_named_range("Col_Desc", ws_lists, f"$E$2:$E${max_r}") # Actually unused in OFFSET, we calculate range

# --- TRA CUU Sheet ---
ws_ui = wb.create_sheet("TRA_CUU")
ws_ui.column_dimensions['B'].width = 25
ws_ui.column_dimensions['C'].width = 60

ws_ui['B2'] = "TRA CỨU SUẤT VỐN ĐẦU TƯ"
ws_ui['B2'].font = Font(size=14, bold=True)

ws_ui['B4'] = "1. Chọn Nhóm:"
ws_ui['C4'].border = Border(bottom=Side(style='thin'))
dv1 = DataValidation(type="list", formula1="=List_Codes", allow_blank=True)
ws_ui.add_data_validation(dv1)
dv1.add(ws_ui['C4'])

ws_ui['B5'] = "2. Chọn Chi tiết:"
ws_ui['C5'].border = Border(bottom=Side(style='thin'))

# Fix formula string escaping
dv2_formula = "=OFFSET(LISTS!$E$2;MATCH($C$4;Col_Group;0)-1;0;COUNTIF(Col_Group;$C$4);1)"
dv2 = DataValidation(type="list", formula1=dv2_formula, allow_blank=True)
dv2.errorTitle = "" 
ws_ui.add_data_validation(dv2)
dv2.add(ws_ui['C5'])

labels = ["Mã hiệu", "Đơn vị", "Suất vốn đầu tư", "Chi phí Xây dựng", "Chi phí Thiết bị"]
# Map labels to DATA columns: Code(A)=1, Unit(C)=3, Total(D)=4, Const(E)=5, Equip(F)=6
col_indices = [1, 3, 4, 5, 6]

for i, lbl in enumerate(labels):
    r = 8 + i
    ws_ui[f'B{r}'] = lbl
    ws_ui[f'B{r}'].font = Font(bold=True)
    
    col_idx = col_indices[i]
    col_letter = get_column_letter(col_idx) # A, C, D...
    
    # Formula explanation:
    # DATA!A:A is Code. C4 is "Code - Name". Extraction: LEFT(C4, FIND(" ", C4)-1)
    # DATA!B:B is Description. C5 is Description.
    
    if i == 0: # Code display
        f = f"=LEFT($C$4;FIND(\" \";$C$4)-1)"
    elif i == 1: # Unit
        # Hard to lookup text with SUMIFS. Use generic.
        f = "1000đ/m2 sàn (hoặc theo ĐV)"
    else: # Values (Sumifs)
        # SUMIFS(ValuesCol, CodeCol, CodeExtracted, DescCol, DescSelected)
        f = f"=SUMIFS(DATA!{col_letter}:{col_letter}; DATA!$A:$A; LEFT($C$4;FIND(\" \";$C$4)-1); DATA!$B:$B; $C$5)"
        ws_ui[f'C{r}'].number_format = '#,##0' # Integer format for VND

    ws_ui[f'C{r}'] = f 

print(f"Saving to {output_excel}...")
wb.save(output_excel)
print("Done.")
