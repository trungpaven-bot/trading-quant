
import pandas as pd
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.formatting.rule import FormulaRule
from openpyxl.workbook.defined_name import DefinedName
import re

# Source data
input_excel = r"g:\My Drive\30_RESOURCES (Kho tai nguyen)\30.01_Tinh hieu qua du an\02_Tra dinh muc theo TT 09 va 12\Du_lieu_Co_Tieu_De_Bang.xlsx"
output_excel = r"g:\My Drive\30_RESOURCES (Kho tai nguyen)\30.01_Tinh hieu qua du an\02_Tra dinh muc theo TT 09 va 12\File_Tra_Cuu_Highlight_TT09.xlsx"

print("Reading Data...")
df = pd.read_excel(input_excel, header=None)

# Clean Data
data_rows = []
for idx, row in df.iterrows():
    r = [str(x).strip() if not pd.isna(x) else "" for x in row]
    # table, code, desc...
    # Ensure numeric columns 3,4,5
    vals = r[3:6]
    nums = []
    for v in vals:
        try: nums.append(float(v.replace('.','').replace(',','.')))
        except: nums.append(0)
    
    # Structure for Visual Table: 
    # [Table(HiddenHelper), Code, Desc, Unit, Total, Const, Equip]
    data_rows.append([r[0], r[1], r[2], "1000đ/m2", nums[0] if len(nums)>0 else 0, nums[1] if len(nums)>1 else 0, nums[2] if len(nums)>2 else 0])

wb = openpyxl.Workbook()

# --- SHEET: TRA_CUU (The Main Interface) ---
ws = wb.active
ws.title = "TRA_CUU"

# 1. INPUT SECTION (Top)
ws.row_dimensions[1].height = 20
ws.merge_cells('B2:F2')
ws['B2'] = "TRA CỨU SUẤT VỐN ĐẦU TƯ (HIGHLIGHT)"
ws['B2'].font = Font(size=14, bold=True, color="FF0000")
ws['B2'].alignment = Alignment(horizontal='center')

# Controls
ws['B4'] = "Chọn Bảng:"
ws['C4'].border = Border(bottom=Side(style='thin'))

ws['B5'] = "Chọn Công trình:"
ws['C5'].border = Border(bottom=Side(style='thin'))
ws['C5'].alignment = Alignment(wrap_text=True)

# Result Box (Simulating the user's screenshot result area)
ws['E4'] = "KẾT QUẢ ĐÃ CHỌN:"
ws['E4'].font = Font(bold=True)

res_labels = ["Mã hiệu", "Đơn vị", "Suất vốn", "CP Xây dựng", "CP Thiết bị"]
for i, lbl in enumerate(res_labels):
    r = 5 + i
    ws[f'E{r}'] = lbl
    ws[f'F{r}'].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    ws[f'F{r}'].font = Font(bold=True) # The result values

# 2. DATA TABLE SECTION (Below)
start_row = 15
ws['B14'] = "BẢNG DỮ LIỆU TRA CỨU"
ws['B14'].font = Font(bold=True)

headers = ["Bảng (Ẩn)", "Mã hiệu", "Loại công trình", "Đơn vị", "Suất vốn đầu tư", "Chi phí Xây dựng", "CP Thiết bị"]
ws.append([]) # Spacer
ws.append([]) # spacer
ws.append([]) # spacer... actually let's write to specific cells
# Write Headers at Row 15
for col_idx, h in enumerate(headers):
    cell = ws.cell(row=start_row, column=col_idx+1) # A15, B15...
    cell.value = h
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")

# Write Data
for i, row_data in enumerate(data_rows):
    r = start_row + 1 + i
    for j, val in enumerate(row_data):
        cell = ws.cell(row=r, column=j+1)
        cell.value = val
        cell.border = Border(bottom=Side(style='dotted'))
        if j >= 4: # Number format
            cell.number_format = '#,##0'

max_row = start_row + len(data_rows)
table_range = f"$A${start_row+1}:$G${max_row}"

# 3. HIGHLIGHT LOGIC (Conditional Formatting)
# Formula: If (Col A = C4) AND (Col C = C5) -> Fill Grey
# Table Col A is Hidden Helper. Col C is Desc.
gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
rule = FormulaRule(formula=[f'AND($A{start_row+1}=$C$4, $C{start_row+1}=$C$5)'], stopIfTrue=True, fill=gray_fill)
ws.conditional_formatting.add(table_range, rule)

# 4. DATA VALIDATION (Try simple approach)
# Lists Sheet
ws_list = wb.create_sheet("LISTS")
ws_list.sheet_state = 'hidden'

# Table List
tables = sorted(list(set(r[0] for r in data_rows)))
ws_list['A1']="Tables"
for i, t in enumerate(tables): ws_list[f'A{i+2}'] = t 
wb.defined_names.add(DefinedName("ListTables", attr_text=f"LISTS!$A$2:$A${len(tables)+1}"))

dv1 = DataValidation(type="list", formula1="=ListTables", allow_blank=True)
ws.add_data_validation(dv1)
dv1.add(ws['C4'])

# Desc Helper (Table | Desc)
ws_list['C1']="Tbl"; ws_list['D1']="Desc"
sorted_rows = sorted(data_rows, key=lambda x: x[0])
for i, r in enumerate(sorted_rows):
    ws_list[f'C{i+2}'] = r[0]
    ws_list[f'D{i+2}'] = r[2]

max_l = len(sorted_rows)+1
wb.defined_names.add(DefinedName("ColTbl", attr_text=f"LISTS!$C$2:$C${max_l}"))
# Dependent DV
dv2_f = "=OFFSET(LISTS!$D$2;MATCH($C$4;ColTbl;0)-1;0;COUNTIF(ColTbl;$C$4);1)"
dv2 = DataValidation(type="list", formula1=dv2_f, allow_blank=True)
dv2.errorTitle=""
ws.add_data_validation(dv2)
dv2.add(ws['C5'])

# 5. LINK RESULTS TO HIGHLIGHTED ROW
# We use SUMIFS because it filters based on C4 and C5
# Result Cells: F5 (Code), F6 (Unit), F7 (Total), F8 (Const), F9 (Equip)
# Code is text, SUMIFS won't work. Use LOOKUP logic.
# XLOOKUP is best but requires newer Excel.
# Safe bet: INDEX/MATCH with Helper or Array formula. 
# We already have A (Table) and C (Desc) in Data Table range.
# Let's use LOOKUP(2, 1/(...)) pattern which is robust.
# Formula: =LOOKUP(2, 1/(($A$16:$A$1000=$C$4)*($C$16:$C$1000=$C$5)), $B$16:$B$1000)

limit = max_row
# F5: Code (Col B)
ws['F5'] = f"=LOOKUP(2, 1/(($A${start_row+1}:$A${limit}=$C$4)*($C${start_row+1}:$C${limit}=$C$5)), $B${start_row+1}:$B${limit})"
# F6: Unit (Col D)
ws['F6'] = f"=LOOKUP(2, 1/(($A${start_row+1}:$A${limit}=$C$4)*($C${start_row+1}:$C${limit}=$C$5)), $D${start_row+1}:$D${limit})"
# F7: Total (Col E)
ws['F7'] = f"=LOOKUP(2, 1/(($A${start_row+1}:$A${limit}=$C$4)*($C${start_row+1}:$C${limit}=$C$5)), $E${start_row+1}:$E${limit})"
ws['F7'].number_format = '#,##0'
# F8: Const (Col F)
ws['F8'] = f"=LOOKUP(2, 1/(($A${start_row+1}:$A${limit}=$C$4)*($C${start_row+1}:$C${limit}=$C$5)), $F${start_row+1}:$F${limit})"
ws['F8'].number_format = '#,##0'
# F9: Equip (Col G)
ws['F9'] = f"=LOOKUP(2, 1/(($A${start_row+1}:$A${limit}=$C$4)*($C${start_row+1}:$C${limit}=$C$5)), $G${start_row+1}:$G${limit})"
ws['F9'].number_format = '#,##0'

# Formatting Widths
ws.column_dimensions['A'].hidden = True # Hide Table Name col in Data View
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 50
ws.column_dimensions['E'].width = 20
ws.column_dimensions['F'].width = 20

print(f"Saving to {output_excel}...")
wb.save(output_excel)
print("Done.")
