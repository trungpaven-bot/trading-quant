
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import copy

file_path = r'g:\My Drive\30_RESOURCES (Kho tai nguyen)\30.01_Tinh hieu qua du an\02_Tra dinh muc theo TT 09 va 12\FS CCN Phan Dinh Phung 1.xlsx'
output_path = r'g:\My Drive\30_RESOURCES (Kho tai nguyen)\30.01_Tinh hieu qua du an\02_Tra dinh muc theo TT 09 va 12\FS CCN Phan Dinh Phung 1_Refined.xlsx'

try:
    wb = openpyxl.load_workbook(file_path)
    if 'PA 3 (ok)' not in wb.sheetnames:
        print("Sheet 'PA 3 (ok)' not found!")
        exit()
    
    ws_source = wb['PA 3 (ok)']
    ws = wb.copy_worksheet(ws_source)
    ws.title = "PA 3 (Refined)"
    
    # --- 1. Fix Table Numbering ---
    # Row 61 (index 61 in 1-based openpyxl?) -> Dump says Row 61 (0-based) is line 62. So Excel Row 62.
    # Let's search for the cell containing "Bảng 1.2 : Bảng tính doanh thu"
    
    revenue_title_cell = None
    for row in ws.iter_rows(min_row=50, max_row=70, min_col=1, max_col=5):
        for cell in row:
            if cell.value and isinstance(cell.value, str) and "Bảng tính doanh thu" in cell.value:
                revenue_title_cell = cell
                break
        if revenue_title_cell: break
    
    if revenue_title_cell:
        print(f"Renaming Revenue Table at {revenue_title_cell.coordinate}")
        revenue_title_cell.value = revenue_title_cell.value.replace("Bảng  1.2", "Bảng 1.3").replace("Bảng 1.2", "Bảng 1.3")

    # --- 2. Re-organize Inputs (Move Right Block to Bottom) ---
    # The block "Nhóm thông tin về chi phí hoạt động hàng năm" is approx F4/F5 to H17/H20
    # Let's find the header
    op_cost_header = None
    for row in ws.iter_rows(min_row=1, max_row=10, min_col=5, max_col=15):
        for cell in row:
            if cell.value and isinstance(cell.value, str) and "chi phí hoạt động" in cell.value:
                op_cost_header = cell
                break
        if op_cost_header: break
        
    if op_cost_header:
        print(f"Found Operating Cost Input at {op_cost_header.coordinate}")
        # Assuming block size approx 15 rows x 5 columns
        min_r = op_cost_header.row
        min_c = op_cost_header.column
        max_r = min_r + 15 
        max_c = min_c + 5 
        
        # Insert rows after the "Capital Structure" section.
        # Capital structure ends around row 20.
        insert_pos = 22
        ws.insert_rows(insert_pos, amount=18)
        
        # Add a section header
        ws.cell(row=insert_pos, column=2).value = "IV. CHI PHÍ HOẠT ĐỘNG KHÁC (Di chuyển từ cột bên phải)"
        ws.cell(row=insert_pos, column=2).font = Font(bold=True, color="FF0000")
        
        # Move the range
        # target is B(insert_pos+2)
        target_row = insert_pos + 2
        
        # We move cell by cell to be safe, or use move_range
        # move_range string format: "F4:H20"
        # Calculate range string
        from openpyxl.utils import get_column_letter
        range_str = f"{get_column_letter(min_c)}{min_r}:{get_column_letter(max_c)}{max_r}"
        
        # Offset: Target B (Col 2) - Current Min Col
        col_offset = 2 - min_c
        row_offset = target_row - min_r
        
        print(f"Moving {range_str} by R:{row_offset} C:{col_offset}")
        ws.move_range(range_str, rows=row_offset, cols=col_offset, translate=True)
        
    # --- 3. Formatting Improvements ---
    # Make headers bold
    header_font = Font(bold=True, size=11)
    for row in ws.iter_rows(min_row=1, max_row=150):
        cell = row[1] # Column B
        if cell.value and isinstance(cell.value, str) and (cell.value.startswith("Bảng") or cell.value.isupper()):
            cell.font = header_font
            
    # Auto-fit Column B (approx)
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 15

    wb.save(output_path)
    print(f"Saved refined file to {output_path}")

except Exception as e:
    print(f"Error: {e}")
