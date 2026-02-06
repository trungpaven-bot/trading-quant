
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import re

input_excel = r"g:\My Drive\30_RESOURCES (Kho tai nguyen)\30.01_Tinh hieu qua du an\02_Tra dinh muc theo TT 09 va 12\Du_lieu_Suat_von_dau_tu_2024.xlsx"
output_excel = r"g:\My Drive\30_RESOURCES (Kho tai nguyen)\30.01_Tinh hieu qua du an\02_Tra dinh muc theo TT 09 va 12\File_Tra_Cuu_Simple.xlsx"

print("Reading Data...")
df = pd.read_excel(input_excel, header=None)

# 1. CLEAN DATA
refined_data = []
current_code = ""

for idx, row in df.iterrows():
    row_strs = [str(x).strip() if not pd.isna(x) else "" for x in row]
    if "--- BẢNG" in row_strs[0]: continue
    
    code = row_strs[0]
    desc = row_strs[1]
    
    if code: current_code = code 
    vals = row_strs[2:6]
    has_number = any(re.search(r'[\d.,]+', v) for v in vals)
    
    if desc and has_number:
        unit = "1000đ/m2/đơn vị"
        val1, val2, val3 = 0, 0, 0
        nums = []
        for v in vals:
            clean = v.replace('.', '').replace(',', '.')
            try: nums.append(float(clean))
            except: pass
        if len(nums) >= 1: val1 = nums[0]
        if len(nums) >= 2: val2 = nums[1]
        if len(nums) >= 3: val3 = nums[2]
        refined_data.append([current_code, desc, unit, val1, val2, val3])

# 2. SETUP WORKBOOK
wb = openpyxl.Workbook()

# --- SHEET: DATA ---
ws_data = wb.active
ws_data.title = "DATA"
headers = ["Mã hiệu", "Loại công trình", "Đơn vị tính", "Suất vốn đầu tư", "Chi phí Xây dựng", "Chi phí Thiết bị"]
ws_data.append(headers)
for row in refined_data:
    ws_data.append(row)

# Styling
header_font = Font(bold=True)
header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
for cell in ws_data[1]: 
    cell.font = header_font
    cell.fill = header_fill
ws_data.column_dimensions['B'].width = 50

# --- SHEET: TRA_CUU ---
ws_ui = wb.create_sheet("TRA_CUU")
ws_ui.column_dimensions['B'].width = 20
ws_ui.column_dimensions['C'].width = 40

ws_ui['B2'] = "TRA CỨU SUẤT VỐN ĐẦU TƯ"
ws_ui['B2'].font = Font(size=14, bold=True, color="0000FF")

# Inputs (No Data Validation yet to avoid corruption)
ws_ui['B4'] = "1. Chọn Mã hiệu:"
ws_ui['B4'].font = Font(bold=True)
ws_ui['C4'].border = Border(bottom=Side(style='thin'))
ws_ui['C4'].value = refined_data[0][0] # Default value

ws_ui['B5'] = "2. Chọn Loại công trình:"
ws_ui['B5'].font = Font(bold=True)
ws_ui['C5'].border = Border(bottom=Side(style='thin'))
ws_ui['C5'].alignment = Alignment(wrap_text=True)
ws_ui['C5'].value = refined_data[0][1] # Default value

# Results Formulas (Simple SUMIFS)
# Note: Using comma separator for standard Excel file generation. 
# Excel will automatically convert to semicolon if opened in VN region.
res_labels = ["Đơn vị tính", "Suất vốn đầu tư", "Chi phí Xây dựng", "Chi phí Thiết bị"]
col_map = ["C", "D", "E", "F"] # Columns in DATA

ws_ui['B8'] = "KẾT QUẢ:"
ws_ui['B8'].font = Font(bold=True, underline="single")

for i, lbl in enumerate(res_labels):
    r = 9 + i
    ws_ui[f'B{r}'] = lbl
    ws_ui[f'B{r}'].font = Font(bold=True)
    
    col_letter = col_map[i]
    
    if i == 0: 
        f = "1000đ/m2"
    else:
        # SUMIFS(ValCol, CodeCol, C4, DescCol, C5)
        # Using comma, let Excel handle locale.
        f = f"=SUMIFS(DATA!{col_letter}:{col_letter}, DATA!A:A, C4, DATA!B:B, C5)"
        ws_ui[f'C{r}'].number_format = '#,##0'
        
    ws_ui[f'C{r}'] = f

print(f"Saving to {output_excel}...")
wb.save(output_excel)
print("Done.")
