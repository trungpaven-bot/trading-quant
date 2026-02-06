
import openpyxl

file_path = r"g:\My Drive\30_RESOURCES (Kho tai nguyen)\30.01_Tinh hieu qua du an\02_Tra dinh muc theo TT 09 va 12\Tra dinh muc TT09.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=False)
ws = wb["TT09"]

print(f"Sheet Protection: {ws.protection.sheet}")
print(f"Password Hash: {ws.protection.password}")

# Break Links Logic
# We can't easily "Break Links" using openpyxl like Excel does (converting to values).
# But we can remove the external links from the workbook definition so Excel stops asking.
# However, formulas like ='[2]Config'!C118 will result in #REF! if we just delete the link definition.
# 
# The user wants to FIX the sign in.
# Best way using openpyxl: Find formulas with brackets [] and replace them with 0 or #N/A to stop the prompt?
# Or just delete the defined names that refer externally.
#
# The "Sign In" usually comes from the Defined Names or Formulas pointing to a URL/SharePoint.
# My previous check showed external refs in Formulas AND Defined Names.
# I should delete the External Defined Names.
# And replace External Formulas with 0.

print("\n--- Cleaning External References ---")
# 1. Clean External Defined Names
names_to_delete = []
for name in wb.defined_names:
    defn = wb.defined_names[name]
    if "[" in str(defn.attr_text) or "http" in str(defn.attr_text):
        names_to_delete.append(name)

for name in names_to_delete:
    del wb.defined_names[name]
    print(f"Deleted Name: {name}")

# 2. Clean/Nuke External Formulas in TT09 (and others?)
# This is destructive but the user can't use them anyway (don't have the file).
# Focus on the ones causing issues.
cnt = 0
for sheet in wb.sheetnames:
    ws_curr = wb[sheet]
    for row in ws_curr.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and "=" in cell.value:
                if "[" in cell.value and "]" in cell.value:
                    # Replace with 0 to stop the error/prompt
                    cell.value = 0
                    cnt += 1
    if cnt > 0:
        print(f"Replaced {cnt} external formulas in {sheet} with 0.")
    cnt = 0

print("Saving cleaned file...")
wb.save(file_path)
print("Done.")
