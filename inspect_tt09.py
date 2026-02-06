
import pandas as pd
import openpyxl

file_path = r"g:\My Drive\00_INBOX (xu ly nhanh)\tra TMDT.xlsx"
sheet_name = "TT09"

print("--- ANALYZING SHEET TT09 ---")

# 1. Find the header row using pandas
df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
print(f"Shape: {df.shape}")

# Find first row with meaningful content (e.g. more than 3 non-null columns)
header_row_idx = 0
for i, row in df.iterrows():
    if row.count() > 3:
        header_row_idx = i
        print(f"Likely header at row {i+1}:")
        print(row.dropna().tolist())
        break

# Reload with header
df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row_idx)
print("\n--- Columns ---")
print(df.columns.tolist())

print("\n--- Sample Data (Rows 1-5 relative to data) ---")
print(df.head(5).to_string())

# 2. Check formulas using openpyxl
print("\n--- Formulas Check ---")
wb = openpyxl.load_workbook(file_path, data_only=False)
ws = wb[sheet_name]

# Check the row after the header
data_start_row = header_row_idx + 2 # 1-based, +1 for header
print(f"Checking formulas at row {data_start_row}")

for r in range(data_start_row, data_start_row + 5):
    row_formulas = []
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=r, column=c)
        val = cell.value
        if isinstance(val, str) and val.startswith("="):
            # Get column letter
            col_letter = openpyxl.utils.get_column_letter(c)
            row_formulas.append(f"{col_letter}{r}: {val}")
    if row_formulas:
        print(f"Row {r} Formulas: {row_formulas}")

print("\n--- Merged Cells ---")
print(list(ws.merged_cells.ranges)[:10])
